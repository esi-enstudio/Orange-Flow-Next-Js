import base64
import logging
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request, Response
from pydantic import BaseModel, Field, field_validator
from sqlalchemy import select, func, or_, and_, cast, String
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload, selectinload
from datetime import datetime, date, timedelta
from calendar import monthrange

from app.routers.deps import get_db, has_permission, get_house_context, get_current_user
from app.schemas.pagination import PaginationParams, PaginatedResponse, PaginationMeta
from app.models.user import User
from app.models.activation import Activation
from app.models.live_activation import LiveActivation
from app.models.house import House
from app.models.itopup_detail import ITopUpDetail
from app.models.scratch_card_issue import ScratchCardIssue
from app.models.sim_issue import SimIssue
from app.models.ga_filter import RetailerFilter, FilterTag, RetailerFilter as RetailerFilterModel
from app.models.retailer import Retailer
from app.models.employee import Employee
from app.models.active_lso_config import ActiveLsoConfig
from app.models.active_sso_config import ActiveSsoConfig
from app.models.bp_retailer_code import BpRetailerCode
from app.models.role import Role
from app.utils.access_control import is_admin_user
from app.utils.activity_logger import log_activity
from app.services.whatsapp_service_client import whatsapp_service_client, WhatsAppServiceError
from app.services.whatsapp_token import resolve_house_wa_target
from app.utils.activation_rules import get_excluded_codes, exclude_clause
from app.services.Automation.activation_excel import export_activations_excel
from app.services.Automation.dms_report_excel import export_itopup_details_excel
from app.services.Automation.live_activation_excel import export_live_activations_excel
from app.services.Automation.ga_live_performance_excel import export_ga_live_performance_excel
from app.services.Automation.issue_reports_excel import export_scratch_card_excel, export_sim_issue_excel
from app.services.target_achievement_service import TargetAchievementService
from app.services.active_lso_report_service import (
    ActiveLsoReportService,
    DEFAULT_ACTIVE_LSO_DAYS,
    get_active_lso_filters as _get_active_lso_filters,
    get_active_lso_thresholds,
    get_employee_profile,
    subordinate_rso_ids,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api", tags=["reports"])

@router.get("/activations")
async def get_activations(
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,

    activation_date_from: Optional[str] = None,
    activation_date_to: Optional[str] = None,
    activation_time: Optional[str] = None,
    retailer_code: Optional[str] = None,
    retailer_name: Optional[str] = None,
    bts_code: Optional[str] = None,
    thana: Optional[str] = None,
    promotion: Optional[str] = None,
    product_code: Optional[str] = None,
    product_codes: Optional[str] = None,
    product_name: Optional[str] = None,
    sim_no: Optional[str] = None,
    msisdn: Optional[str] = None,
    selling_price_min: Optional[str] = None,
    selling_price_max: Optional[str] = None,
    bp_flag: Optional[str] = None,
    bp_number: Optional[str] = None,
    fc_bts_code: Optional[str] = None,
    bio_bts_code: Optional[str] = None,
    dh_lifting_date: Optional[str] = None,
    issue_date: Optional[str] = None,
    subscription_type: Optional[str] = None,
    service_class: Optional[str] = None,
    customer_second_contact: Optional[str] = None,
    employee_id: Optional[int] = None,
    filter_house_id: Optional[int] = Query(None, alias="house_id"),

    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("activations.view")),
    header_house_id: Optional[int] = Depends(get_house_context)
):
    effective_house_id = filter_house_id or header_house_id
    query = select(Activation).options(
        joinedload(Activation.house),
        joinedload(Activation.retailer).joinedload(Retailer.employee).joinedload(Employee.user)
    )
    if effective_house_id: query = query.where(Activation.house_id == effective_house_id)

    if employee_id:
        retailer_ids_subq = select(Retailer.id).where(Retailer.employee_id == employee_id)
        query = query.where(Activation.retailer_id.in_(retailer_ids_subq))

    if activation_date_from:
        try: sd = datetime.strptime(activation_date_from, "%Y-%m-%d").date()
        except: pass
        else: query = query.where(Activation.activation_date >= sd)
    if activation_date_to:
        try: ed = datetime.strptime(activation_date_to, "%Y-%m-%d").date()
        except: pass
        else: query = query.where(Activation.activation_date <= ed)
    if activation_time:
        p = f"%{activation_time}%"
        query = query.where(Activation.activation_time.ilike(p))
    if retailer_code:
        p = f"%{retailer_code}%"
        query = query.where(Activation.retailer_code.ilike(p))
    if retailer_name:
        p = f"%{retailer_name}%"
        query = query.where(Activation.retailer_name.ilike(p))
    if bts_code:
        p = f"%{bts_code}%"
        query = query.where(Activation.bts_code.ilike(p))
    if thana:
        p = f"%{thana}%"
        query = query.where(Activation.thana.ilike(p))
    if promotion:
        query = query.where(Activation.promotion == promotion)
    if product_code:
        query = query.where(Activation.product_code == product_code)
    if product_codes:
        codes_list = [c.strip() for c in product_codes.split(",") if c.strip()]
        if codes_list:
            query = query.where(Activation.product_code.in_(codes_list))
    if product_name:
        query = query.where(Activation.product_name == product_name)
    if sim_no:
        p = f"%{sim_no}%"
        query = query.where(Activation.sim_no.ilike(p))
    if msisdn:
        p = f"%{msisdn}%"
        query = query.where(Activation.msisdn.ilike(p))
    if selling_price_min:
        query = query.where(Activation.selling_price >= selling_price_min)
    if selling_price_max:
        query = query.where(Activation.selling_price <= selling_price_max)
    if bp_flag:
        query = query.where(Activation.bp_flag == bp_flag)
    if bp_number:
        p = f"%{bp_number}%"
        query = query.where(Activation.bp_number.ilike(p))
    if fc_bts_code:
        p = f"%{fc_bts_code}%"
        query = query.where(Activation.fc_bts_code.ilike(p))
    if bio_bts_code:
        p = f"%{bio_bts_code}%"
        query = query.where(Activation.bio_bts_code.ilike(p))
    if dh_lifting_date:
        p = f"%{dh_lifting_date}%"
        query = query.where(Activation.dh_lifting_date.ilike(p))
    if issue_date:
        p = f"%{issue_date}%"
        query = query.where(Activation.issue_date.ilike(p))
    if subscription_type:
        query = query.where(Activation.subscription_type == subscription_type)
    if service_class:
        query = query.where(Activation.service_class == service_class)
    if customer_second_contact:
        p = f"%{customer_second_contact}%"
        query = query.where(Activation.customer_second_contact.ilike(p))

    if search:
        p = f"%{search}%"
        query = query.where(
            (Activation.sim_no.ilike(p)) | (Activation.retailer_code.ilike(p)) |
            (Activation.retailer_name.ilike(p)) | (Activation.msisdn.ilike(p))
        )
    count_query = select(func.count()).select_from(query.subquery())
    total = await db.execute(count_query)
    total_count = total.scalar()
    result = await db.execute(query.offset(skip).limit(limit).order_by(Activation.id.desc()))
    records = result.unique().scalars().all()

    data = []
    for r in records:
        item = {
            "id": r.id, "sim_no": r.sim_no, "activation_date": r.activation_date,
            "activation_time": r.activation_time, "retailer_code": r.retailer_code,
            "retailer_name": r.retailer_name, "bts_code": r.bts_code, "thana": r.thana,
            "promotion": r.promotion, "product_code": r.product_code, "product_name": r.product_name,
            "msisdn": r.msisdn, "selling_price": r.selling_price,
            "bp_flag": r.bp_flag, "bp_number": r.bp_number,
            "fc_bts_code": r.fc_bts_code, "bio_bts_code": r.bio_bts_code,
            "dh_lifting_date": r.dh_lifting_date, "issue_date": r.issue_date,
            "subscription_type": r.subscription_type, "service_class": r.service_class,
            "customer_second_contact": r.customer_second_contact,
            "house_id": r.house_id, "house": None, "rso_name": None, "rso_employee_id": None,
            "rso_dms_code": None, "rso_itop_number": None,
        }
        if r.house:
            item["house"] = {"id": r.house.id, "name": r.house.name, "code": r.house.code}
        if r.retailer and r.retailer.employee:
            emp = r.retailer.employee
            item["rso_name"] = emp.user.name if emp.user else emp.dms_code
            item["rso_employee_id"] = emp.id
            item["rso_dms_code"] = emp.dms_code
            item["rso_itop_number"] = emp.itop_number
        data.append(item)

    return {"total": total_count, "data": data}

@router.get("/activations/filter-options")
async def get_activation_filter_options(
    filter_house_id: Optional[int] = Query(None, alias="house_id"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("activations.view")),
    header_house_id: Optional[int] = Depends(get_house_context)
):
    effective_house_id = filter_house_id or header_house_id
    base = select(Activation)
    if effective_house_id: base = base.where(Activation.house_id == effective_house_id)

    async def get_distinct(column):
        q = select(column).distinct().where(column.isnot(None)).where(column != "").order_by(column)
        result = await db.execute(q)
        return [row[0] for row in result.all()]

    promotions = await get_distinct(Activation.promotion)
    product_codes = await get_distinct(Activation.product_code)
    product_names = await get_distinct(Activation.product_name)
    subscription_types = await get_distinct(Activation.subscription_type)
    service_classes = await get_distinct(Activation.service_class)
    bp_flags = await get_distinct(Activation.bp_flag)

    return {
        "promotions": promotions,
        "product_codes": product_codes,
        "product_names": product_names,
        "subscription_types": subscription_types,
        "service_classes": service_classes,
        "bp_flags": bp_flags,
    }

@router.get("/activations/rso-list")
async def get_activation_rso_list(
    filter_house_id: Optional[int] = Query(None, alias="house_id"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("activations.view")),
    header_house_id: Optional[int] = Depends(get_house_context)
):
    effective_house_id = filter_house_id or header_house_id
    base = select(Employee).options(joinedload(Employee.user)).where(Employee.employee_type == "rso")
    if effective_house_id:
        base = base.where(Employee.house_id == effective_house_id)
    result = await db.execute(base.order_by(Employee.id))
    employees = result.scalars().unique().all()
    return [
        {
            "id": e.id,
            "name": e.user.name if e.user else e.dms_code,
            "employee_id": e.employee_id,
            "dms_code": e.dms_code,
        }
        for e in employees
    ]

@router.get("/activations/export")
async def export_activations(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("activations.export")),
    house_id: Optional[int] = Depends(get_house_context)
):
    query = select(Activation).options(joinedload(Activation.house))
    if house_id: query = query.where(Activation.house_id == house_id)
    if start_date:
        try: sd = datetime.strptime(start_date, "%Y-%m-%d").date()
        except: return Response("Invalid start_date format", status_code=400)
        query = query.where(Activation.activation_date >= sd)
    if end_date:
        try: ed = datetime.strptime(end_date, "%Y-%m-%d").date()
        except: return Response("Invalid end_date format", status_code=400)
        query = query.where(Activation.activation_date <= ed)
    result = await db.execute(query.order_by(Activation.id.desc()))
    records = result.scalars().all()
    excel_data = await export_activations_excel(records)
    return Response(
        content=excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=activations.xlsx"}
    )

@router.get("/itopup-details")
async def get_itopup_details(
    pagination: PaginationParams = Depends(),
    report_type: Optional[str] = Query(None, description="Filter by report type (C2C, C2S, Balance)"),
    start_date: Optional[str] = Query(None, description="Filter by report_date >= start_date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="Filter by report_date <= end_date (YYYY-MM-DD)"),
    retailer_search: Optional[str] = Query(None, description="Search by retailer code or name"),
    filter_house_id: Optional[int] = Query(None, alias="house_id", description="Filter by house ID"),
    rso_id: Optional[int] = Query(None, description="Filter by RSO employee ID"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("itopup.view")),
    house_id: Optional[int] = Depends(get_house_context)
):
    query = select(ITopUpDetail).options(
        joinedload(ITopUpDetail.house),
        joinedload(ITopUpDetail.retailer).selectinload(Retailer.employee).selectinload(Employee.user)
    )

    conditions = []

    effective_house_id = filter_house_id or house_id
    if effective_house_id:
        conditions.append(ITopUpDetail.house_id == effective_house_id)

    if rso_id:
        conditions.append(
            ITopUpDetail.retailer.has(Retailer.employee_id == rso_id)
        )

    if report_type:
        conditions.append(ITopUpDetail.report_type == report_type)

    if start_date:
        try:
            sd = datetime.strptime(start_date, "%Y-%m-%d").date()
            conditions.append(ITopUpDetail.report_date >= sd)
        except:
            raise HTTPException(status_code=400, detail="Invalid start_date format (use YYYY-MM-DD)")

    if end_date:
        try:
            ed = datetime.strptime(end_date, "%Y-%m-%d").date()
            conditions.append(ITopUpDetail.report_date <= ed)
        except:
            raise HTTPException(status_code=400, detail="Invalid end_date format (use YYYY-MM-DD)")

    if pagination.search:
        p = f"%{pagination.search}%"
        conditions.append(
            or_(
                ITopUpDetail.report_type.ilike(p),
                cast(ITopUpDetail.report_date, String).ilike(p),
                cast(ITopUpDetail.daily_value, String).ilike(p),
                ITopUpDetail.house.has(
                    or_(House.name.ilike(p), House.code.ilike(p))
                ),
                ITopUpDetail.retailer.has(
                    or_(
                        Retailer.name.ilike(p),
                        Retailer.retailer_code.ilike(p),
                        Retailer.itop_number.ilike(p),
                        Retailer.employee.has(
                            or_(
                                Employee.dms_code.ilike(p),
                                Employee.itop_number.ilike(p),
                                Employee.user.has(User.name.ilike(p)),
                            )
                        ),
                    )
                ),
            )
        )

    if retailer_search:
        p = f"%{retailer_search}%"
        conditions.append(
            ITopUpDetail.retailer.has(
                or_(Retailer.retailer_code.ilike(p), Retailer.name.ilike(p))
            )
        )

    if conditions:
        query = query.where(and_(*conditions))

    base_count = select(ITopUpDetail.id)
    if conditions:
        base_count = base_count.where(and_(*conditions))
    if effective_house_id:
        base_count = base_count.where(ITopUpDetail.house_id == effective_house_id)
    count_query = select(func.count()).select_from(base_count.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    sort_map = {
        "id": ITopUpDetail.id,
        "report_type": ITopUpDetail.report_type,
        "report_date": ITopUpDetail.report_date,
        "daily_value": ITopUpDetail.daily_value,
    }
    sort_col = sort_map.get(pagination.sort_by, ITopUpDetail.id)
    order = sort_col.desc() if pagination.sort_order == "desc" else sort_col.asc()

    offset = (pagination.page - 1) * pagination.per_page
    result = await db.execute(query.order_by(order).offset(offset).limit(pagination.per_page))
    records = result.scalars().all()

    total_pages = max(1, (total + pagination.per_page - 1) // pagination.per_page)

    return {
        "success": True,
        "data": records,
        "pagination": {
            "page": pagination.page,
            "per_page": pagination.per_page,
            "total": total,
            "total_pages": total_pages,
            "has_next": pagination.page < total_pages,
            "has_prev": pagination.page > 1,
        }
    }

@router.get("/itopup-details/rso-list")
async def get_itopup_rso_list(
    filter_house_id: Optional[int] = Query(None, alias="house_id"),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("itopup.view")),
    header_house_id: Optional[int] = Depends(get_house_context)
):
    effective_house_id = filter_house_id or header_house_id
    base = select(Employee).options(joinedload(Employee.user)).where(
        Employee.employee_type == "rso",
        Employee.status == "Active"
    )
    if effective_house_id:
        base = base.where(Employee.house_id == effective_house_id)
    if search:
        p = f"%{search}%"
        base = base.where(
            or_(
                Employee.dms_code.ilike(p),
                Employee.itop_number.ilike(p),
                Employee.user.has(User.name.ilike(p)),
            )
        )
    result = await db.execute(base.order_by(Employee.id))
    employees = result.scalars().unique().all()
    return [
        {
            "id": e.id,
            "name": e.user.name if e.user else e.dms_code,
            "dms_code": e.dms_code,
            "itop_number": e.itop_number,
        }
        for e in employees
    ]

@router.get("/itopup-details/export")
async def export_itopup_details(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    report_type: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("itopup.export")),
    house_id: Optional[int] = Depends(get_house_context)
):
    query = select(ITopUpDetail)
    if house_id: query = query.where(ITopUpDetail.house_id == house_id)
    if report_type: query = query.where(ITopUpDetail.report_type == report_type)
    if start_date:
        try: sd = datetime.strptime(start_date, "%Y-%m-%d").date()
        except: return Response("Invalid start_date", status_code=400)
        query = query.where(ITopUpDetail.report_date >= sd)
    if end_date:
        try: ed = datetime.strptime(end_date, "%Y-%m-%d").date()
        except: return Response("Invalid end_date", status_code=400)
        query = query.where(ITopUpDetail.report_date <= ed)
    result = await db.execute(query.order_by(ITopUpDetail.id.desc()).limit(50000))
    records = result.scalars().all()
    excel_data = await export_itopup_details_excel(records)
    return Response(
        content=excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=itopup_details.xlsx"}
    )

@router.get("/live-activations")
async def get_live_activations(
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    activation_date_from: Optional[str] = None,
    activation_date_to: Optional[str] = None,
    activation_time: Optional[str] = None,
    retailer_code: Optional[str] = None,
    retailer_name: Optional[str] = None,
    bts_code: Optional[str] = None,
    thana: Optional[str] = None,
    promotion: Optional[str] = None,
    product_code: Optional[str] = None,
    product_name: Optional[str] = None,
    sim_no: Optional[str] = None,
    msisdn: Optional[str] = None,
    selling_price_min: Optional[str] = None,
    selling_price_max: Optional[str] = None,
    bp_flag: Optional[str] = None,
    bp_number: Optional[str] = None,
    fc_bts_code: Optional[str] = None,
    bio_bts_code: Optional[str] = None,
    dh_lifting_date: Optional[str] = None,
    issue_date: Optional[str] = None,
    subscription_type: Optional[str] = None,
    service_class: Optional[str] = None,
    customer_second_contact: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("live_activations.view")),
    house_id: Optional[int] = Depends(get_house_context)
):
    query = select(LiveActivation).options(joinedload(LiveActivation.house))
    if house_id: query = query.where(LiveActivation.house_id == house_id)
    if search:
        p = f"%{search}%"
        query = query.where(
            (LiveActivation.sim_no.ilike(p)) | (LiveActivation.retailer_code.ilike(p)) |
            (LiveActivation.retailer_name.ilike(p)) | (LiveActivation.msisdn.ilike(p))
        )
    if activation_date_from:
        try:
            sd = datetime.strptime(activation_date_from, "%Y-%m-%d").date()
            query = query.where(LiveActivation.activation_date >= sd)
        except:
            pass
    if activation_date_to:
        try:
            ed = datetime.strptime(activation_date_to, "%Y-%m-%d").date()
            query = query.where(LiveActivation.activation_date <= ed)
        except:
            pass
    if activation_time: query = query.where(LiveActivation.activation_time.ilike(f"%{activation_time}%"))
    if retailer_code: query = query.where(LiveActivation.retailer_code.ilike(f"%{retailer_code}%"))
    if retailer_name: query = query.where(LiveActivation.retailer_name.ilike(f"%{retailer_name}%"))
    if bts_code: query = query.where(LiveActivation.bts_code.ilike(f"%{bts_code}%"))
    if thana: query = query.where(LiveActivation.thana.ilike(f"%{thana}%"))
    if promotion: query = query.where(LiveActivation.promotion == promotion)
    if product_code: query = query.where(LiveActivation.product_code.ilike(f"%{product_code}%"))
    if product_name: query = query.where(LiveActivation.product_name.ilike(f"%{product_name}%"))
    if sim_no: query = query.where(LiveActivation.sim_no.ilike(f"%{sim_no}%"))
    if msisdn: query = query.where(LiveActivation.msisdn.ilike(f"%{msisdn}%"))
    if selling_price_min: query = query.where(LiveActivation.selling_price >= selling_price_min)
    if selling_price_max: query = query.where(LiveActivation.selling_price <= selling_price_max)
    if bp_flag: query = query.where(LiveActivation.bp_flag == bp_flag)
    if bp_number: query = query.where(LiveActivation.bp_number.ilike(f"%{bp_number}%"))
    if fc_bts_code: query = query.where(LiveActivation.fc_bts_code.ilike(f"%{fc_bts_code}%"))
    if bio_bts_code: query = query.where(LiveActivation.bio_bts_code.ilike(f"%{bio_bts_code}%"))
    if dh_lifting_date: query = query.where(LiveActivation.dh_lifting_date.ilike(f"%{dh_lifting_date}%"))
    if issue_date: query = query.where(LiveActivation.issue_date.ilike(f"%{issue_date}%"))
    if subscription_type: query = query.where(LiveActivation.subscription_type == subscription_type)
    if service_class: query = query.where(LiveActivation.service_class == service_class)
    if customer_second_contact: query = query.where(LiveActivation.customer_second_contact.ilike(f"%{customer_second_contact}%"))
    count_query = select(func.count()).select_from(query.subquery())
    total = await db.execute(count_query)
    total_count = total.scalar()
    result = await db.execute(query.offset(skip).limit(limit).order_by(LiveActivation.id.desc()))
    records = result.scalars().all()
    return {"total": total_count, "data": records}

@router.get("/live-activations/filter-options")
async def get_live_activation_filter_options(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("live_activations.view")),
    house_id: Optional[int] = Depends(get_house_context)
):
    base = select(LiveActivation)
    if house_id: base = base.where(LiveActivation.house_id == house_id)

    async def get_distinct(column):
        q = select(column).distinct().where(column.isnot(None)).where(column != "").order_by(column)
        result = await db.execute(q)
        return [row[0] for row in result.all()]

    promotions = await get_distinct(LiveActivation.promotion)
    product_codes = await get_distinct(LiveActivation.product_code)
    product_names = await get_distinct(LiveActivation.product_name)
    subscription_types = await get_distinct(LiveActivation.subscription_type)
    service_classes = await get_distinct(LiveActivation.service_class)
    bp_flags = await get_distinct(LiveActivation.bp_flag)

    return {
        "promotions": promotions,
        "product_codes": product_codes,
        "product_names": product_names,
        "subscription_types": subscription_types,
        "service_classes": service_classes,
        "bp_flags": bp_flags,
    }

@router.get("/live-activations/export")
async def export_live_activations(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    house_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("live_activations.export")),
    header_house_id: Optional[int] = Depends(get_house_context),
):
    if not house_id:
        house_id = header_house_id
    query = select(LiveActivation).options(selectinload(LiveActivation.house), selectinload(LiveActivation.retailer))
    if house_id: query = query.where(LiveActivation.house_id == house_id)
    if start_date:
        try: sd = datetime.strptime(start_date, "%Y-%m-%d").date()
        except: return Response("Invalid start_date", status_code=400)
        query = query.where(LiveActivation.activation_date >= sd)
    if end_date:
        try: ed = datetime.strptime(end_date, "%Y-%m-%d").date()
        except: return Response("Invalid end_date", status_code=400)
        query = query.where(LiveActivation.activation_date <= ed)
    result = await db.execute(query.order_by(LiveActivation.id.desc()).limit(50000))
    records = result.scalars().all()
    excel_data = await export_live_activations_excel(records)
    return Response(
        content=excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=live_activations.xlsx"}
    )


@router.get("/reports/live-activations/export-performance")
async def export_ga_live_performance(
    house_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("live_activations.export")),
):
    is_admin = is_admin_user(current_user)
    user_houses_raw = current_user.houses

    target_house_id = house_id
    if not target_house_id:
        if is_admin:
            raise HTTPException(status_code=400, detail="house_id is required for admin users")
        if len(user_houses_raw) == 1:
            target_house_id = user_houses_raw[0].id
        else:
            raise HTTPException(status_code=400, detail="house_id is required when user has multiple houses")
    else:
        if not is_admin:
            user_house_ids = [h.id for h in user_houses_raw]
            if target_house_id not in user_house_ids:
                raise HTTPException(status_code=403, detail="You do not have access to this house")

    today = date.today()
    excel_data = await export_ga_live_performance_excel(db, target_house_id, today)
    filename = f"ga_live_performance_{today}.xlsx"
    return Response(
        content=excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.delete("/live-activations/truncate")
async def truncate_live_activations(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("live_activations.import")),
):
    await db.execute(LiveActivation.__table__.delete())
    await db.commit()
    return {"message": "All live activations deleted successfully"}

@router.delete("/activations/truncate")
async def truncate_activations(
    house_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("activations.import")),
):
    if house_id:
        await db.execute(Activation.__table__.delete().where(Activation.house_id == house_id))
        await db.commit()
        return {"message": f"Activations deleted for house {house_id}"}
    await db.execute(Activation.__table__.delete())
    await db.commit()
    return {"message": "All activations deleted successfully"}

@router.delete("/itopup-details/truncate")
async def truncate_itopup_details(
    house_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("itopup.import")),
):
    if house_id:
        await db.execute(ITopUpDetail.__table__.delete().where(ITopUpDetail.house_id == house_id))
        await db.commit()
        return {"message": f"iTopUp details deleted for house {house_id}"}
    await db.execute(ITopUpDetail.__table__.delete())
    await db.commit()
    return {"message": "All iTopUp details deleted successfully"}

@router.get("/scratch-card")
async def get_scratch_card(
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("scratch_card.view")),
    house_id: Optional[int] = Depends(get_house_context)
):
    query = select(ScratchCardIssue)
    if house_id: query = query.where(ScratchCardIssue.house_id == house_id)
    if search:
        p = f"%{search}%"
        query = query.where(
            (ScratchCardIssue.distributor_code.ilike(p)) | (ScratchCardIssue.retailer_code.ilike(p)) |
            (ScratchCardIssue.retailer_name.ilike(p)) | (ScratchCardIssue.product_name.ilike(p))
        )
    count_query = select(func.count()).select_from(query.subquery())
    total = await db.execute(count_query)
    total_count = total.scalar()
    result = await db.execute(query.offset(skip).limit(limit).order_by(ScratchCardIssue.id.desc()))
    records = result.scalars().all()
    return {"total": total_count, "data": records}

@router.get("/scratch-card/export")
async def export_scratch_card(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("scratch_card.export")),
    house_id: Optional[int] = Depends(get_house_context)
):
    query = select(ScratchCardIssue)
    if house_id: query = query.where(ScratchCardIssue.house_id == house_id)
    if start_date:
        try: sd = datetime.strptime(start_date, "%Y-%m-%d").date()
        except: return Response("Invalid start_date", status_code=400)
        query = query.where(ScratchCardIssue.issue_date >= sd)
    if end_date:
        try: ed = datetime.strptime(end_date, "%Y-%m-%d").date()
        except: return Response("Invalid end_date", status_code=400)
        query = query.where(ScratchCardIssue.issue_date <= ed)
    result = await db.execute(query.order_by(ScratchCardIssue.id.desc()).limit(50000))
    records = result.scalars().all()
    excel_data = await export_scratch_card_excel(records)
    return Response(
        content=excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=scratch_card.xlsx"}
    )

@router.get("/sim-issues")
async def get_sim_issues(
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("sim_issues.view")),
    house_id: Optional[int] = Depends(get_house_context)
):
    query = select(SimIssue)
    if house_id: query = query.where(SimIssue.house_id == house_id)
    if search:
        p = f"%{search}%"
        query = query.where(
            (SimIssue.sim_no.ilike(p)) | (SimIssue.distributor_code.ilike(p)) |
            (SimIssue.retailer_code.ilike(p)) | (SimIssue.retailer_name.ilike(p)) |
            (SimIssue.product_name.ilike(p))
        )
    count_query = select(func.count()).select_from(query.subquery())
    total = await db.execute(count_query)
    total_count = total.scalar()
    result = await db.execute(query.offset(skip).limit(limit).order_by(SimIssue.id.desc()))
    records = result.scalars().all()
    return {"total": total_count, "data": records}

@router.get("/sim-issues/export")
async def export_sim_issues(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("sim_issues.export")),
    house_id: Optional[int] = Depends(get_house_context)
):
    query = select(SimIssue)
    if house_id: query = query.where(SimIssue.house_id == house_id)
    if start_date:
        try: sd = datetime.strptime(start_date, "%Y-%m-%d").date()
        except: return Response("Invalid start_date", status_code=400)
        query = query.where(SimIssue.issue_date >= sd)
    if end_date:
        try: ed = datetime.strptime(end_date, "%Y-%m-%d").date()
        except: return Response("Invalid end_date", status_code=400)
        query = query.where(SimIssue.issue_date <= ed)
    result = await db.execute(query.order_by(SimIssue.id.desc()).limit(50000))
    records = result.scalars().all()
    excel_data = await export_sim_issue_excel(records)
    return Response(
        content=excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sim_issues.xlsx"}
    )

@router.get("/activations/report")
async def get_activation_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    exclude_tags: Optional[str] = Query(None, description="Comma-separated tag names to exclude (e.g. DRC,RSP,BSP)"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("reports.view")),
    house_id: Optional[int] = Depends(get_house_context),
    q_house_id: Optional[int] = Query(None, alias="house_id"),
):
    from app.services.cache_service import cache_service
    target_house_id = q_house_id or house_id
    if not search and page == 1 and page_size <= 50:
        cache_key = cache_service.cache_key("report", start_date, end_date, exclude_tags, target_house_id)
        cached = await cache_service.get(cache_key)
        if cached:
            return cached
    if q_house_id and not is_admin_user(current_user):
        user_house_ids = [h.id for h in current_user.houses]
        if q_house_id not in user_house_ids:
            raise HTTPException(status_code=403, detail="You do not have access to this house")
    excluded_codes = await get_excluded_codes(db)
    query = select(Activation)
    is_admin = is_admin_user(current_user)
    effective_house_id = target_house_id
    user_house_ids = []
    if not effective_house_id and not is_admin:
        user_house_ids = [h.id for h in current_user.houses]
        if user_house_ids:
            query = query.where(Activation.house_id.in_(user_house_ids))
    elif effective_house_id:
        query = query.where(Activation.house_id == effective_house_id)
    clause = exclude_clause(Activation, excluded_codes)
    if clause is not None:
        query = query.where(clause)
    today_dt = date.today()
    if not start_date:
        start_date = today_dt.replace(day=1).isoformat()
    if not end_date:
        last = monthrange(today_dt.year, today_dt.month)[1]
        end_date = today_dt.replace(day=last).isoformat()
    if start_date:
        try: sd = datetime.strptime(start_date, "%Y-%m-%d").date()
        except: raise HTTPException(status_code=400, detail="Invalid start_date format, use YYYY-MM-DD")
        query = query.where(Activation.activation_date >= sd)
    if end_date:
        try: ed = datetime.strptime(end_date, "%Y-%m-%d").date()
        except: raise HTTPException(status_code=400, detail="Invalid end_date format, use YYYY-MM-DD")
        query = query.where(Activation.activation_date <= ed)
    if search:
        p = f"%{search}%"
        query = query.where((Activation.sim_no.ilike(p)) | (Activation.retailer_code.ilike(p)) | (Activation.retailer_name.ilike(p)) | (Activation.msisdn.ilike(p)))
    excluded_tags_list = []
    excluded_retailer_ids = []
    if exclude_tags:
        excluded_tags_list = [t.strip() for t in exclude_tags.split(",") if t.strip()]
        if excluded_tags_list:
            house_ids_for_exclusion = [effective_house_id] if effective_house_id else (user_house_ids if not is_admin else None)
            excl_query = select(RetailerFilter.retailer_id).join(FilterTag, RetailerFilter.tag_id == FilterTag.id).where(FilterTag.name.in_(excluded_tags_list))
            if house_ids_for_exclusion:
                excl_query = excl_query.where(RetailerFilter.house_id.in_(house_ids_for_exclusion))
            excluded_ids_result = await db.execute(excl_query)
            excluded_retailer_ids = [row[0] for row in excluded_ids_result.all()]
            if excluded_retailer_ids:
                query = query.where(Activation.retailer_id.notin_(excluded_retailer_ids))
    # Single count query with all filters applied
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await db.execute(count_query)
    total_count = total_result.scalar()
    excluded_count = len(excluded_retailer_ids)  # approximate, not exact
    offset = (page - 1) * page_size
    result = await db.execute(query.offset(offset).limit(page_size).order_by(Activation.id.desc()))
    records = result.scalars().all()

    # Bulk fetch retailer tags
    retailer_ids = [r.retailer_id for r in records if r.retailer_id]
    tags_map: dict[int, list[str]] = {}
    if retailer_ids:
        tag_rows = await db.execute(
            select(RetailerFilter.retailer_id, FilterTag.name)
            .join(FilterTag, RetailerFilter.tag_id == FilterTag.id)
            .where(RetailerFilter.retailer_id.in_(retailer_ids))
        )
        for rid, tname in tag_rows.all():
            tags_map.setdefault(rid, []).append(tname)

    # Bulk fetch RSO info
    rso_map: dict[int, dict] = {}
    if retailer_ids:
        rso_rows = await db.execute(
            select(Retailer.id, Employee.user_id, User.name, Employee.itop_number)
            .join(Employee, Retailer.employee_id == Employee.id)
            .join(User, Employee.user_id == User.id)
            .where(Retailer.id.in_(retailer_ids))
        )
        for rid, uid, uname, itop in rso_rows.all():
            rso_map[rid] = {"name": uname, "itop": itop}

    data = []
    for r in records:
        item = {
            "id": r.id, "house_id": r.house_id, "retailer_id": r.retailer_id,
            "activation_date": r.activation_date.isoformat() if r.activation_date else None,
            "retailer_code": r.retailer_code, "retailer_name": r.retailer_name,
            "retailer_tags": tags_map.get(r.retailer_id, []) if r.retailer_id else [],
            "rso": rso_map.get(r.retailer_id),
            "sim_no": r.sim_no, "msisdn": r.msisdn, "product_name": r.product_name,
            "product_code": r.product_code, "selling_price": r.selling_price, "thana": r.thana,
            "house": {"id": r.house.id, "name": r.house.name, "code": r.house.code} if r.house else None
        }
        data.append(item)
    result = {
        "total_activations": total_count, "excluded_count": excluded_count,
        "filtered_total": total_count - excluded_count, "excluded_tags": excluded_tags_list,
        "page": page, "page_size": page_size, "data": data
    }
    if not search and page == 1 and page_size <= 50:
        await cache_service.set(cache_key, result, ttl=120)
    return result


@router.get("/activations/daily-stats")
async def get_activation_daily_stats(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    exclude_tags: Optional[str] = Query(None),
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("reports.view")),
    house_id: Optional[int] = Depends(get_house_context),
    q_house_id: Optional[int] = Query(None, alias="house_id"),
):
    from app.services.cache_service import cache_service
    target_house_id = q_house_id or house_id
    if q_house_id and not is_admin_user(current_user):
        user_house_ids = [h.id for h in current_user.houses]
        if q_house_id not in user_house_ids:
            raise HTTPException(status_code=403, detail="You do not have access to this house")
    cache_key = cache_service.cache_key("daily_stats", year, month, exclude_tags, target_house_id)
    if not search:
        cached = await cache_service.get(cache_key)
        if cached:
            return cached
    from calendar import monthrange
    start_date = date(year, month, 1)
    last_day = monthrange(year, month)[1]
    end_date = date(year, month, last_day)

    excluded_codes = await get_excluded_codes(db)
    query = (
        select(Activation.activation_date, func.count())
        .select_from(Activation)
        .where(Activation.activation_date >= start_date)
        .where(Activation.activation_date <= end_date)
        .group_by(Activation.activation_date)
        .order_by(Activation.activation_date)
    )
    is_admin = is_admin_user(current_user)
    if not target_house_id and not is_admin:
        user_house_ids = [h.id for h in current_user.houses]
        if user_house_ids:
            query = query.where(Activation.house_id.in_(user_house_ids))
    elif target_house_id:
        query = query.where(Activation.house_id == target_house_id)
    clause = exclude_clause(Activation, excluded_codes)
    if clause is not None:
        query = query.where(clause)
    if search:
        p = f"%{search}%"
        query = query.where((Activation.sim_no.ilike(p)) | (Activation.retailer_code.ilike(p)) | (Activation.retailer_name.ilike(p)) | (Activation.msisdn.ilike(p)))
    excluded_tags_list = []
    if exclude_tags:
        excluded_tags_list = [t.strip() for t in exclude_tags.split(",") if t.strip()]
        if excluded_tags_list:
            house_ids_for_exclusion = [target_house_id] if target_house_id else ([h.id for h in current_user.houses] if not is_admin else None)
            excl_query = select(RetailerFilter.retailer_id).join(RetailerFilter.tag).where(FilterTag.name.in_(excluded_tags_list))
            if house_ids_for_exclusion:
                excl_query = excl_query.where(RetailerFilter.house_id.in_(house_ids_for_exclusion))
            excluded_ids_result = await db.execute(excl_query)
            excluded_retailer_ids = [row[0] for row in excluded_ids_result.all()]
            if excluded_retailer_ids:
                query = query.where(Activation.retailer_id.notin_(excluded_retailer_ids))
    rows = (await db.execute(query)).all()
    data_map: dict[str, int] = {}
    for row in rows:
        d = row.activation_date
        ds = d.isoformat() if isinstance(d, date) else str(d)
        data_map[ds] = row[1]
    result = []
    d = start_date
    while d <= end_date:
        ds = d.isoformat()
        result.append({"date": ds, "count": data_map.get(ds, 0)})
        d += timedelta(days=1)
    if not search:
        await cache_service.set(cache_key, result, ttl=300)
    return result


@router.get("/reports/live-activations")
async def get_ga_live_report(
    house_id: Optional[int] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("live_activations.view")),
):
    from app.services.ga_live_service import GaLiveQueryBuilder

    is_admin = is_admin_user(current_user)
    user_houses_raw = current_user.houses

    target_house_id = house_id
    if not target_house_id:
        if is_admin:
            raise HTTPException(status_code=400, detail="house_id is required for admin users")
        if len(user_houses_raw) == 1:
            target_house_id = user_houses_raw[0].id
        else:
            raise HTTPException(status_code=400, detail="house_id is required when user has multiple houses")
    else:
        if not is_admin:
            user_house_ids = [h.id for h in user_houses_raw]
            if target_house_id not in user_house_ids:
                raise HTTPException(status_code=403, detail="You do not have access to this house")

    today = date.today()
    if not start_date:
        start_date = today
    else:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
    if not end_date:
        end_date = today
    else:
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

    try:
        builder = GaLiveQueryBuilder(db, target_house_id, start_date, end_date)
        result = await builder.build_all()
        return result
    except Exception as e:
        logger.error(f"GaLive report failed for house_id={target_house_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {str(e)}")


@router.get("/reports/live-activations/live-activations-details")
async def get_employee_activation_details(
    employee_id: int = Query(..., description="Employee (not user) ID"),
    role_type: str = Query(..., description="rso or bp"),
    house_id: Optional[int] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("live_activations.view")),
):
    is_admin = is_admin_user(current_user)
    user_houses_raw = current_user.houses

    target_house_id = house_id
    if not target_house_id:
        if is_admin:
            raise HTTPException(status_code=400, detail="house_id is required for admin users")
        if len(user_houses_raw) == 1:
            target_house_id = user_houses_raw[0].id
        else:
            raise HTTPException(status_code=400, detail="house_id is required when user has multiple houses")
    else:
        if not is_admin:
            user_house_ids = [h.id for h in user_houses_raw]
            if target_house_id not in user_house_ids:
                raise HTTPException(status_code=403, detail="You do not have access to this house")

    today = date.today()
    if not start_date:
        start_date = today
    else:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
    if not end_date:
        end_date = today
    else:
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

    emp = await db.execute(
        select(Employee).where(
            Employee.id == employee_id,
            Employee.house_id == target_house_id,
        )
    )
    emp = emp.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    base_q = select(LiveActivation).where(
        LiveActivation.house_id == target_house_id,
        LiveActivation.activation_date >= start_date,
        LiveActivation.activation_date <= end_date,
    )

    retailer_codes: list[str] = []
    retailer_ids: list[int] = []

    if role_type == "rso":
        if emp.assisted_retailer_code:
            retailer_codes.append(emp.assisted_retailer_code)
        ret_rows = await db.execute(
            select(Retailer.id, Retailer.retailer_code, Retailer.name)
            .where(Retailer.house_id == target_house_id, Retailer.employee_id == employee_id)
        )
        for r in ret_rows.all():
            if r.id:
                retailer_ids.append(r.id)
            if r.retailer_code and r.retailer_code not in retailer_codes:
                retailer_codes.append(r.retailer_code)
    elif role_type == "bp":
        bp_code_rows = await db.execute(
            select(BpRetailerCode.retailer_code).where(
                BpRetailerCode.house_id == target_house_id,
                BpRetailerCode.bp_employee_id == employee_id,
            )
        )
        for (code,) in bp_code_rows.all():
            if code:
                retailer_codes.append(code)
        if emp.assisted_retailer_code:
            retailer_codes.append(emp.assisted_retailer_code)
    else:
        raise HTTPException(status_code=400, detail="Invalid role_type. Use 'rso' or 'bp'.")

    retailer_codes = list(set(c for c in retailer_codes if c))

    if not retailer_codes and not retailer_ids:
        return {"employee": {"id": emp.id, "name": emp.dms_code or f"#{emp.id}", "assisted_code": emp.assisted_retailer_code}, "groups": []}

    filters = []
    if retailer_codes:
        filters.append(LiveActivation.retailer_code.in_(retailer_codes))
    if retailer_ids:
        filters.append(LiveActivation.retailer_id.in_(retailer_ids))

    base_q = base_q.where(or_(*filters))
    base_q = base_q.order_by(LiveActivation.retailer_code, LiveActivation.product_code, LiveActivation.activation_time)

    result = await db.execute(base_q)
    records = result.scalars().all()

    emp_name = (
        (await db.execute(select(User.name).where(User.id == emp.user_id))).scalar()
        if emp.user_id else emp.dms_code or f"#{emp.id}"
    )

    groups: dict[str, dict] = {}
    for rec in records:
        rc = rec.retailer_code or "UNKNOWN"
        pc = rec.product_code or "UNKNOWN"
        if rc not in groups:
            groups[rc] = {"retailer_code": rc, "retailer_name": rec.retailer_name or "", "products": {}}
        if pc not in groups[rc]["products"]:
            groups[rc]["products"][pc] = {"product_code": pc, "product_name": rec.product_name or "", "records": []}
        groups[rc]["products"][pc]["records"].append({
            "retailer_code": rec.retailer_code,
            "retailer_name": rec.retailer_name,
            "activation_time": rec.activation_time,
            "product_code": rec.product_code,
            "product_name": rec.product_name,
            "sim_no": rec.sim_no,
            "msisdn": rec.msisdn,
            "dh_lifting_date": rec.dh_lifting_date,
            "issue_date": rec.issue_date,
            "selling_price": rec.selling_price,
            "subscription_type": rec.subscription_type,
            "service_class": rec.service_class,
        })

    result_groups = []
    for rc in sorted(groups.keys()):
        g = groups[rc]
        product_list = []
        for pc in sorted(g["products"].keys()):
            p = g["products"][pc]
            product_list.append({
                "product_code": p["product_code"],
                "product_name": p["product_name"],
                "count": len(p["records"]),
                "records": p["records"],
            })
        result_groups.append({
            "retailer_code": g["retailer_code"],
            "retailer_name": g["retailer_name"],
            "is_assisted": emp.assisted_retailer_code is not None and g["retailer_code"] == emp.assisted_retailer_code,
            "count": sum(pr["count"] for pr in product_list),
            "products": product_list,
        })

    return {
        "employee": {"id": emp.id, "name": emp_name, "assisted_code": emp.assisted_retailer_code},
        "groups": result_groups,
        "total_count": sum(g["count"] for g in result_groups),
    }


@router.get("/reports/target-achievement")
async def get_target_achievement(
    house_id: Optional[int] = Query(None),
    target_date: Optional[str] = Query(None),
    role_type: Optional[str] = Query(None, description="house, supervisor, or rso"),
    employee_id: Optional[int] = Query(None, description="Employee ID for supervisor or rso level"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("reports.target_achievement")),
):
    is_admin = is_admin_user(current_user)
    user_houses_raw = current_user.houses

    target_house_id = house_id
    if not target_house_id:
        if is_admin:
            raise HTTPException(status_code=400, detail="house_id is required")
        if len(user_houses_raw) == 1:
            target_house_id = user_houses_raw[0].id
        else:
            raise HTTPException(status_code=400, detail="house_id is required when user has multiple houses")
    else:
        if not is_admin:
            user_house_ids = [h.id for h in user_houses_raw]
            if target_house_id not in user_house_ids:
                raise HTTPException(status_code=403, detail="You do not have access to this house")

    if target_date:
        td = datetime.strptime(target_date, "%Y-%m-%d").date()
    else:
        today = date.today()
        td = date(today.year, today.month, 1)

    if td.day != 1:
        td = date(td.year, td.month, 1)

    service = TargetAchievementService(db, target_house_id, td)

    if role_type == "supervisor":
        if not employee_id:
            raise HTTPException(status_code=400, detail="employee_id is required for supervisor level")
        result = await service.get_supervisor_progress(employee_id)
    elif role_type == "rso":
        if not employee_id:
            raise HTTPException(status_code=400, detail="employee_id is required for rso level")
        result = await service.get_rso_progress(employee_id)
    else:
        result = await service.get_house_progress()

    return {"success": True, "data": result}


@router.get("/reports/target-achievement/export")
async def export_target_achievement(
    house_id: Optional[int] = Query(None),
    target_date: Optional[str] = Query(None),
    role_type: Optional[str] = Query(None),
    employee_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("reports.target_achievement")),
):
    from io import BytesIO
    from openpyxl import Workbook

    is_admin = is_admin_user(current_user)
    user_houses_raw = current_user.houses

    target_house_id = house_id
    if not target_house_id:
        if is_admin:
            raise HTTPException(status_code=400, detail="house_id is required")
        if len(user_houses_raw) == 1:
            target_house_id = user_houses_raw[0].id
        else:
            raise HTTPException(status_code=400, detail="house_id is required")

    if target_date:
        td = datetime.strptime(target_date, "%Y-%m-%d").date()
    else:
        today = date.today()
        td = date(today.year, today.month, 1)
    if td.day != 1:
        td = date(td.year, td.month, 1)

    service = TargetAchievementService(db, target_house_id, td)

    if role_type == "supervisor" and employee_id:
        result = await service.get_supervisor_progress(employee_id)
    elif role_type == "rso" and employee_id:
        result = await service.get_rso_progress(employee_id)
    else:
        result = await service.get_house_progress()

    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Target Achievement"
    ws.append(["Category", "Target", "Achieved", "Percentage", "Remaining", "Projected", "Status"])

    for cat in result.get("categories", []):
        ws.append([
            cat["label_en"],
            cat["target"],
            cat["achieved"],
            cat["percentage"],
            cat["remaining"],
            round(cat["projected"], 1),
            cat["status"],
        ])

    ws.append([])
    summary = result.get("summary", {})
    ws.append(["Overall", summary.get("total_target"), summary.get("total_achieved"),
               summary.get("overall_percentage"), "", "", ""])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=target_achievement_{td.isoformat()}.xlsx"},
    )


@router.get("/reports/activations/dashboard")
async def get_activation_dashboard(
    month: int = Query(None, ge=1, le=12),
    year: int = Query(None, ge=2020),
    exclude_tags: Optional[str] = Query(None, description="Comma-separated tag names to exclude for Achievement (e.g. DRC,RSP,BSP)"),
    exclude_codes: Optional[str] = Query(None, description="Comma-separated product codes to exclude for Achievement (e.g. SIMSWAP,EV-SWAP)"),
    rso_exclude_tags: Optional[str] = Query(None, description="Comma-separated tag names to exclude for RSO Performance"),
    rso_exclude_codes: Optional[str] = Query(None, description="Comma-separated product codes to exclude for RSO Performance"),
    rso_achieved_exclude_tags: Optional[str] = Query(None, description="Comma-separated tag names to exclude for RSO Achieved column"),
    rso_market_exclude_tags: Optional[str] = Query(None, description="Comma-separated tag names to exclude for RSO Market column"),
    rso_active_days_threshold: int = Query(1, ge=1, description="Minimum activations per day to count as active day for RSO"),
    bp_exclude_tags: Optional[str] = Query(None, description="Comma-separated tag names to exclude for BP Performance"),
    bp_exclude_codes: Optional[str] = Query(None, description="Comma-separated product codes to exclude for BP Performance"),
    cc_exclude_tags: Optional[str] = Query(None, description="Comma-separated tag names to exclude for CC Performance"),
    cc_exclude_codes: Optional[str] = Query(None, description="Comma-separated product codes to exclude for CC Performance"),
    supervisor_exclude_tags: Optional[str] = Query(None, description="Comma-separated tag names to exclude for Supervisor Performance"),
    supervisor_exclude_codes: Optional[str] = Query(None, description="Comma-separated product codes to exclude for Supervisor Performance"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("reports.view")),
    house_id: Optional[int] = Depends(get_house_context),
    q_house_id: Optional[int] = Query(None, alias="house_id"),
):
    from app.services.activation_report_service import ActivationReportService

    target_house_id = q_house_id or house_id
    if q_house_id and not is_admin_user(current_user):
        user_house_ids = [h.id for h in current_user.houses]
        if q_house_id not in user_house_ids:
            raise HTTPException(status_code=403, detail="You do not have access to this house")

    if not target_house_id:
        user_house_ids = [h.id for h in current_user.houses]
        if user_house_ids:
            target_house_id = user_house_ids[0]

    if not target_house_id:
        from app.models.house import House
        house_res = await db.execute(select(House.id).limit(1))
        first_house_id = house_res.scalar_one_or_none()
        if first_house_id:
            target_house_id = first_house_id

    if not target_house_id:
        raise HTTPException(status_code=400, detail="No house context available. Please select a house.")

    today = date.today()
    target_month = month or today.month
    target_year = year or today.year

    achievement_tag_list = [t.strip() for t in exclude_tags.split(",") if t.strip()] if exclude_tags else []

    if exclude_codes:
        achievement_code_set = {c.strip() for c in exclude_codes.split(",") if c.strip()}
    else:
        achievement_code_set = await get_excluded_codes(db)

    rso_tag_list = [t.strip() for t in rso_exclude_tags.split(",") if t.strip()] if rso_exclude_tags else []
    rso_code_set = {c.strip() for c in rso_exclude_codes.split(",") if c.strip()} if rso_exclude_codes else await get_excluded_codes(db)
    rso_achieved_tag_list = [t.strip() for t in rso_achieved_exclude_tags.split(",") if t.strip()] if rso_achieved_exclude_tags else []
    rso_market_tag_list = [t.strip() for t in rso_market_exclude_tags.split(",") if t.strip()] if rso_market_exclude_tags else []

    bp_tag_list = [t.strip() for t in bp_exclude_tags.split(",") if t.strip()] if bp_exclude_tags else []
    bp_code_set = {c.strip() for c in bp_exclude_codes.split(",") if c.strip()} if bp_exclude_codes else await get_excluded_codes(db)

    cc_tag_list = [t.strip() for t in cc_exclude_tags.split(",") if t.strip()] if cc_exclude_tags else []
    cc_code_set = {c.strip() for c in cc_exclude_codes.split(",") if c.strip()} if cc_exclude_codes else await get_excluded_codes(db)

    supervisor_tag_list = [t.strip() for t in supervisor_exclude_tags.split(",") if t.strip()] if supervisor_exclude_tags else []
    supervisor_code_set = {c.strip() for c in supervisor_exclude_codes.split(",") if c.strip()} if supervisor_exclude_codes else await get_excluded_codes(db)

    achievement_service = ActivationReportService(
        db, target_house_id, target_month, target_year,
        exclude_tag_names=achievement_tag_list,
        exclude_product_codes=achievement_code_set,
    )
    summary = await achievement_service.get_summary()
    daily_trend = await achievement_service.get_daily_trend()

    rso_service = ActivationReportService(
        db, target_house_id, target_month, target_year,
        exclude_tag_names=rso_tag_list,
        exclude_product_codes=rso_code_set,
        achieved_exclude_tag_names=rso_achieved_tag_list,
        market_exclude_tag_names=rso_market_tag_list,
        active_days_threshold=rso_active_days_threshold,
    )
    bp_service = ActivationReportService(
        db, target_house_id, target_month, target_year,
        exclude_tag_names=bp_tag_list,
        exclude_product_codes=bp_code_set,
    )
    cc_service = ActivationReportService(
        db, target_house_id, target_month, target_year,
        exclude_tag_names=cc_tag_list,
        exclude_product_codes=cc_code_set,
    )
    supervisor_service = ActivationReportService(
        db, target_house_id, target_month, target_year,
        exclude_tag_names=supervisor_tag_list,
        exclude_product_codes=supervisor_code_set,
    )
    rso = await rso_service.get_rso_performance()
    bp = await bp_service.get_bp_performance()
    cc = await cc_service.get_cc_performance()
    supervisor = await supervisor_service.get_supervisor_performance()
    top_performers = await rso_service.get_top_performers(rso, bp, cc, supervisor)

    return {
        "success": True,
        "summary": summary,
        "rso_performance": rso,
        "bp_performance": bp,
        "cc_performance": cc,
        "supervisor_performance": supervisor,
        "daily_trend": daily_trend,
        "top_performers": top_performers,
    }


class ActivationWhatsAppSendPayload(BaseModel):
    whatsapp_chat_id: str
    whatsapp_chat_name: Optional[str] = None
    caption: Optional[str] = None
    format: str = "image"  # image | text
    image_base64: Optional[str] = None
    text: Optional[str] = None
    month: int = Field(ge=1, le=12)
    year: int = Field(ge=2020)

    @field_validator("format")
    @classmethod
    def _validate_format(cls, v: str) -> str:
        if v not in ("image", "text"):
            raise ValueError("format must be 'image' or 'text'")
        return v

    @field_validator("whatsapp_chat_id")
    @classmethod
    def _validate_chat(cls, v: str) -> str:
        if not v.strip():
            raise ValueError("whatsapp_chat_id is required")
        return v.strip()


@router.post("/reports/activations/whatsapp")
async def send_activation_dashboard_whatsapp(
    payload: ActivationWhatsAppSendPayload,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("reports.whatsapp_share")),
    house_id: Optional[int] = Depends(get_house_context),
):
    """Share the Activation Dashboard report to a WhatsApp chat.

    The report content (rendered PNG or formatted text summary) is generated by
    the client so the recipient receives exactly what the user sees, including
    their exclusion-filter configuration. This endpoint only validates and
    forwards it through the house's WhatsApp gateway device.
    """
    target_house_id = house_id
    if not target_house_id:
        user_house_ids = [h.id for h in current_user.houses]
        if user_house_ids:
            target_house_id = user_house_ids[0]
    if not target_house_id:
        raise HTTPException(status_code=400, detail="No house context available. Please select a house.")

    house_res = await db.execute(select(House).where(House.id == target_house_id))
    house = house_res.scalar_one_or_none()
    if not house:
        raise HTTPException(status_code=400, detail="House not found.")
    wa_target = await resolve_house_wa_target(db, house)
    if not wa_target or not wa_target.jwt_token:
        raise HTTPException(status_code=400, detail="WhatsApp not configured for this house. Run setup first.")

    chat_name = payload.whatsapp_chat_name or payload.whatsapp_chat_id

    async def _log(action: str, status_code: int, extra: Optional[dict] = None):
        await log_activity(
            db,
            user_id=current_user.id,
            user_name=current_user.name,
            module="reports",
            action=action,
            record_identifier=chat_name,
            new_values={
                "house_id": target_house_id,
                "house_name": house.name,
                "month": payload.month,
                "year": payload.year,
                "chat": chat_name,
                "format": payload.format,
                **(extra or {}),
            },
            request=request,
            status_code=status_code,
        )

    if payload.format == "text":
        text = (payload.text or "").strip()
        if not text:
            raise HTTPException(status_code=400, detail="text is required for text format")
        if len(text) > 20000:
            raise HTTPException(status_code=400, detail="Text is too long (max 20000 characters)")
        chunks = [text[i:i + 3500] for i in range(0, len(text), 3500)]

        try:
            for chunk in chunks:
                await whatsapp_service_client.send_text(wa_target.jwt_token, payload.whatsapp_chat_id, chunk)
        except WhatsAppServiceError as e:
            await _log("whatsapp_send_failed", 502, {"error": f"{e.code}: {e.message}"})
            raise HTTPException(status_code=502, detail=f"{e.code}: {e.message}")

        await _log("whatsapp_send", 200, {"messages": len(chunks)})
        return {"success": True, "data": {"format": "text", "messages": len(chunks)}}

    # format == "image"
    raw_b64 = payload.image_base64 or ""
    b64_data = raw_b64.split(",", 1)[-1]
    try:
        image_bytes = base64.b64decode(b64_data, validate=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid base64 image data")
    if not image_bytes:
        raise HTTPException(status_code=400, detail="image_base64 is required for image format")
    if len(image_bytes) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Image too large (max 10MB)")
    caption = (payload.caption or f"{house.name} • Activation Report • {payload.month}/{payload.year}")[:1000]

    try:
        await whatsapp_service_client.send_image(
            jwt_token=wa_target.jwt_token,
            chat_jid=payload.whatsapp_chat_id,
            filename=f"activation_report_{payload.year}_{payload.month:02d}.png",
            image_bytes=image_bytes,
            caption=caption,
        )
    except WhatsAppServiceError as e:
        await _log("whatsapp_send_failed", 502, {"error": f"{e.code}: {e.message}"})
        raise HTTPException(status_code=502, detail=f"{e.code}: {e.message}")

    await _log("whatsapp_send", 200)
    return {"success": True, "data": {"format": "image"}}


@router.get("/reports/activations/dashboard/export")
async def export_activation_dashboard(
    month: int = Query(None, ge=1, le=12),
    year: int = Query(None, ge=2020),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("reports.download")),
    house_id: Optional[int] = Depends(get_house_context),
    q_house_id: Optional[int] = Query(None, alias="house_id"),
):
    from app.services.activation_report_service import ActivationReportService
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    target_house_id = q_house_id or house_id
    if q_house_id and not is_admin_user(current_user):
        user_house_ids = [h.id for h in current_user.houses]
        if q_house_id not in user_house_ids:
            raise HTTPException(status_code=403, detail="You do not have access to this house")

    if not target_house_id:
        user_house_ids = [h.id for h in current_user.houses]
        if user_house_ids:
            target_house_id = user_house_ids[0]

    if not target_house_id:
        from app.models.house import House
        house_res = await db.execute(select(House.id).limit(1))
        first_house_id = house_res.scalar_one_or_none()
        if first_house_id:
            target_house_id = first_house_id

    if not target_house_id:
        raise HTTPException(status_code=400, detail="No house context available. Please select a house.")

    today = date.today()
    target_month = month or today.month
    target_year = year or today.year

    excluded_product_codes = await get_excluded_codes(db)
    service = ActivationReportService(db, target_house_id, target_month, target_year, exclude_product_codes=excluded_product_codes)
    data = await service.build_dashboard()

    wb = Workbook()
    ws = wb.active
    ws.title = "Activation Report"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.cell(row=1, column=1, value=f"Activation Report - {target_year}-{target_month:02d}").font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')

    s = data["summary"]
    ws.cell(row=3, column=1, value="Summary").font = Font(bold=True, size=12)
    summary_headers = ["Metric", "Value"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    summary_rows = [
        ("Monthly Target", s["monthly_target"]),
        ("Achievement", s["achievement"]),
        ("Achievement %", f'{s["achievement_percentage"]}%'),
        ("Remaining", s["remaining"]),
        ("Daily Required", s["daily_required"]),
        ("Daily Average", s["daily_average"]),
        ("Projection", s["projection"]),
        ("Expected %", f'{s["expected_percentage"]}%'),
        ("Days Elapsed", s["days_elapsed"]),
        ("Days Remaining", s["days_remaining"]),
    ]
    for i, (label, val) in enumerate(summary_rows):
        ws.cell(row=5+i, column=1, value=label).border = thin_border
        ws.cell(row=5+i, column=2, value=val).border = thin_border

    for section_name, section_key in [("RSO Performance", "rso_performance"), ("BP Performance", "bp_performance"), ("CC Performance", "cc_performance"), ("Supervisor Performance", "supervisor_performance")]:
        items = data.get(section_key, [])
        row_offset = 20
        ws.cell(row=row_offset, column=1, value=section_name).font = Font(bold=True, size=12)
        cols = ["Name", "Target", "Achievement", "%", "Remaining", "Daily Avg", "Projection", "Status"]
        for col, h in enumerate(cols, 1):
            cell = ws.cell(row=row_offset+1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        for i, item in enumerate(items):
            r = row_offset + 2 + i
            ws.cell(row=r, column=1, value=item["name"]).border = thin_border
            ws.cell(row=r, column=2, value=item["target"]).border = thin_border
            ws.cell(row=r, column=3, value=item["achievement"]).border = thin_border
            ws.cell(row=r, column=4, value=f'{item["percentage"]}%').border = thin_border
            ws.cell(row=r, column=5, value=item["remaining"]).border = thin_border
            ws.cell(row=r, column=6, value=item["daily_average"]).border = thin_border
            ws.cell(row=r, column=7, value=item["projection"]).border = thin_border
            ws.cell(row=r, column=8, value=item["status"].replace("_", " ").title()).border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=activation_dashboard_{target_year}_{target_month:02d}.xlsx"},
    )


@router.get("/reports/recharge/dashboard")
async def get_recharge_dashboard(
    month: int = Query(None, ge=1, le=12),
    year: int = Query(None, ge=2020),
    report_type: str = Query("recharge", pattern="^(recharge|ev_secondary)$"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("reports.view")),
    house_id: Optional[int] = Depends(get_house_context),
    q_house_id: Optional[int] = Query(None, alias="house_id"),
):
    from app.services.recharge_report_service import RechargeReportService

    target_house_id = q_house_id or house_id
    if q_house_id and not is_admin_user(current_user):
        user_house_ids = [h.id for h in current_user.houses]
        if q_house_id not in user_house_ids:
            raise HTTPException(status_code=403, detail="You do not have access to this house")

    if not target_house_id:
        user_house_ids = [h.id for h in current_user.houses]
        if user_house_ids:
            target_house_id = user_house_ids[0]

    if not target_house_id:
        from app.models.house import House
        house_res = await db.execute(select(House.id).limit(1))
        first_house_id = house_res.scalar_one_or_none()
        if first_house_id:
            target_house_id = first_house_id

    if not target_house_id:
        raise HTTPException(status_code=400, detail="No house context available. Please select a house.")

    today = date.today()
    target_month = month or today.month
    target_year = year or today.year

    service = RechargeReportService(db, target_house_id, target_month, target_year, report_type)
    summary = await service.get_summary()
    daily_trend = await service.get_daily_trend()
    rso = await service.get_rso_performance()
    supervisor = await service.get_supervisor_performance()
    top_performers = await service.get_top_performers(rso, supervisor)

    return {
        "success": True,
        "summary": summary,
        "rso_performance": rso,
        "supervisor_performance": supervisor,
        "daily_trend": daily_trend,
        "top_performers": top_performers,
    }


@router.get("/reports/recharge/dashboard/export")
async def export_recharge_dashboard(
    month: int = Query(None, ge=1, le=12),
    year: int = Query(None, ge=2020),
    report_type: str = Query("recharge", pattern="^(recharge|ev_secondary)$"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("reports.download")),
    house_id: Optional[int] = Depends(get_house_context),
    q_house_id: Optional[int] = Query(None, alias="house_id"),
):
    from app.services.recharge_report_service import RechargeReportService
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    target_house_id = q_house_id or house_id
    if q_house_id and not is_admin_user(current_user):
        user_house_ids = [h.id for h in current_user.houses]
        if q_house_id not in user_house_ids:
            raise HTTPException(status_code=403, detail="You do not have access to this house")

    if not target_house_id:
        user_house_ids = [h.id for h in current_user.houses]
        if user_house_ids:
            target_house_id = user_house_ids[0]

    if not target_house_id:
        from app.models.house import House
        house_res = await db.execute(select(House.id).limit(1))
        first_house_id = house_res.scalar_one_or_none()
        if first_house_id:
            target_house_id = first_house_id

    if not target_house_id:
        raise HTTPException(status_code=400, detail="No house context available. Please select a house.")

    today = date.today()
    target_month = month or today.month
    target_year = year or today.year

    service = RechargeReportService(db, target_house_id, target_month, target_year, report_type)
    summary = await service.get_summary()
    rso = await service.get_rso_performance()
    supervisor = await service.get_supervisor_performance()

    wb = Workbook()
    ws = wb.active
    ws.title = "Recharge Report"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    report_label = "EV C2C Report" if report_type == "ev_secondary" else "Recharge Report (C2C)"
    ws.cell(row=1, column=1, value=f"{report_label} - {target_year}-{target_month:02d}").font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')

    s = summary
    ws.cell(row=3, column=1, value="Summary").font = Font(bold=True, size=12)
    summary_headers = ["Metric", "Value"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    summary_rows = [
        ("Total Target" if report_type == "ev_secondary" else "Total Recharge Target", s["monthly_target"]),
        ("EV C2C Target", s["ev_c2c_target"]),
        ("SC Primary Target", s["sc_primary_target"]),
        ("Achievement", s["achievement"]),
        ("Achievement %", f'{s["achievement_percentage"]}%'),
        ("Remaining", s["remaining"]),
        ("Daily Required", s["daily_required"]),
        ("Daily Average", s["daily_average"]),
        ("Projection", s["projection"]),
        ("Expected %", f'{s["expected_percentage"]}%'),
        ("Days Elapsed", s["days_elapsed"]),
        ("Days Remaining", s["days_remaining"]),
    ]
    for i, (label, val) in enumerate(summary_rows):
        ws.cell(row=5+i, column=1, value=label).border = thin_border
        ws.cell(row=5+i, column=2, value=val).border = thin_border

    for section_name, section_key in [("RSO Performance", "rso_performance"), ("Supervisor Performance", "supervisor_performance")]:
        items = rso if section_key == "rso_performance" else supervisor
        row_offset = 20
        ws.cell(row=row_offset, column=1, value=section_name).font = Font(bold=True, size=12)
        cols = ["Name", "Target", "EV Target", "SC Target", "Achievement", "%", "Remaining", "Daily Avg", "Projection", "Status"]
        for col, h in enumerate(cols, 1):
            cell = ws.cell(row=row_offset+1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        for i, item in enumerate(items):
            r = row_offset + 2 + i
            ws.cell(row=r, column=1, value=item["name"]).border = thin_border
            ws.cell(row=r, column=2, value=item["target"]).border = thin_border
            ws.cell(row=r, column=3, value=item.get("ev_target", 0)).border = thin_border
            ws.cell(row=r, column=4, value=item.get("sc_target", 0)).border = thin_border
            ws.cell(row=r, column=5, value=item["achievement"]).border = thin_border
            ws.cell(row=r, column=6, value=f'{item["percentage"]}%').border = thin_border
            ws.cell(row=r, column=7, value=item["remaining"]).border = thin_border
            ws.cell(row=r, column=8, value=item["daily_average"]).border = thin_border
            ws.cell(row=r, column=9, value=item["projection"]).border = thin_border
            ws.cell(row=r, column=10, value=item["status"].replace("_", " ").title()).border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=recharge_dashboard_{target_year}_{target_month:02d}.xlsx"},
    )


@router.get("/reports/my-target-progress")
async def get_my_target_progress(
    target_date: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("reports.target_achievement")),
):
    user_houses_raw = current_user.houses
    if not user_houses_raw:
        raise HTTPException(status_code=400, detail="User has no houses assigned")

    if target_date:
        td = datetime.strptime(target_date, "%Y-%m-%d").date()
    else:
        today = date.today()
        td = date(today.year, today.month, 1)

    if td.day != 1:
        td = date(td.year, td.month, 1)

    target_house_id = user_houses_raw[0].id
    service = TargetAchievementService(db, target_house_id, td)

    role_names = [r.name.lower() for r in current_user.roles]
    emp_profile = current_user.employee_profile

    if not emp_profile:
        return {"success": True, "data": await service.get_house_progress(), "user_role": "house"}

    if "rso" in role_names:
        result = await service.get_rso_progress(emp_profile.id)
        return {"success": True, "data": result, "user_role": "rso"}
    elif "supervisor" in role_names:
        result = await service.get_supervisor_progress(emp_profile.id)
        return {"success": True, "data": result, "user_role": "supervisor"}
    else:
        result = await service.get_house_progress()
        return {"success": True, "data": result, "user_role": "house"}


# --------------------------------------------------------------------------- #
# Active LSO (Target vs Achievement) report
# --------------------------------------------------------------------------- #

def _forbidden(msg: str = "You do not have access to this report data"):
    raise HTTPException(status_code=403, detail=msg)


async def _resolve_active_lso_scope(db, current_user, house_id, manager_id, supervisor_id, rso_id):
    """Role-aware scope resolution for the Active LSO report.

    Returns (house_id, manager_id, supervisor_id, rso_id) that the service may use.
    RSO/Supervisor accounts are forced to their own data; admin/manager accounts
    are limited to their assigned houses.
    """
    role_names = {r.name.strip().lower() for r in current_user.roles}
    emp = await get_employee_profile(db, current_user.id)

    if "rso" in role_names:
        if not emp or emp.employee_type != "rso" or not emp.house_id or emp.status != "Active":
            raise HTTPException(status_code=403, detail="No active RSO employee profile linked to this account")
        if house_id and house_id != emp.house_id:
            _forbidden()
        if manager_id or supervisor_id:
            _forbidden()
        if rso_id and rso_id != emp.id:
            _forbidden()
        return emp.house_id, None, None, emp.id

    if "supervisor" in role_names:
        if not emp or emp.employee_type != "supervisor" or not emp.house_id or emp.status != "Active":
            raise HTTPException(status_code=403, detail="No active supervisor employee profile linked to this account")
        if house_id and house_id != emp.house_id:
            _forbidden()
        if manager_id:
            _forbidden()
        if rso_id:
            sub_ids = await subordinate_rso_ids(db, emp)
            if rso_id not in sub_ids:
                _forbidden()
        return emp.house_id, None, emp.id, rso_id

    accessible = [h.id for h in current_user.houses]
    if not accessible:
        if is_admin_user(current_user):
            q = select(House.id)
            accessible = [r[0] for r in (await db.execute(q)).all()]
    if not accessible:
        raise HTTPException(status_code=403, detail="No house assigned to this account")
    if not house_id:
        house_id = accessible[0]
    if house_id not in accessible:
        _forbidden("Access denied to this house")

    if manager_id:
        m = await db.get(Employee, manager_id)
        if not m or m.house_id != house_id or m.employee_type != "manager" or m.status != "Active":
            raise HTTPException(status_code=400, detail="Invalid manager selection")
    if supervisor_id:
        s = await db.get(Employee, supervisor_id)
        if not s or s.house_id != house_id or s.employee_type != "supervisor" or s.status != "Active":
            raise HTTPException(status_code=400, detail="Invalid supervisor selection")
    if rso_id:
        r = await db.get(Employee, rso_id)
        if not r or r.house_id != house_id or r.employee_type != "rso" or r.status != "Active":
            raise HTTPException(status_code=400, detail="Invalid RSO selection")
        if supervisor_id:
            sup_emp = await db.get(Employee, supervisor_id)
            sub_ids = await subordinate_rso_ids(db, sup_emp)
            if rso_id not in sub_ids:
                _forbidden("RSO does not belong to the selected supervisor")
    return house_id, manager_id, supervisor_id, rso_id


def _resolve_active_lso_period(start_date, end_date):
    today = date.today()
    if not start_date and not end_date:
        start_date = date(today.year, today.month, 1)
        end_date = today
    elif not start_date:
        start_date = date(end_date.year, end_date.month, 1)
    elif not end_date:
        end_date = today
    if start_date > end_date:
        raise HTTPException(status_code=400, detail="start_date cannot be after end_date")
    return start_date, end_date


async def _build_active_lso_result(db, current_user, start_date, end_date, house_id, manager_id, supervisor_id, rso_id, status):
    if status and status not in VALID_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Allowed: {', '.join(VALID_STATUSES)}")
    start_date, end_date = _resolve_active_lso_period(start_date, end_date)
    house_id, manager_id, supervisor_id, rso_id = await _resolve_active_lso_scope(
        db, current_user, house_id, manager_id, supervisor_id, rso_id
    )
    service = ActiveLsoReportService(
        db, house_id, start_date, end_date,
        supervisor_id=supervisor_id, rso_id=rso_id, status_filter=status,
    )
    return await service.build_dashboard()


@router.get("/reports/active-lso/filters")
async def get_active_lso_report_filters(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("active_lso.view")),
):
    result = await _get_active_lso_filters(db, current_user)
    if result is None:
        raise HTTPException(status_code=403, detail="You do not have access to this report data")
    return result


@router.get("/reports/active-lso")
async def get_active_lso_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    house_id: Optional[int] = Query(None),
    manager_id: Optional[int] = Query(None),
    supervisor_id: Optional[int] = Query(None),
    rso_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("active_lso.view")),
    house_ctx: Optional[int] = Depends(get_house_context),
):
    resolved_house = house_id or house_ctx
    return await _build_active_lso_result(
        db, current_user, start_date, end_date,
        resolved_house, manager_id, supervisor_id, rso_id, status,
    )


@router.get("/reports/active-lso/config")
async def get_active_lso_report_config(
    house_id: Optional[int] = Query(None),
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020, le=2100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("active_lso.view")),
    house_ctx: Optional[int] = Depends(get_house_context),
):
    resolved_house = house_id or house_ctx
    if not resolved_house:
        user_house_ids = [h.id for h in current_user.houses]
        if not user_house_ids and is_admin_user(current_user):
            first_house = (await db.execute(select(House.id).order_by(House.id).limit(1))).scalar()
            resolved_house = first_house
        else:
            resolved_house = user_house_ids[0] if user_house_ids else None
    if not resolved_house:
        raise HTTPException(status_code=400, detail="No house context available. Please select a house.")

    month_start = date(year, month, 1)
    days, amount = await get_active_lso_thresholds(db, resolved_house, month_start)
    cfg = (
        await db.execute(
            select(ActiveLsoConfig).where(
                ActiveLsoConfig.house_id == resolved_house,
                ActiveLsoConfig.target_month == month_start,
            )
        )
    ).scalar_one_or_none()
    return {
        "success": True,
        "data": {
            "house_id": resolved_house,
            "month": month,
            "year": year,
            "days_threshold": days,
            "amount_threshold": amount,
            "is_custom": cfg is not None,
        },
    }


class ActiveLsoConfigPayload(BaseModel):
    house_id: Optional[int] = None
    month: int = Field(ge=1, le=12)
    year: int = Field(ge=2020, le=2100)
    days_threshold: int = Field(ge=1, le=31)
    amount_threshold: float = Field(ge=0)

    @field_validator("amount_threshold")
    @classmethod
    def _validate_amount(cls, v: float) -> float:
        if v > 10_000_000:
            raise ValueError("amount_threshold is too large")
        return round(float(v), 2)


@router.put("/reports/active-lso/config")
async def upsert_active_lso_report_config(
    payload: ActiveLsoConfigPayload,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("active_lso.config")),
    house_ctx: Optional[int] = Depends(get_house_context),
):
    resolved_house = payload.house_id or house_ctx
    if not resolved_house:
        user_house_ids = [h.id for h in current_user.houses]
        if not user_house_ids and is_admin_user(current_user):
            raise HTTPException(status_code=400, detail="house_id is required")
        resolved_house = user_house_ids[0] if user_house_ids else None
    if not resolved_house:
        raise HTTPException(status_code=400, detail="No house context available. Please select a house.")
    if not is_admin_user(current_user) and resolved_house not in [h.id for h in current_user.houses]:
        raise HTTPException(status_code=403, detail="You do not have access to this house")

    month_start = date(payload.year, payload.month, 1)
    old_cfg = (
        await db.execute(
            select(ActiveLsoConfig).where(
                ActiveLsoConfig.house_id == resolved_house,
                ActiveLsoConfig.target_month == month_start,
            )
        )
    ).scalar_one_or_none()

    old_values = None
    if old_cfg:
        old_values = {
            "days_threshold": old_cfg.days_threshold,
            "amount_threshold": old_cfg.amount_threshold,
        }
        old_cfg.days_threshold = payload.days_threshold
        old_cfg.amount_threshold = payload.amount_threshold
        old_cfg.updated_by = current_user.id
        config = old_cfg
    else:
        config = ActiveLsoConfig(
            house_id=resolved_house,
            target_month=month_start,
            days_threshold=payload.days_threshold,
            amount_threshold=payload.amount_threshold,
            created_by=current_user.id,
        )
        db.add(config)
    await db.commit()

    new_values = {
        "days_threshold": payload.days_threshold,
        "amount_threshold": payload.amount_threshold,
    }
    await log_activity(
        db,
        user_id=current_user.id,
        user_name=current_user.name,
        module="active_lso",
        action="edit",
        record_id=config.id,
        record_identifier=f"{payload.year}-{payload.month:02d}",
        old_values=old_values,
        new_values=new_values,
        request=request,
        status_code=200,
    )

    return {
        "success": True,
        "message": "Active LSO criteria updated",
        "data": {
            "house_id": resolved_house,
            "month": payload.month,
            "year": payload.year,
            "days_threshold": payload.days_threshold,
            "amount_threshold": payload.amount_threshold,
            "is_custom": True,
        },
    }


@router.get("/reports/active-lso/retailers/export")
async def export_inactive_retailers(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    house_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("active_lso.view")),
    house_ctx: Optional[int] = Depends(get_house_context),
):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from datetime import datetime

    resolved_house = house_id or house_ctx
    start_date, end_date = _resolve_active_lso_period(start_date, end_date)
    if not resolved_house:
        user_house_ids = [h.id for h in current_user.houses]
        if not user_house_ids and is_admin_user(current_user):
            first_house = (await db.execute(select(House.id).order_by(House.id).limit(1))).scalar()
            resolved_house = first_house
        else:
            resolved_house = user_house_ids[0] if user_house_ids else None
    if not resolved_house:
        raise HTTPException(status_code=400, detail="No house context available.")

    service = ActiveLsoReportService(db, resolved_house, start_date, end_date)
    groups = await service.get_all_inactive_retailers_grouped()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inactive Retailers"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    group_font = Font(bold=True, size=11, color="1F4E79")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    col_headers = [
        "House", "Retailer Code", "Retailer Name", "iTopUp No",
        "Days Sold", "Sales (BDT)", "Req Days", "Req Sales (BDT)", "Prev Month Inactive",
    ]
    col_widths = [15, 30, 15, 12, 10, 15, 10, 15, 18]

    today_str = datetime.now().strftime("%d %b %Y")
    current_row = 1

    for grp in groups:
        left_text = f"{today_str}  |  RSO: {grp['rso_name']} ({grp['dms_code']} - {grp['itop_number']})  |  Supervisor: {grp['supervisor_name']}"
        last_col = len(col_headers)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=last_col - 1)
        cell = ws.cell(row=current_row, column=1, value=left_text)
        cell.font = group_font
        cell.alignment = Alignment(horizontal="left")
        title_cell = ws.cell(row=current_row, column=last_col, value="Active LSO Report")
        title_cell.font = group_font
        title_cell.alignment = Alignment(horizontal="right")
        current_row += 1

        for ci, h in enumerate(col_headers, 1):
            cell = ws.cell(row=current_row, column=ci, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        for r in grp["retailers"]:
            remaining_sales = max(0, r["required_sales_amount"] - r["sales_amount"])
            remaining_days = max(0, r["required_selling_days"] - r["days_sold"])

            row_data = [
                r["house_code"], r["retailer_code"],
                r["name"], r["itop_number"],
                r["days_sold"],
                r["sales_amount"],
                f'{remaining_days}/{r["required_selling_days"]}',
                f'{format(round(remaining_sales), ",")}/{format(round(r["required_sales_amount"]), ",")}',
                r["inactive_last_month"] or "",
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=ci, value=val)
                cell.border = thin_border
                if ci in (5, 6, 7, 8):
                    cell.alignment = Alignment(horizontal="right")
            current_row += 1

        current_row += 1

    for ci, w in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from starlette.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=inactive_retailers_{today_str.replace(' ', '_')}.xlsx"
        },
    )


@router.get("/reports/active-lso/retailers/{employee_id}/export")
async def export_inactive_retailers_single(
    employee_id: int,
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    house_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("active_lso.view")),
    house_ctx: Optional[int] = Depends(get_house_context),
):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    resolved_house = house_id or house_ctx
    start_date, end_date = _resolve_active_lso_period(start_date, end_date)
    if not resolved_house:
        user_house_ids = [h.id for h in current_user.houses]
        if not user_house_ids and is_admin_user(current_user):
            first_house = (await db.execute(select(House.id).order_by(House.id).limit(1))).scalar()
            resolved_house = first_house
        else:
            resolved_house = user_house_ids[0] if user_house_ids else None
    if not resolved_house:
        raise HTTPException(status_code=400, detail="No house context available.")

    service = ActiveLsoReportService(db, resolved_house, start_date, end_date)
    retailers = await service.get_inactive_retailers(employee_id)

    emp = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalars().first()
    rso_name = ""
    dms_code = ""
    itop_number = ""
    supervisor_name = ""
    if emp:
        from app.services.active_lso_report_service import employee_name_map, supervisor_map
        nm = await employee_name_map(db, [emp.id])
        rso_name = nm.get(emp.id, "")
        dms_code = emp.dms_code or ""
        itop_number = emp.itop_number or ""
        target_date = date(start_date.year, start_date.month, 1)
        sup_map = await supervisor_map(db, [emp.id], target_date)
        sup_id = sup_map.get(emp.id)
        if sup_id:
            sup_nm = await employee_name_map(db, [sup_id])
            supervisor_name = sup_nm.get(sup_id, "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Inactive Retailers"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    group_font = Font(bold=True, size=11, color="1F4E79")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    col_headers = [
        "House", "Retailer Code", "Retailer Name", "iTopUp No",
        "Days Sold", "Sales (BDT)", "Req Days", "Req Sales (BDT)", "Prev Month Inactive",
    ]
    col_widths = [15, 30, 15, 12, 10, 15, 10, 15, 18]

    today_str = datetime.now().strftime("%d %b %Y")
    current_row = 1

    left_text = f"{today_str}  |  RSO: {rso_name} ({dms_code} - {itop_number})  |  Supervisor: {supervisor_name}"
    last_col = len(col_headers)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=last_col - 1)
    cell = ws.cell(row=current_row, column=1, value=left_text)
    cell.font = group_font
    cell.alignment = Alignment(horizontal="left")
    title_cell = ws.cell(row=current_row, column=last_col, value="Active LSO Report")
    title_cell.font = group_font
    title_cell.alignment = Alignment(horizontal="right")
    current_row += 1

    for ci, h in enumerate(col_headers, 1):
        cell = ws.cell(row=current_row, column=ci, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    current_row += 1

    for r in retailers:
        remaining_sales = max(0, r["required_sales_amount"] - r["sales_amount"])
        remaining_days = max(0, r["required_selling_days"] - r["days_sold"])
        row_data = [
            r["house_code"], r["retailer_code"], r["name"], r["itop_number"],
            r["days_sold"], r["sales_amount"],
            f'{remaining_days}/{r["required_selling_days"]}',
            f'{format(round(remaining_sales), ",")}/{format(round(r["required_sales_amount"]), ",")}',
            r["inactive_last_month"] or "",
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=ci, value=val)
            cell.border = thin_border
            if ci in (5, 6, 7, 8):
                cell.alignment = Alignment(horizontal="right")
        current_row += 1

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from starlette.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=inactive_retailers_{dms_code}_{today_str.replace(' ', '_')}.xlsx"
        },
    )


@router.get("/reports/active-lso/retailers/{employee_id}")
async def get_active_lso_inactive_retailers(
    employee_id: int,
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    house_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("active_lso.view")),
    house_ctx: Optional[int] = Depends(get_house_context),
):
    resolved_house = house_id or house_ctx
    start_date, end_date = _resolve_active_lso_period(start_date, end_date)
    if not resolved_house:
        user_house_ids = [h.id for h in current_user.houses]
        if not user_house_ids and is_admin_user(current_user):
            first_house = (await db.execute(select(House.id).order_by(House.id).limit(1))).scalar()
            resolved_house = first_house
        else:
            resolved_house = user_house_ids[0] if user_house_ids else None
    if not resolved_house:
        raise HTTPException(status_code=400, detail="No house context available.")
    service = ActiveLsoReportService(db, resolved_house, start_date, end_date)
    retailers = await service.get_inactive_retailers(employee_id)
    return {"success": True, "data": retailers}


@router.get("/reports/active-lso/export")
async def export_active_lso_report(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    house_id: Optional[int] = Query(None),
    manager_id: Optional[int] = Query(None),
    supervisor_id: Optional[int] = Query(None),
    rso_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("active_lso.export")),
    house_ctx: Optional[int] = Depends(get_house_context),
):
    from io import BytesIO, StringIO
    import csv as csv_module
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    resolved_house = house_id or house_ctx
    result = await _build_active_lso_result(
        db, current_user, start_date, end_date,
        resolved_house, manager_id, supervisor_id, rso_id, status,
    )

    period = result["period"]
    period_label = f"{period['start_date']} to {period['end_date']}"
    filename = f"active_lso_report_{period['start_date']}_to_{period['end_date']}.{format}"

    threshold_days = int(period.get("active_threshold_days") or DEFAULT_ACTIVE_LSO_DAYS)
    day_headers = [f"{i} Day" for i in range(threshold_days)]
    count_keys = [f"day_{i}" for i in range(threshold_days)] + \
        ["days_no_sales", "inactive_last_month", "reactivated"]

    headers = ["RSO", "Supervisor", "Target", "Achieved", "Ach %", "Remain",
               "DRR", "D.Avg", "Projection", "Status", "Retailers"] + day_headers + \
        ["Days No Sales", "Inactive Last Month", "Reactivated"]

    def row_values(item, include_name=True):
        base = [item.get("name", "Grand Total"), item.get("supervisor_name") or ""] if include_name else []
        base += [
            item["target"], item["achieved"], item["ach_pct"], item["remaining"],
            item["drr"], item["daily_avg"], item["projection"],
            item["status"].replace("_", " ").title(), item["retailer_count"],
        ]
        rc = item["retailer_counts"]
        base += [rc[k] for k in count_keys]
        return base

    if format == "csv":
        buf = StringIO()
        writer = csv_module.writer(buf)
        writer.writerow(["Active LSO Target vs Achievement Report"])
        writer.writerow([f"Period: {period_label}"])
        writer.writerow([f"Active threshold: {period['active_threshold_days']} days / {period['active_threshold_amount']} taka"])
        writer.writerow([])
        writer.writerow(headers)
        for r in result["rows"]:
            writer.writerow(row_values(r))
        writer.writerow(row_values(result["summary"]))
        writer.writerow([])
        writer.writerow(["Supervisor Summary"])
        sup_headers = ["Supervisor", "RSO Count", "Retailers", "Target", "Achieved",
                       "Ach %", "Remain", "DRR", "D.Avg", "Projection", "Status"]
        sup_headers += count_keys
        writer.writerow(sup_headers)
        for s in result["supervisor_summary"]:
            vals = [s["supervisor_name"], s["rso_count"], s["retailer_count"],
                    s["target"], s["achieved"], s["ach_pct"], s["remaining"],
                    s["drr"], s["daily_avg"], s["projection"],
                    s["status"].replace("_", " ").title()]
            rc = s["retailer_counts"]
            vals += [rc[k] for k in count_keys]
            writer.writerow(vals)
        return Response(
            content=buf.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Active LSO Report"

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    ws.cell(row=1, column=1, value="Active LSO Target vs Achievement Report").font = title_font
    ws.cell(row=2, column=1, value=f"Period: {period_label}").font = Font(bold=True, size=10)
    ws.cell(row=3, column=1, value=f"Active threshold: {period['active_threshold_days']} days / {period['active_threshold_amount']} taka").font = Font(size=10, italic=True)

    header_row = 5
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, r in enumerate(result["rows"]):
        for col, val in enumerate(row_values(r), 1):
            cell = ws.cell(row=header_row + 1 + i, column=col, value=val)
            cell.border = thin_border

    total_row = header_row + 1 + len(result["rows"])
    for col, val in enumerate(row_values(result["summary"]), 1):
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.border = thin_border
        cell.fill = total_fill
        cell.font = Font(bold=True)

    sup_start = total_row + 3
    ws.cell(row=sup_start, column=1, value="Supervisor Summary").font = section_font
    sup_headers = ["Supervisor", "RSO Count", "Retailers", "Target", "Achieved",
                   "Ach %", "Remain", "DRR", "D.Avg", "Projection", "Status"] + count_keys
    for col, h in enumerate(sup_headers, 1):
        cell = ws.cell(row=sup_start + 1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, s in enumerate(result["supervisor_summary"]):
        vals = [s["supervisor_name"], s["rso_count"], s["retailer_count"],
                s["target"], s["achieved"], s["ach_pct"], s["remaining"],
                s["drr"], s["daily_avg"], s["projection"],
                s["status"].replace("_", " ").title()]
        rc = s["retailer_counts"]
        vals += [rc[k] for k in count_keys]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=sup_start + 2 + i, column=col, value=val)
            cell.border = thin_border

    widths = [26, 22, 10, 10, 9, 9, 9, 9, 10, 14, 10] + [8] * threshold_days + [14, 16, 14]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# =====================================================================
#  ACTIVE SSO REPORT
# =====================================================================
from app.services.active_sso_report_service import (
    ActiveSsoReportService,
    get_active_sso_filters,
    get_active_sso_thresholds,
)


def _resolve_active_sso_period(start_date: Optional[date], end_date: Optional[date]) -> tuple:
    today = date.today()
    if start_date and end_date:
        return start_date, end_date
    return date(today.year, today.month, 1), today


@router.get("/reports/active-sso/filters")
async def active_sso_filters(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("active_sso.view")),
):
    result = await get_active_sso_filters(db, current_user)
    if not result:
        raise HTTPException(status_code=403, detail="Access denied")
    return result


@router.get("/reports/active-sso/config")
async def get_active_sso_config(
    house_id: Optional[int] = Query(None),
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020, le=2100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("active_sso.view")),
    house_ctx: Optional[int] = Depends(get_house_context),
):
    resolved_house = house_id or house_ctx
    if not resolved_house:
        user_house_ids = [h.id for h in current_user.houses]
        if not user_house_ids and is_admin_user(current_user):
            first_house = (await db.execute(select(House.id).order_by(House.id).limit(1))).scalar()
            resolved_house = first_house
        else:
            resolved_house = user_house_ids[0] if user_house_ids else None
    if not resolved_house:
        raise HTTPException(status_code=400, detail="No house context available.")

    month_start = date(year, month, 1)
    threshold = await get_active_sso_thresholds(db, resolved_house, month_start)

    res = await db.execute(
        select(ActiveSsoConfig).where(
            ActiveSsoConfig.house_id == resolved_house,
            ActiveSsoConfig.target_month == month_start,
        )
    )
    config = res.scalars().first()

    return {
        "success": True,
        "data": {
            "house_id": resolved_house,
            "month": month,
            "year": year,
            "activations_threshold": threshold,
            "is_custom": config is not None,
        },
    }


class ActiveSsoConfigPayload(BaseModel):
    house_id: Optional[int] = None
    month: int = Field(..., ge=1, le=12)
    year: int = Field(..., ge=2020, le=2100)
    activations_threshold: int = Field(..., ge=1, le=100)


@router.put("/reports/active-sso/config")
async def update_active_sso_config(
    payload: ActiveSsoConfigPayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("active_sso.config")),
    house_ctx: Optional[int] = Depends(get_house_context),
):
    resolved_house = payload.house_id or house_ctx
    if not resolved_house:
        user_house_ids = [h.id for h in current_user.houses]
        if not user_house_ids and is_admin_user(current_user):
            first_house = (await db.execute(select(House.id).order_by(House.id).limit(1))).scalar()
            resolved_house = first_house
        else:
            resolved_house = user_house_ids[0] if user_house_ids else None
    if not resolved_house:
        raise HTTPException(status_code=400, detail="No house context available.")

    month_start = date(payload.year, payload.month, 1)

    res = await db.execute(
        select(ActiveSsoConfig).where(
            ActiveSsoConfig.house_id == resolved_house,
            ActiveSsoConfig.target_month == month_start,
        )
    )
    config = res.scalars().first()

    if config:
        config.activations_threshold = payload.activations_threshold
        config.updated_by = current_user.id
    else:
        config = ActiveSsoConfig(
            house_id=resolved_house,
            target_month=month_start,
            activations_threshold=payload.activations_threshold,
            created_by=current_user.id,
            updated_by=current_user.id,
        )
        db.add(config)
    await db.commit()

    return {
        "success": True,
        "message": "Active SSO config updated successfully",
        "data": {
            "house_id": resolved_house,
            "month": payload.month,
            "year": payload.year,
            "activations_threshold": payload.activations_threshold,
            "is_custom": True,
        },
    }


@router.get("/reports/active-sso")
async def active_sso_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    house_id: Optional[int] = Query(None),
    supervisor_id: Optional[int] = Query(None),
    rso_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("active_sso.view")),
    house_ctx: Optional[int] = Depends(get_house_context),
):
    resolved_house = house_id or house_ctx
    start_date, end_date = _resolve_active_sso_period(start_date, end_date)
    if not resolved_house:
        user_house_ids = [h.id for h in current_user.houses]
        if not user_house_ids and is_admin_user(current_user):
            first_house = (await db.execute(select(House.id).order_by(House.id).limit(1))).scalar()
            resolved_house = first_house
        else:
            resolved_house = user_house_ids[0] if user_house_ids else None
    if not resolved_house:
        raise HTTPException(status_code=400, detail="No house context available.")

    service = ActiveSsoReportService(
        db, resolved_house, start_date, end_date,
        supervisor_id=supervisor_id, rso_id=rso_id, status_filter=status,
    )
    return await service.build_dashboard()


@router.get("/reports/active-sso/retailers/export")
async def export_inactive_sso_retailers(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    house_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("active_sso.view")),
    house_ctx: Optional[int] = Depends(get_house_context),
):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    resolved_house = house_id or house_ctx
    start_date, end_date = _resolve_active_sso_period(start_date, end_date)
    if not resolved_house:
        user_house_ids = [h.id for h in current_user.houses]
        if not user_house_ids and is_admin_user(current_user):
            first_house = (await db.execute(select(House.id).order_by(House.id).limit(1))).scalar()
            resolved_house = first_house
        else:
            resolved_house = user_house_ids[0] if user_house_ids else None
    if not resolved_house:
        raise HTTPException(status_code=400, detail="No house context available.")

    service = ActiveSsoReportService(db, resolved_house, start_date, end_date)
    groups = await service.get_all_inactive_retailers_grouped()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inactive SSO Retailers"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    group_font = Font(bold=True, size=11, color="1F4E79")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    col_headers = [
        "House", "Retailer Code", "Retailer Name", "iTopUp No",
        "Activations Done", "Req. Activations", "Prev Month Inactive",
    ]
    col_widths = [12, 15, 30, 15, 15, 15, 18]

    today_str = datetime.now().strftime("%d %b %Y")
    current_row = 1
    last_col = len(col_headers)

    for grp in groups:
        left_text = f"{today_str}  |  RSO: {grp['rso_name']} ({grp['dms_code']} - {grp['itop_number']})  |  Supervisor: {grp['supervisor_name']}"
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=last_col - 1)
        cell = ws.cell(row=current_row, column=1, value=left_text)
        cell.font = group_font
        cell.alignment = Alignment(horizontal="left")
        title_cell = ws.cell(row=current_row, column=last_col, value="Active SSO Report")
        title_cell.font = group_font
        title_cell.alignment = Alignment(horizontal="right")
        current_row += 1

        for ci, h in enumerate(col_headers, 1):
            cell = ws.cell(row=current_row, column=ci, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        for r in grp["retailers"]:
            row_data = [
                r["house_code"], r["retailer_code"], r["name"], r["itop_number"],
                r["activations_done"], r["required_activations"],
                r["inactive_last_month"] or "",
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=ci, value=val)
                cell.border = thin_border
                if ci in (5, 6):
                    cell.alignment = Alignment(horizontal="right")
            current_row += 1

        current_row += 1

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from starlette.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=inactive_sso_retailers_{today_str.replace(' ', '_')}.xlsx"
        },
    )


@router.get("/reports/active-sso/retailers/{employee_id}/export")
async def export_inactive_sso_retailers_single(
    employee_id: int,
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    house_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("active_sso.view")),
    house_ctx: Optional[int] = Depends(get_house_context),
):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    resolved_house = house_id or house_ctx
    start_date, end_date = _resolve_active_sso_period(start_date, end_date)
    if not resolved_house:
        user_house_ids = [h.id for h in current_user.houses]
        if not user_house_ids and is_admin_user(current_user):
            first_house = (await db.execute(select(House.id).order_by(House.id).limit(1))).scalar()
            resolved_house = first_house
        else:
            resolved_house = user_house_ids[0] if user_house_ids else None
    if not resolved_house:
        raise HTTPException(status_code=400, detail="No house context available.")

    service = ActiveSsoReportService(db, resolved_house, start_date, end_date)
    retailers = await service.get_inactive_retailers(employee_id)

    emp = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalars().first()
    rso_name = ""
    dms_code = ""
    itop_number = ""
    supervisor_name = ""
    if emp:
        from app.services.active_lso_report_service import employee_name_map, supervisor_map
        nm = await employee_name_map(db, [emp.id])
        rso_name = nm.get(emp.id, "")
        dms_code = emp.dms_code or ""
        itop_number = emp.itop_number or ""
        target_date = date(start_date.year, start_date.month, 1)
        sup_map = await supervisor_map(db, [emp.id], target_date)
        sup_id = sup_map.get(emp.id)
        if sup_id:
            sup_nm = await employee_name_map(db, [sup_id])
            supervisor_name = sup_nm.get(sup_id, "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Inactive SSO Retailers"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    group_font = Font(bold=True, size=11, color="1F4E79")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    col_headers = [
        "House", "Retailer Code", "Retailer Name", "iTopUp No",
        "Activations Done", "Req. Activations", "Prev Month Inactive",
    ]
    col_widths = [12, 15, 30, 15, 15, 15, 18]

    today_str = datetime.now().strftime("%d %b %Y")
    current_row = 1
    last_col = len(col_headers)

    left_text = f"{today_str}  |  RSO: {rso_name} ({dms_code} - {itop_number})  |  Supervisor: {supervisor_name}"
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=last_col - 1)
    cell = ws.cell(row=current_row, column=1, value=left_text)
    cell.font = group_font
    cell.alignment = Alignment(horizontal="left")
    title_cell = ws.cell(row=current_row, column=last_col, value="Active SSO Report")
    title_cell.font = group_font
    title_cell.alignment = Alignment(horizontal="right")
    current_row += 1

    for ci, h in enumerate(col_headers, 1):
        cell = ws.cell(row=current_row, column=ci, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    current_row += 1

    for r in retailers:
        row_data = [
            r["house_code"], r["retailer_code"], r["name"], r["itop_number"],
            r["activations_done"], r["required_activations"],
            r["inactive_last_month"] or "",
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=ci, value=val)
            cell.border = thin_border
            if ci in (5, 6):
                cell.alignment = Alignment(horizontal="right")
        current_row += 1

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from starlette.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=inactive_sso_{dms_code}_{today_str.replace(' ', '_')}.xlsx"
        },
    )


@router.get("/reports/active-sso/retailers/{employee_id}")
async def get_active_sso_inactive_retailers(
    employee_id: int,
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    house_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("active_sso.view")),
    house_ctx: Optional[int] = Depends(get_house_context),
):
    resolved_house = house_id or house_ctx
    start_date, end_date = _resolve_active_sso_period(start_date, end_date)
    if not resolved_house:
        user_house_ids = [h.id for h in current_user.houses]
        if not user_house_ids and is_admin_user(current_user):
            first_house = (await db.execute(select(House.id).order_by(House.id).limit(1))).scalar()
            resolved_house = first_house
        else:
            resolved_house = user_house_ids[0] if user_house_ids else None
    if not resolved_house:
        raise HTTPException(status_code=400, detail="No house context available.")
    service = ActiveSsoReportService(db, resolved_house, start_date, end_date)
    retailers = await service.get_inactive_retailers(employee_id)
    return {"success": True, "data": retailers}
