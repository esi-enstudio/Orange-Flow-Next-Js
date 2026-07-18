from typing import Optional
from datetime import datetime, date, timedelta
import logging
import io
from collections import defaultdict

import os
import shutil
from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile, Response
from pydantic import BaseModel
from sqlalchemy import select, func, and_, case, delete as sa_delete, literal_column, String
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload, selectinload

from app.routers.deps import get_db, has_permission, get_house_context, get_current_user
from app.schemas.zoom_in import (
    ZoomInEventTypeSchema, ZoomInEventTypeCreate, ZoomInEventTypeUpdate,
    ZoomInActivitySchema, ZoomInActivityCreate, ZoomInActivityUpdate,
    ZoomInAllocationCreate, ZoomInAllocationUpdate, ZoomInAllocationResponse,
    ZoomInEventCreate, ZoomInEventUpdate, ZoomInEventResponse,
    CurrentMonthSummary, AllocationCheckResponse, DashboardSummaryResponse,
    BulkAllocationCreate, BulkAllocationItem,
)
from app.schemas.pagination import PaginationParams, PaginatedResponse, PaginationMeta
from app.models.zoom_in import (
    ZoomInEventType, ZoomInActivity,
    ZoomInAllocation, ZoomInEvent,
    ZoomInEventBTS, ZoomInEventRSO, ZoomInEventBP, ZoomInEventRetailer,
)
from app.models.mela import MelaEligibleBTS
from app.models.bts import BTS
from app.models.employee import Employee
from app.models.retailer import Retailer
from app.models.house import House
from app.models.activation import Activation
from app.models.live_activation import LiveActivation
from app.models.product_exclusion import ExcludedProductCode
from app.models.user import User
from app.utils.access_control import is_admin_user
from app.utils.validation import safe_filename, validate_excel
from app.services.Automation.mela_eligible_bts_excel import process_eligible_bts_excel, generate_eligible_bts_sample_bytes

class EligibleBTSAttach(BaseModel):
    bts_code: str
    house_code: str


logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/zoom-in", tags=["zoom_in"])


# ─── Config Endpoints ───────────────────────────────────────────

@router.get("/event-types", response_model=list[ZoomInEventTypeSchema])
async def get_event_types(
    include_inactive: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
):
    query = select(ZoomInEventType)
    if not include_inactive:
        query = query.where(ZoomInEventType.is_active == True)
    result = await db.execute(query.order_by(ZoomInEventType.id))
    return result.scalars().all()


@router.post("/event-types", status_code=201)
async def create_event_type(
    data: ZoomInEventTypeCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.create")),
):
    existing = await db.execute(
        select(ZoomInEventType).where(ZoomInEventType.name == data.name)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Event type with this name already exists")

    event_type = ZoomInEventType(name=data.name, name_bn=data.name_bn)
    db.add(event_type)
    await db.commit()
    await db.refresh(event_type)
    return {"success": True, "message": "Event type created", "id": event_type.id}


@router.put("/event-types/{event_type_id}")
async def update_event_type(
    event_type_id: int,
    data: ZoomInEventTypeUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.edit")),
):
    event_type = await db.get(ZoomInEventType, event_type_id)
    if not event_type:
        raise HTTPException(status_code=404, detail="Event type not found")

    if data.name is not None:
        existing = await db.execute(
            select(ZoomInEventType).where(
                ZoomInEventType.name == data.name,
                ZoomInEventType.id != event_type_id,
            )
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=409, detail="Event type with this name already exists")
        event_type.name = data.name
    if data.name_bn is not None:
        event_type.name_bn = data.name_bn
    if data.is_active is not None:
        event_type.is_active = data.is_active

    await db.commit()
    return {"success": True, "message": "Event type updated"}


@router.delete("/event-types/{event_type_id}")
async def delete_event_type(
    event_type_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.delete")),
):
    event_type = await db.get(ZoomInEventType, event_type_id)
    if not event_type:
        raise HTTPException(status_code=404, detail="Event type not found")

    await db.delete(event_type)
    await db.commit()
    return {"success": True, "message": "Event type deleted"}


@router.get("/activities", response_model=list[ZoomInActivitySchema])
async def get_activities(
    include_inactive: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
):
    query = select(ZoomInActivity)
    if not include_inactive:
        query = query.where(ZoomInActivity.is_active == True)
    result = await db.execute(query.order_by(ZoomInActivity.id))
    return result.scalars().all()


@router.post("/activities", status_code=201)
async def create_activity(
    data: ZoomInActivityCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.create")),
):
    existing = await db.execute(
        select(ZoomInActivity).where(ZoomInActivity.name == data.name)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Activity with this name already exists")

    activity = ZoomInActivity(name=data.name, name_bn=data.name_bn)
    db.add(activity)
    await db.commit()
    await db.refresh(activity)
    return {"success": True, "message": "Activity created", "id": activity.id}


@router.put("/activities/{activity_id}")
async def update_activity(
    activity_id: int,
    data: ZoomInActivityUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.edit")),
):
    activity = await db.get(ZoomInActivity, activity_id)
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found")

    if data.name is not None:
        existing = await db.execute(
            select(ZoomInActivity).where(
                ZoomInActivity.name == data.name,
                ZoomInActivity.id != activity_id,
            )
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=409, detail="Activity with this name already exists")
        activity.name = data.name
    if data.name_bn is not None:
        activity.name_bn = data.name_bn
    if data.is_active is not None:
        activity.is_active = data.is_active

    await db.commit()
    return {"success": True, "message": "Activity updated"}


@router.delete("/activities/{activity_id}")
async def delete_activity(
    activity_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.delete")),
):
    activity = await db.get(ZoomInActivity, activity_id)
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found")

    await db.delete(activity)
    await db.commit()
    return {"success": True, "message": "Activity deleted"}


# ─── Lookup Endpoints ───────────────────────────────────────────

@router.get("/thanas")
async def get_thanas(
    house_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
):
    query = select(BTS.thana).where(
        BTS.is_eligible == True,
        BTS.thana.isnot(None),
        BTS.thana != "",
    ).distinct().order_by(BTS.thana)
    if house_id:
        query = query.where(BTS.house_id == house_id)
    result = await db.execute(query)
    return [row[0] for row in result.all()]


@router.get("/bts-by-thana/{thana}")
async def get_bts_by_thana(
    thana: str,
    house_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
):
    query = select(BTS).where(
        BTS.thana.ilike(thana),
        BTS.is_eligible == True,
    )
    if house_id:
        query = query.where(BTS.house_id == house_id)
    result = await db.execute(query.order_by(BTS.site_id))
    bts_list = result.scalars().all()
    return [
        {"id": b.id, "site_id": b.site_id, "bts_code": b.bts_code, "address": b.short_address or b.address}
        for b in bts_list
    ]


@router.get("/rsos-by-house/{house_id}")
async def get_rsos_by_house(
    house_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
):
    result = await db.execute(
        select(Employee)
        .options(joinedload(Employee.user))
        .where(
            Employee.house_id == house_id,
            Employee.employee_type == "rso",
            Employee.status == "Active",
        ).order_by(Employee.dms_code)
    )
    employees = result.scalars().all()
    return [
        {
            "id": e.id,
            "dms_code": e.dms_code,
            "itop_number": e.itop_number,
            "pool_number": e.pool_number,
            "name": e.user.name if e.user else None,
            "employee_id": e.employee_id,
            "assisted_retailer_code": e.assisted_retailer_code,
        }
        for e in employees
    ]


@router.get("/bps-by-house/{house_id}")
async def get_bps_by_house(
    house_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
):
    result = await db.execute(
        select(Employee)
        .options(joinedload(Employee.user))
        .where(
            Employee.house_id == house_id,
            Employee.employee_type == "bp",
            Employee.status == "Active",
        ).order_by(Employee.dms_code)
    )
    employees = result.scalars().all()
    return [
        {
            "id": e.id,
            "dms_code": e.dms_code,
            "itop_number": e.itop_number,
            "pool_number": e.pool_number,
            "name": e.user.name if e.user else None,
            "employee_id": e.employee_id,
            "assisted_retailer_code": e.assisted_retailer_code,
        }
        for e in employees
    ]


@router.get("/retailers-by-rso")
async def get_retailers_by_rso(
    house_id: int = Query(..., description="House ID to fetch all retailers for"),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
):
    excl_result = await db.execute(select(ExcludedProductCode.product_code))
    excluded_codes = {row[0] for row in excl_result.all()}

    today = date.today()
    first_of_month = today.replace(day=1)
    last_month_end = first_of_month - timedelta(days=1)
    last_month_start = last_month_end.replace(day=1)

    emp_result = await db.execute(
        select(Employee.id, Employee.itop_number, Employee.assisted_retailer_code).where(
            Employee.house_id == house_id,
            Employee.employee_type.in_(["rso", "bp"]),
            Employee.status == "Active",
        )
    )
    employee_rows = emp_result.all()
    employee_ids = [row.id for row in employee_rows]
    emp_itop: dict[int, str | None] = {row.id: row.itop_number for row in employee_rows}

    emp_exclude_codes: set[str] = set()
    for row in employee_rows:
        if row.assisted_retailer_code:
            emp_exclude_codes.add(row.assisted_retailer_code)

    query = select(Retailer).where(
        Retailer.employee_id.in_(employee_ids),
        Retailer.enabled == "Yes",
        Retailer.sim_seller == "Yes",
    )
    if emp_exclude_codes:
        query = query.where(~Retailer.retailer_code.in_(list(emp_exclude_codes)))
    if search:
        pattern = f"%{search}%"
        query = query.where(
            (Retailer.retailer_code.ilike(pattern)) |
            (Retailer.name.ilike(pattern))
        )
    result = await db.execute(query)
    retailers = result.scalars().all()

    retailer_codes = [r.retailer_code for r in retailers]
    activation_counts: dict[str, int] = {}
    if retailer_codes:
        count_query = select(
            Activation.retailer_code,
            func.count(Activation.id).label("cnt")
        ).where(
            Activation.retailer_code.in_(retailer_codes),
            Activation.activation_date >= last_month_start,
            Activation.activation_date <= last_month_end,
        )
        if excluded_codes:
            count_query = count_query.where(
                ~Activation.product_code.in_(list(excluded_codes))
            )
        count_query = count_query.group_by(Activation.retailer_code)
        count_result = await db.execute(count_query)
        for row in count_result.all():
            activation_counts[row[0]] = row[1]

    result_data = [
        {
            "retailer_code": r.retailer_code,
            "name": r.name,
            "itop_number": r.itop_number,
            "sim_seller": r.sim_seller,
            "rso_itop_last3": emp_itop.get(r.employee_id, "")[-3:] if emp_itop.get(r.employee_id) else None,
            "activation_count": activation_counts.get(r.retailer_code, 0),
        }
        for r in retailers
    ]
    result_data.sort(key=lambda x: x["activation_count"], reverse=True)
    return result_data


# ─── Allocation Endpoints ──────────────────────────────────────

@router.get("/allocations/summary")
async def get_allocations_summary(
    month: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
    house_id: Optional[int] = Depends(get_house_context),
):
    if month:
        target_month = datetime.strptime(month, "%Y-%m").date()
    else:
        today = date.today()
        target_month = date(today.year, today.month, 1)

    query = select(ZoomInAllocation).options(
        joinedload(ZoomInAllocation.event_type),
        joinedload(ZoomInAllocation.house),
    ).where(
        ZoomInAllocation.month == target_month,
        ZoomInAllocation.is_deleted == False,
    )

    effective_house_id = house_id
    if not effective_house_id:
        user_house_ids = [h.id for h in current_user.houses]
        if not is_admin_user(current_user) and user_house_ids:
            effective_house_id = user_house_ids[0]

    if effective_house_id:
        query = query.where(ZoomInAllocation.house_id == effective_house_id)

    result = await db.execute(query)
    allocations = result.unique().scalars().all()

    total_budget = sum(a.total_budget for a in allocations)
    event_type_summaries = {}
    for a in allocations:
        name = a.event_type.name if a.event_type else f"Type {a.event_type_id}"
        if name not in event_type_summaries:
            event_type_summaries[name] = {"event_type": name, "count": 0, "budget": 0}
        event_type_summaries[name]["count"] += a.count
        event_type_summaries[name]["budget"] += a.total_budget

    return {
        "total_budget": total_budget,
        "event_type_summaries": list(event_type_summaries.values()),
    }


@router.post("/allocations")
async def create_allocation(
    data: ZoomInAllocationCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.create")),
):
    existing = await db.execute(
        select(ZoomInAllocation).where(
            ZoomInAllocation.house_id == data.house_id,
            ZoomInAllocation.month == data.month,
            ZoomInAllocation.event_type_id == data.event_type_id,
            ZoomInAllocation.thana == data.thana,
            ZoomInAllocation.is_deleted == False,
        )
    )
    existing_allocation = existing.scalar_one_or_none()
    if existing_allocation:
        existing_allocation.count = data.count
        existing_allocation.budget_per_unit = data.budget_per_unit
        existing_allocation.total_budget = data.count * data.budget_per_unit
        existing_allocation.updated_by = current_user.id
        await db.commit()
        await db.refresh(existing_allocation)
        return {"success": True, "message": "Allocation updated", "id": existing_allocation.id}

    allocation = ZoomInAllocation(
        house_id=data.house_id,
        month=data.month,
        event_type_id=data.event_type_id,
        thana=data.thana,
        count=data.count,
        budget_per_unit=data.budget_per_unit,
        total_budget=data.count * data.budget_per_unit,
        created_by=current_user.id,
    )
    db.add(allocation)
    await db.commit()
    await db.refresh(allocation)
    return {"success": True, "message": "Allocation created", "id": allocation.id}


@router.post("/allocations/bulk")
async def bulk_create_allocations(
    data: BulkAllocationCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.create")),
):
    created = 0
    updated = 0
    for item in data.allocations:
        existing = await db.execute(
            select(ZoomInAllocation).where(
                ZoomInAllocation.house_id == data.house_id,
                ZoomInAllocation.month == data.month,
                ZoomInAllocation.event_type_id == item.event_type_id,
                ZoomInAllocation.thana == item.thana,
                ZoomInAllocation.is_deleted == False,
            )
        )
        existing_allocation = existing.scalar_one_or_none()
        if existing_allocation:
            existing_allocation.count = item.count
            existing_allocation.budget_per_unit = item.budget_per_unit
            existing_allocation.total_budget = item.count * item.budget_per_unit
            existing_allocation.updated_by = current_user.id
            updated += 1
        else:
            allocation = ZoomInAllocation(
                house_id=data.house_id,
                month=data.month,
                event_type_id=item.event_type_id,
                thana=item.thana,
                count=item.count,
                budget_per_unit=item.budget_per_unit,
                total_budget=item.count * item.budget_per_unit,
                created_by=current_user.id,
            )
            db.add(allocation)
            created += 1
    await db.commit()
    return {
        "success": True,
        "message": f"{created} created, {updated} updated",
        "created": created,
        "updated": updated,
    }


@router.delete("/allocations/{allocation_id}")
async def delete_allocation(
    allocation_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.delete")),
):
    allocation = await db.get(ZoomInAllocation, allocation_id)
    if not allocation or allocation.is_deleted:
        raise HTTPException(status_code=404, detail="Allocation not found")
    allocation.is_deleted = True
    allocation.deleted_at = datetime.utcnow()
    allocation.deleted_by = current_user.id
    await db.commit()
    return {"success": True, "message": "Allocation deleted"}


@router.get("/allocations")
async def get_allocations(
    month: str = Query(default=None, description="Month (YYYY-MM)"),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query("month"),
    sort_order: Optional[str] = Query("desc"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
    house_id: Optional[int] = Depends(get_house_context),
):
    query = select(ZoomInAllocation).options(
        joinedload(ZoomInAllocation.event_type),
        joinedload(ZoomInAllocation.house),
    ).where(
        ZoomInAllocation.is_deleted == False,
    )

    if month:
        try:
            target_month = datetime.strptime(month, "%Y-%m").date()
            query = query.where(ZoomInAllocation.month == target_month)
        except ValueError:
            pass

    effective_house_id = house_id
    if not effective_house_id:
        user_house_ids = [h.id for h in current_user.houses]
        if not is_admin_user(current_user) and user_house_ids:
            effective_house_id = user_house_ids[0]

    if effective_house_id:
        query = query.where(ZoomInAllocation.house_id == effective_house_id)

    if search:
        search_pattern = f"%{search}%"
        query = query.where(
            ZoomInAllocation.month.cast(String).ilike(search_pattern)
        )

    count_query = select(func.count()).select_from(query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    sort_col = getattr(ZoomInAllocation, sort_by, ZoomInAllocation.month)
    order = sort_col.asc() if sort_order == "asc" else sort_col.desc()
    offset = (page - 1) * per_page
    result = await db.execute(query.offset(offset).limit(per_page).order_by(order))
    allocations = result.unique().scalars().all()

    total_pages = max(1, -(-total // per_page))
    data = []
    for a in allocations:
        data.append({
            "id": a.id,
            "house_id": a.house_id,
            "month": str(a.month),
            "event_type_id": a.event_type_id,
            "thana": a.thana,
            "count": a.count,
            "budget_per_unit": a.budget_per_unit,
            "total_budget": a.total_budget,
            "house_name": a.house.name if a.house else None,
            "event_type_name": a.event_type.name if a.event_type else None,
        })

    return PaginatedResponse(
        success=True,
        data=data,
        pagination=PaginationMeta(
            page=page,
            per_page=per_page,
            total=total,
            total_pages=total_pages,
            has_next=page < total_pages,
            has_prev=page > 1,
        ),
    )


@router.put("/allocations/{allocation_id}")
async def update_allocation(
    allocation_id: int,
    data: ZoomInAllocationUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.edit")),
):
    allocation = await db.get(ZoomInAllocation, allocation_id)
    if not allocation or allocation.is_deleted:
        raise HTTPException(status_code=404, detail="Allocation not found")
    allocation.count = data.count
    allocation.budget_per_unit = data.budget_per_unit
    allocation.total_budget = data.count * data.budget_per_unit
    allocation.updated_by = current_user.id
    await db.commit()
    await db.refresh(allocation)
    return {"success": True, "message": "Allocation updated", "id": allocation.id}


@router.get("/allocations/check", response_model=AllocationCheckResponse)
async def check_allocation(
    house_id: int = Query(...),
    event_type_id: int = Query(...),
    thana: str = Query(...),
    month: str = Query(..., description="Month (YYYY-MM)"),
    exclude_event_id: Optional[int] = Query(None, description="Exclude this event ID from count (for edit)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
):
    try:
        target_month = datetime.strptime(month, "%Y-%m").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")

    alloc_result = await db.execute(
        select(ZoomInAllocation).options(
            joinedload(ZoomInAllocation.house),
            joinedload(ZoomInAllocation.event_type),
        ).where(
            ZoomInAllocation.house_id == house_id,
            ZoomInAllocation.month == target_month,
            ZoomInAllocation.event_type_id == event_type_id,
            ZoomInAllocation.thana == thana,
            ZoomInAllocation.is_deleted == False,
        )
    )
    allocation = alloc_result.scalar_one_or_none()

    if not allocation:
        return AllocationCheckResponse(
            has_allocation=False,
            allocated_count=0,
            created_count=0,
            remaining=0,
            month=month,
        )

    event_count_query = select(func.count(ZoomInEvent.id)).where(
        ZoomInEvent.house_id == house_id,
        func.date_trunc("month", ZoomInEvent.date) == target_month,
        ZoomInEvent.event_type_id == event_type_id,
        ZoomInEvent.thana == thana,
    )
    if exclude_event_id:
        event_count_query = event_count_query.where(ZoomInEvent.id != exclude_event_id)

    count_result = await db.execute(event_count_query)
    created_count = count_result.scalar() or 0

    remaining = max(0, allocation.count - created_count)

    return AllocationCheckResponse(
        has_allocation=True,
        allocated_count=allocation.count,
        created_count=created_count,
        remaining=remaining,
        month=month,
        house_name=allocation.house.name if allocation.house else None,
        event_type_name=allocation.event_type.name if allocation.event_type else None,
        thana=allocation.thana,
    )


@router.get("/dashboard/summary", response_model=DashboardSummaryResponse)
async def get_dashboard_summary(
    month: Optional[str] = Query(None, description="Month (YYYY-MM), defaults to current"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
    house_id: Optional[int] = Depends(get_house_context),
):
    if month:
        target_month = datetime.strptime(month, "%Y-%m").date()
    else:
        today = date.today()
        target_month = date(today.year, today.month, 1)

    next_month = date(target_month.year + target_month.month // 12, target_month.month % 12 + 1, 1)
    month_end = next_month - timedelta(days=1)

    user_house_ids = [h.id for h in current_user.houses]
    is_admin = is_admin_user(current_user)

    effective_house_id = house_id
    if not effective_house_id and not is_admin and user_house_ids:
        effective_house_id = user_house_ids[0]

    house_filter = []
    if effective_house_id:
        house_filter = [ZoomInEvent.house_id == effective_house_id]
    elif not is_admin:
        if user_house_ids:
            house_filter = [ZoomInEvent.house_id.in_(user_house_ids)]
        else:
            house_filter = [ZoomInEvent.house_id == -1]

    # ─── Total Events ───────────────────────────────────────────
    count_query = select(func.count(ZoomInEvent.id)).where(
        ZoomInEvent.date >= target_month,
        ZoomInEvent.date <= month_end,
        *house_filter,
    )
    total_result = await db.execute(count_query)
    total_events = total_result.scalar() or 0

    # ─── Total Activations ──────────────────────────────────────
    # Get all event IDs for the month → their retailer codes → count Activation/LiveActivation
    total_activations = 0
    if total_events > 0:
        events_query = select(ZoomInEvent.id, ZoomInEvent.date).where(
            ZoomInEvent.date >= target_month,
            ZoomInEvent.date <= month_end,
            *house_filter,
        )
        evt_result = await db.execute(events_query)
        month_events = evt_result.all()

        evt_ids = [row[0] for row in month_events]
        evt_dates = {row[0]: row[1] for row in month_events}

        if evt_ids:
            retailer_result = await db.execute(
                select(ZoomInEventRetailer.zoom_in_event_id, ZoomInEventRetailer.retailer_code).where(
                    ZoomInEventRetailer.zoom_in_event_id.in_(evt_ids),
                )
            )
            event_codes: dict[int, set[str]] = {}
            for row in retailer_result.all():
                event_codes.setdefault(row.zoom_in_event_id, set()).add(row.retailer_code)

            excl_result = await db.execute(select(ExcludedProductCode.product_code))
            excluded_codes = {row[0] for row in excl_result.all()}

            today = date.today()
            for eid, edate in evt_dates.items():
                codes = event_codes.get(eid, set())
                if not codes:
                    continue
                model = LiveActivation if edate == today else Activation
                q = select(func.count()).where(
                    model.retailer_code.in_(list(codes)),
                    model.activation_date == edate,
                )
                if excluded_codes:
                    q = q.where(~model.product_code.in_(list(excluded_codes)))
                cnt_result = await db.execute(q)
                total_activations += cnt_result.scalar() or 0

    # ─── Daily Event Counts ──────────────────────────────────────
    daily_query = select(
        ZoomInEvent.date,
        func.count(ZoomInEvent.id).label("cnt"),
    ).where(
        ZoomInEvent.date >= target_month,
        ZoomInEvent.date <= month_end,
        *house_filter,
    ).group_by(ZoomInEvent.date).order_by(ZoomInEvent.date)

    daily_result = await db.execute(daily_query)
    daily_map: dict[str, int] = {}
    for row in daily_result.all():
        daily_map[str(row.date)] = row.cnt

    daily_events = []
    current = target_month
    while current <= month_end:
        daily_events.append({"date": str(current), "count": daily_map.get(str(current), 0)})
        current += timedelta(days=1)

    # ─── Allocations ────────────────────────────────────────────
    alloc_query = select(ZoomInAllocation).options(
        joinedload(ZoomInAllocation.event_type),
    ).where(
        ZoomInAllocation.month == target_month,
        ZoomInAllocation.is_deleted == False,
    )
    if effective_house_id:
        alloc_query = alloc_query.where(ZoomInAllocation.house_id == effective_house_id)

    alloc_result = await db.execute(alloc_query)
    allocations = alloc_result.unique().scalars().all()

    total_allocated = sum(a.count for a in allocations)

    # ─── Events Per Event Type ──────────────────────────────────
    event_type_query = select(
        ZoomInEvent.event_type_id,
        func.count(ZoomInEvent.id).label("cnt"),
    ).where(
        ZoomInEvent.date >= target_month,
        ZoomInEvent.date <= month_end,
        *house_filter,
    ).group_by(ZoomInEvent.event_type_id)

    event_type_result = await db.execute(event_type_query)
    event_type_counts: dict[int, int] = {}
    for row in event_type_result.all():
        event_type_counts[row.event_type_id] = row.cnt

    # ─── Events Per (Event Type + Thana) ────────────────────────
    event_thana_query = select(
        ZoomInEvent.event_type_id,
        ZoomInEvent.thana,
        func.count(ZoomInEvent.id).label("cnt"),
    ).where(
        ZoomInEvent.date >= target_month,
        ZoomInEvent.date <= month_end,
        *house_filter,
    ).group_by(ZoomInEvent.event_type_id, ZoomInEvent.thana)

    et_result = await db.execute(event_thana_query)
    event_thana_counts: dict[tuple[int, str], int] = {}
    for row in et_result.all():
        event_thana_counts[(row.event_type_id, row.thana)] = row.cnt

    # ─── Build Breakdown (per allocation row, with thana) ───────
    event_type_breakdown = []
    for a in allocations:
        et_name = a.event_type.name if a.event_type else f"Type {a.event_type_id}"
        key = (a.event_type_id, a.thana)
        created = event_thana_counts.get(key, 0)
        remaining = max(0, a.count - created)
        event_type_breakdown.append({
            "event_type": et_name,
            "thana": a.thana,
            "allocated": a.count,
            "created": created,
            "remaining": remaining,
        })
    total_created_events_from_alloc = sum(b["created"] for b in event_type_breakdown)
    remaining_allocations = max(0, total_allocated - total_created_events_from_alloc)
    allocation_used_pct = round((total_created_events_from_alloc / total_allocated * 100) if total_allocated > 0 else 0, 1)

    return DashboardSummaryResponse(
        total_events=total_events,
        total_activations=total_activations,
        total_allocated=total_allocated,
        remaining_allocations=remaining_allocations,
        allocation_used_pct=allocation_used_pct,
        event_type_breakdown=event_type_breakdown,
        daily_events=daily_events,
    )


# ─── Filter Options ────────────────────────────────────────────

@router.get("/filter-options")
async def get_zoom_in_filter_options(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
):
    event_types_result = await db.execute(
        select(ZoomInEventType).where(ZoomInEventType.is_active == True).order_by(ZoomInEventType.name)
    )
    event_types = event_types_result.scalars().all()

    activities_result = await db.execute(
        select(ZoomInActivity).where(ZoomInActivity.is_active == True).order_by(ZoomInActivity.name)
    )
    activities = activities_result.scalars().all()

    return {
        "event_types": [{"id": et.id, "name": et.name, "name_bn": et.name_bn} for et in event_types],
        "activities": [{"id": a.id, "name": a.name, "name_bn": a.name_bn} for a in activities],
    }


# ─── Event Endpoints ────────────────────────────────────────────

@router.get("/events")
async def get_events(
    pagination: PaginationParams = Depends(),
    date_from: Optional[date] = Query(None, description="Filter start date (inclusive)"),
    date_to: Optional[date] = Query(None, description="Filter end date (inclusive)"),
    event_type_id: Optional[int] = Query(None, description="Filter by event type"),
    activity_id: Optional[int] = Query(None, description="Filter by activity"),
    thana: Optional[str] = Query(None, description="Filter by thana"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
    house_id: Optional[int] = Depends(get_house_context),
):
    base_query = select(ZoomInEvent).options(
        joinedload(ZoomInEvent.house),
        joinedload(ZoomInEvent.event_type),
        joinedload(ZoomInEvent.activity),
        joinedload(ZoomInEvent.bts_list),
        joinedload(ZoomInEvent.rsos),
        joinedload(ZoomInEvent.bps),
        joinedload(ZoomInEvent.retailers),
    )

    is_admin = is_admin_user(current_user)
    if house_id:
        base_query = base_query.where(ZoomInEvent.house_id == house_id)
    elif not is_admin:
        user_house_ids = [h.id for h in current_user.houses]
        if user_house_ids:
            base_query = base_query.where(ZoomInEvent.house_id.in_(user_house_ids))
        else:
            base_query = base_query.where(ZoomInEvent.house_id == -1)

    if date_from:
        base_query = base_query.where(ZoomInEvent.date >= date_from)
    if date_to:
        base_query = base_query.where(ZoomInEvent.date <= date_to)

    if event_type_id:
        base_query = base_query.where(ZoomInEvent.event_type_id == event_type_id)
    if activity_id:
        base_query = base_query.where(ZoomInEvent.activity_id == activity_id)
    if thana:
        base_query = base_query.where(ZoomInEvent.thana.ilike(f"%{thana}%"))

    if pagination.search:
        search_pattern = f"%{pagination.search}%"
        base_query = base_query.where(
            (ZoomInEvent.thana.ilike(search_pattern)) |
            (ZoomInEvent.date.cast(String).ilike(search_pattern))
        )

    count_query = select(func.count()).select_from(base_query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    sort_column = getattr(ZoomInEvent, pagination.sort_by, ZoomInEvent.id)
    order = sort_column.asc() if pagination.sort_order == "asc" else sort_column.desc()
    offset = (pagination.page - 1) * pagination.per_page
    result = await db.execute(
        base_query.offset(offset).limit(pagination.per_page).order_by(order)
    )
    events = result.unique().scalars().all()

    total_pages = max(1, -(-total // pagination.per_page))

    # ─── Batch Activation Counts ────────────────────────────────
    today = date.today()
    all_emp_ids = set()
    all_retailer_codes_by_event: dict[int, set[str]] = {}
    for e in events:
        event_date = e.date
        table = LiveActivation if event_date == today else Activation
        codes: set[str] = set()
        for r in e.rsos:
            all_emp_ids.add(r.employee_id)
        for b in e.bps:
            all_emp_ids.add(b.employee_id)
        for r in e.retailers:
            if r.retailer_code:
                codes.add(r.retailer_code)
        all_retailer_codes_by_event[e.id] = codes

    emp_code_map: dict[int, str] = {}
    if all_emp_ids:
        emp_result = await db.execute(
            select(Employee.id, Employee.assisted_retailer_code).where(
                Employee.id.in_(list(all_emp_ids)),
                Employee.assisted_retailer_code.isnot(None),
            )
        )
        for row in emp_result.all():
            if row.assisted_retailer_code:
                emp_code_map[row.id] = row.assisted_retailer_code

    for e in events:
        codes = all_retailer_codes_by_event[e.id]
        for r in e.rsos:
            code = emp_code_map.get(r.employee_id)
            if code:
                codes.add(code)
        for b in e.bps:
            code = emp_code_map.get(b.employee_id)
            if code:
                codes.add(code)
        all_retailer_codes_by_event[e.id] = codes

    excl_result = await db.execute(select(ExcludedProductCode.product_code))
    excluded_codes = {row[0] for row in excl_result.all()}

    all_activation_counts: dict[int, int] = {}
    today_events: list[ZoomInEvent] = []
    past_events: list[ZoomInEvent] = []
    for e in events:
        if e.date == today:
            today_events.append(e)
        else:
            past_events.append(e)

    async def _batch_counts(evts: list[ZoomInEvent], model):
        if not evts:
            return
        code_set: set[str] = set()
        for e in evts:
            code_set.update(all_retailer_codes_by_event.get(e.id, set()))
        if not code_set:
            return
        dates = {e.date for e in evts}
        query = select(
            model.retailer_code,
            model.activation_date,
            func.count().label("cnt")
        ).where(
            model.retailer_code.in_(list(code_set)),
            model.activation_date.in_(list(dates)),
        )
        if excluded_codes:
            query = query.where(~model.product_code.in_(list(excluded_codes)))
        query = query.group_by(model.retailer_code, model.activation_date)
        count_result = await db.execute(query)
        date_code_counts: dict[tuple[date, str], int] = {}
        for row in count_result.all():
            key = (row[1], row[0])
            date_code_counts[key] = row[2]
        for e in evts:
            total = 0
            for code in all_retailer_codes_by_event.get(e.id, set()):
                total += date_code_counts.get((e.date, code), 0)
            all_activation_counts[e.id] = total

    await _batch_counts(today_events, LiveActivation)
    await _batch_counts(past_events, Activation)

    data = []
    for e in events:
        data.append({
            "id": e.id,
            "house_id": e.house_id,
            "date": str(e.date),
            "event_type_id": e.event_type_id,
            "activity_id": e.activity_id,
            "thana": e.thana,
            "house_name": e.house.name if e.house else None,
            "house_code": e.house.code if e.house else None,
            "event_type_name": e.event_type.name if e.event_type else None,
            "activity_name": e.activity.name if e.activity else None,
            "activation_count": all_activation_counts.get(e.id, 0),
            "bts_ids": [b.bts_id for b in e.bts_list],
            "rso_ids": [r.employee_id for r in e.rsos],
            "bp_ids": [b.employee_id for b in e.bps],
            "retailer_codes": [r.retailer_code for r in e.retailers],
        })

    return PaginatedResponse(
        success=True,
        data=data,
        pagination=PaginationMeta(
            page=pagination.page,
            per_page=pagination.per_page,
            total=total,
            total_pages=total_pages,
            has_next=pagination.page < total_pages,
            has_prev=pagination.page > 1,
        ),
    )


@router.get("/events/export")
async def export_events(
    date_from: Optional[date] = Query(None, description="Filter start date (inclusive)"),
    date_to: Optional[date] = Query(None, description="Filter end date (inclusive)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
    house_id: Optional[int] = Depends(get_house_context),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    base_query = select(ZoomInEvent).options(
        joinedload(ZoomInEvent.house),
        joinedload(ZoomInEvent.event_type),
        joinedload(ZoomInEvent.activity),
        joinedload(ZoomInEvent.bts_list),
        joinedload(ZoomInEvent.rsos),
        joinedload(ZoomInEvent.bps),
        joinedload(ZoomInEvent.retailers),
    )

    is_admin = is_admin_user(current_user)
    if house_id:
        base_query = base_query.where(ZoomInEvent.house_id == house_id)
    elif not is_admin:
        user_house_ids = [h.id for h in current_user.houses]
        if user_house_ids:
            base_query = base_query.where(ZoomInEvent.house_id.in_(user_house_ids))
        else:
            base_query = base_query.where(ZoomInEvent.house_id == -1)

    if date_from:
        base_query = base_query.where(ZoomInEvent.date >= date_from)
    if date_to:
        base_query = base_query.where(ZoomInEvent.date <= date_to)

    base_query = base_query.order_by(ZoomInEvent.date.desc())

    result = await db.execute(base_query)
    events = result.unique().scalars().all()

    # ─── Batch fetch BTS details ───
    all_bts_ids = set()
    for e in events:
        for b in e.bts_list:
            all_bts_ids.add(b.bts_id)
    bts_map: dict[int, tuple] = {}
    if all_bts_ids:
        bts_result = await db.execute(
            select(BTS).where(BTS.id.in_(list(all_bts_ids)))
        )
        for bts in bts_result.scalars().all():
            bts_map[bts.id] = bts

    # ─── Batch fetch employee assisted codes ───
    all_emp_ids = set()
    for e in events:
        for r in e.rsos:
            all_emp_ids.add(r.employee_id)
        for b in e.bps:
            all_emp_ids.add(b.employee_id)
    emp_map: dict[int, dict] = {}
    if all_emp_ids:
        emp_result = await db.execute(
            select(Employee).where(Employee.id.in_(list(all_emp_ids)))
        )
        for emp in emp_result.scalars().all():
            emp_map[emp.id] = {
                "assisted_retailer_code": emp.assisted_retailer_code,
                "dms_code": emp.dms_code,
            }

    today = date.today()

    # ─── Batch Activation Counts (GA) ────────────────────────────
    all_ga_code_event: dict[int, set[str]] = {}
    all_ga_emp_ids = set()
    for e in events:
        codes: set[str] = set()
        for r in e.rsos:
            all_ga_emp_ids.add(r.employee_id)
        for b in e.bps:
            all_ga_emp_ids.add(b.employee_id)
        for r in e.retailers:
            if r.retailer_code:
                codes.add(r.retailer_code)
        all_ga_code_event[e.id] = codes

    ga_emp_code_map: dict[int, str] = {}
    if all_ga_emp_ids:
        ga_emp_result = await db.execute(
            select(Employee.id, Employee.assisted_retailer_code).where(
                Employee.id.in_(list(all_ga_emp_ids)),
                Employee.assisted_retailer_code.isnot(None),
            )
        )
        for row in ga_emp_result.all():
            if row.assisted_retailer_code:
                ga_emp_code_map[row.id] = row.assisted_retailer_code

    for e in events:
        codes = all_ga_code_event[e.id]
        for r in e.rsos:
            code = ga_emp_code_map.get(r.employee_id)
            if code:
                codes.add(code)
        for b in e.bps:
            code = ga_emp_code_map.get(b.employee_id)
            if code:
                codes.add(code)
        all_ga_code_event[e.id] = codes

    excl_result = await db.execute(select(ExcludedProductCode.product_code))
    excluded_codes = {row[0] for row in excl_result.all()}

    ga_counts: dict[int, int] = {}
    today_events: list[ZoomInEvent] = []
    past_events: list[ZoomInEvent] = []
    for e in events:
        if e.date == today:
            today_events.append(e)
        else:
            past_events.append(e)

    async def _batch_ga_counts(evts: list[ZoomInEvent], model):
        if not evts:
            return
        code_set: set[str] = set()
        for e in evts:
            code_set.update(all_ga_code_event.get(e.id, set()))
        if not code_set:
            return
        dates = {e.date for e in evts}
        query = select(
            model.retailer_code,
            model.activation_date,
            func.count().label("cnt")
        ).where(
            model.retailer_code.in_(list(code_set)),
            model.activation_date.in_(list(dates)),
        )
        if excluded_codes:
            query = query.where(~model.product_code.in_(list(excluded_codes)))
        query = query.group_by(model.retailer_code, model.activation_date)
        count_result = await db.execute(query)
        date_code_counts: dict[tuple[date, str], int] = {}
        for row in count_result.all():
            key = (row[1], row[0])
            date_code_counts[key] = row[2]
        for e in evts:
            total = 0
            for code in all_ga_code_event.get(e.id, set()):
                total += date_code_counts.get((e.date, code), 0)
            ga_counts[e.id] = total

    await _batch_ga_counts(today_events, LiveActivation)
    await _batch_ga_counts(past_events, Activation)

    # ─── Build workbook ───
    wb = Workbook()
    ws = wb.active
    ws.title = "Zoom In Events"

    headers = [
        "Cluster", "Region", "DD Code", "Thana",
        "BTS Code 1", "BTS Code 2", "BTS Code 3", "BTS Code 4", "BTS Code 5",
        "Event Location Address",
        "Event Type", "Activity Selection (Only for Zoom IN)",
        "Activity Date (MM-DD-YYYY)",
        "RSO Assisted Code 1", "RSO Assisted Code 2", "RSO Assisted Code 3", "RSO Assisted Code 4", "RSO Assisted Code 5",
        "BP Assisted Code 1", "BP Assisted Code 2", "BP Assisted Code 3", "BP Assisted Code 4",
        "SSO Code 1", "SSO Code 2", "SSO Code 3", "SSO Code 4",
        "GA",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    cell_font = Font(size=10)

    for row_idx, e in enumerate(events, 2):
        bts_codes: list[str] = []
        addresses: list[str] = []
        clusters: set[str] = set()
        regions: set[str] = set()
        for b in e.bts_list:
            bts = bts_map.get(b.bts_id)
            if bts:
                if bts.bts_code:
                    bts_codes.append(bts.bts_code)
                if bts.short_address:
                    addresses.append(bts.short_address)
                if bts.cluster:
                    clusters.add(bts.cluster)
                if bts.region:
                    regions.add(bts.region)

        rso_codes: list[str] = []
        for r in e.rsos:
            emp = emp_map.get(r.employee_id)
            code = emp["assisted_retailer_code"] if emp and emp.get("assisted_retailer_code") else (emp["dms_code"] if emp else "")
            if code:
                rso_codes.append(code)

        bp_codes: list[str] = []
        for b in e.bps:
            emp = emp_map.get(b.employee_id)
            code = emp["assisted_retailer_code"] if emp and emp.get("assisted_retailer_code") else (emp["dms_code"] if emp else "")
            if code:
                bp_codes.append(code)

        sso_codes: list[str] = [r.retailer_code for r in e.retailers if r.retailer_code]

        def _pad(lst: list[str], n: int) -> list[str]:
            return (lst + [""] * n)[:n]

        row_data = [
            ", ".join(sorted(clusters)) if clusters else "",
            ", ".join(sorted(regions)) if regions else "",
            e.house.code if e.house else "",
            e.thana or "",
            *_pad(bts_codes, 5),
            ", ".join(addresses) if addresses else "",
            e.event_type.name if e.event_type else "",
            e.activity.name if e.activity else "",
            e.date.strftime("%m-%d-%Y") if e.date else "",
            *_pad(rso_codes, 5),
            *_pad(bp_codes, 4),
            *_pad(sso_codes, 4),
            ga_counts.get(e.id, 0),
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border

    # ─── Column widths ───
    from openpyxl.utils import get_column_letter
    col_widths = [18, 18, 12, 18, 14, 14, 14, 14, 14, 30, 18, 28, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=zoom_in_events_{date.today().strftime('%Y%m%d')}.xlsx"}
    )


@router.post("/events")
async def create_event(
    data: ZoomInEventCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.create")),
):
    # ─── Allocation Check ────────────────────────────────────────
    target_month = data.date.replace(day=1)
    alloc_result = await db.execute(
        select(ZoomInAllocation).where(
            ZoomInAllocation.house_id == data.house_id,
            ZoomInAllocation.month == target_month,
            ZoomInAllocation.event_type_id == data.event_type_id,
            ZoomInAllocation.thana == data.thana,
            ZoomInAllocation.is_deleted == False,
        )
    )
    allocation = alloc_result.scalar_one_or_none()

    if not allocation:
        raise HTTPException(
            status_code=400,
            detail=f"No allocation found for {data.thana} thana in {target_month.strftime('%B %Y')} with this event type. Please configure an allocation first."
        )

    event_count_query = select(func.count(ZoomInEvent.id)).where(
        ZoomInEvent.house_id == data.house_id,
        func.date_trunc("month", ZoomInEvent.date) == target_month,
        ZoomInEvent.event_type_id == data.event_type_id,
        ZoomInEvent.thana == data.thana,
    )
    count_result = await db.execute(event_count_query)
    created_count = count_result.scalar() or 0

    if created_count >= allocation.count:
        raise HTTPException(
            status_code=400,
            detail=f"Allocation limit reached for {data.thana} thana in {target_month.strftime('%B %Y')}. "
                   f"Allocated: {allocation.count}, Already created: {created_count}, Remaining: 0."
        )

    event = ZoomInEvent(
        house_id=data.house_id,
        date=data.date,
        event_type_id=data.event_type_id,
        activity_id=data.activity_id,
        thana=data.thana,
        created_by=current_user.id,
    )
    db.add(event)
    await db.flush()

    for bts_id in data.bts_ids:
        db.add(ZoomInEventBTS(zoom_in_event_id=event.id, bts_id=bts_id))
    for rso_id in data.rso_ids:
        db.add(ZoomInEventRSO(zoom_in_event_id=event.id, employee_id=rso_id))
    for bp_id in data.bp_ids:
        db.add(ZoomInEventBP(zoom_in_event_id=event.id, employee_id=bp_id))
    for code in data.retailer_codes:
        db.add(ZoomInEventRetailer(zoom_in_event_id=event.id, retailer_code=code))

    await db.commit()
    await db.refresh(event)
    return {"success": True, "message": "Event created", "id": event.id}


@router.get("/events/{event_id}")
async def get_event(
    event_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
):
    result = await db.execute(
        select(ZoomInEvent).options(
            joinedload(ZoomInEvent.house),
            joinedload(ZoomInEvent.event_type),
            joinedload(ZoomInEvent.activity),
            selectinload(ZoomInEvent.bts_list),
            selectinload(ZoomInEvent.rsos),
            selectinload(ZoomInEvent.bps),
            selectinload(ZoomInEvent.retailers),
        ).where(ZoomInEvent.id == event_id)
    )
    event = result.unique().scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    bts_ids = [b.bts_id for b in event.bts_list]
    rso_employee_ids = [r.employee_id for r in event.rsos]
    bp_employee_ids = [b.employee_id for b in event.bps]
    retailer_codes_list = [r.retailer_code for r in event.retailers]

    # ─── Activation Count Setup ──────────────────────────────────
    today = date.today()
    event_date = event.date
    ActivationModel = LiveActivation if event_date == today else Activation

    excl_result = await db.execute(select(ExcludedProductCode.product_code))
    excluded_codes = {row[0] for row in excl_result.all()}

    rso_emp_codes: dict[int, str | None] = {}
    if rso_employee_ids:
        emp_result = await db.execute(
            select(Employee.id, Employee.assisted_retailer_code).where(
                Employee.id.in_(rso_employee_ids),
            )
        )
        for row in emp_result.all():
            if row.assisted_retailer_code:
                rso_emp_codes[row.id] = row.assisted_retailer_code

    bp_emp_codes: dict[int, str | None] = {}
    if bp_employee_ids:
        emp_result = await db.execute(
            select(Employee.id, Employee.assisted_retailer_code).where(
                Employee.id.in_(bp_employee_ids),
            )
        )
        for row in emp_result.all():
            if row.assisted_retailer_code:
                bp_emp_codes[row.id] = row.assisted_retailer_code

    retailer_codes_set = {r.retailer_code for r in event.retailers if r.retailer_code}

    all_retailer_codes = list(
        set(rso_emp_codes.values()) | set(bp_emp_codes.values()) | retailer_codes_set
    )
    retailer_code_counts: dict[str, int] = {}
    if all_retailer_codes:
        query = select(
            ActivationModel.retailer_code,
            func.count().label("cnt")
        ).where(
            ActivationModel.activation_date == event_date,
            ActivationModel.retailer_code.in_(all_retailer_codes),
        )
        if excluded_codes:
            query = query.where(~ActivationModel.product_code.in_(list(excluded_codes)))
        query = query.group_by(ActivationModel.retailer_code)
        count_result = await db.execute(query)
        for row in count_result.all():
            retailer_code_counts[row[0]] = row[1]

    bts_details = []
    if bts_ids:
        bts_result = await db.execute(
            select(BTS).where(BTS.id.in_(bts_ids))
        )
        bts_rows = bts_result.scalars().all()
        bts_map = {b.id: b for b in bts_rows}
        for bid in bts_ids:
            b = bts_map.get(bid)
            bts_details.append({
                "id": bid,
                "bts_code": b.bts_code if b else None,
                "site_id": b.site_id if b else None,
                "address": b.short_address or b.address if b else None,
            })

    rso_details = []
    if rso_employee_ids:
        rso_result = await db.execute(
            select(Employee)
            .options(joinedload(Employee.user))
            .where(Employee.id.in_(rso_employee_ids))
        )
        rso_rows = rso_result.scalars().all()
        rso_map = {e.id: e for e in rso_rows}
        for eid in rso_employee_ids:
            e = rso_map.get(eid)
            code = rso_emp_codes.get(eid)
            rso_details.append({
                "id": eid,
                "dms_code": e.dms_code if e else None,
                "itop_number": e.itop_number if e else None,
                "name": e.user.name if e and e.user else None,
                "assisted_retailer_code": code,
                "activation_count": retailer_code_counts.get(code, 0) if code else 0,
            })

    bp_details = []
    if bp_employee_ids:
        bp_result = await db.execute(
            select(Employee)
            .options(joinedload(Employee.user))
            .where(Employee.id.in_(bp_employee_ids))
        )
        bp_rows = bp_result.scalars().all()
        bp_map = {e.id: e for e in bp_rows}
        for eid in bp_employee_ids:
            e = bp_map.get(eid)
            code = bp_emp_codes.get(eid)
            bp_details.append({
                "id": eid,
                "dms_code": e.dms_code if e else None,
                "pool_number": e.pool_number if e else None,
                "name": e.user.name if e and e.user else None,
                "assisted_retailer_code": code,
                "activation_count": retailer_code_counts.get(code, 0) if code else 0,
            })

    retailer_details = []
    if retailer_codes_list:
        ret_result = await db.execute(
            select(Retailer)
            .options(joinedload(Retailer.employee).joinedload(Employee.user))
            .where(Retailer.retailer_code.in_(retailer_codes_list))
        )
        ret_rows = ret_result.scalars().all()
        ret_map = {r.retailer_code: r for r in ret_rows}
        for code in retailer_codes_list:
            r = ret_map.get(code)
            emp = r.employee if r else None
            retailer_details.append({
                "retailer_code": code,
                "name": r.name if r else None,
                "itop_number": r.itop_number if r else None,
                "employee_name": emp.user.name if emp and emp.user else None,
                "employee_itop_number": emp.itop_number if emp else None,
                "activation_count": retailer_code_counts.get(code, 0) if code else 0,
            })

    # ─── Section Totals ─────────────────────────────────────────
    rso_total = sum(
        retailer_code_counts.get(c, 0)
        for c in rso_emp_codes.values() if c
    )
    bp_total = sum(
        retailer_code_counts.get(c, 0)
        for c in bp_emp_codes.values() if c
    )
    retailer_total = sum(
        retailer_code_counts.get(c, 0)
        for c in retailer_codes_set
    )

    activation_count = rso_total + bp_total + retailer_total

    return {
        "activation_count": activation_count,
        "rso_total_activation_count": rso_total,
        "bp_total_activation_count": bp_total,
        "retailer_total_activation_count": retailer_total,
        "id": event.id,
        "house_id": event.house_id,
        "date": str(event.date),
        "event_type_id": event.event_type_id,
        "activity_id": event.activity_id,
        "thana": event.thana,
        "house_name": event.house.name if event.house else None,
        "house_code": event.house.code if event.house else None,
        "event_type_name": event.event_type.name if event.event_type else None,
        "activity_name": event.activity.name if event.activity else None,
        "created_at": str(event.created_at) if event.created_at else None,
        "bts_details": bts_details,
        "rso_details": rso_details,
        "bp_details": bp_details,
        "retailer_details": retailer_details,
    }


@router.put("/events/{event_id}")
async def update_event(
    event_id: int,
    data: ZoomInEventUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.edit")),
):
    result = await db.execute(
        select(ZoomInEvent).where(ZoomInEvent.id == event_id)
    )
    event = result.unique().scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    new_date = data.date if data.date is not None else event.date
    new_event_type_id = data.event_type_id if data.event_type_id is not None else event.event_type_id
    new_thana = data.thana if data.thana is not None else event.thana

    # ─── Allocation Check on update ──────────────────────────────
    target_month = new_date.replace(day=1)
    alloc_result = await db.execute(
        select(ZoomInAllocation).where(
            ZoomInAllocation.house_id == event.house_id,
            ZoomInAllocation.month == target_month,
            ZoomInAllocation.event_type_id == new_event_type_id,
            ZoomInAllocation.thana == new_thana,
            ZoomInAllocation.is_deleted == False,
        )
    )
    allocation = alloc_result.scalar_one_or_none()

    if not allocation:
        raise HTTPException(
            status_code=400,
            detail=f"No allocation found for {new_thana} thana in {target_month.strftime('%B %Y')} with this event type. Please configure an allocation first."
        )

    event_count_query = select(func.count(ZoomInEvent.id)).where(
        ZoomInEvent.house_id == event.house_id,
        func.date_trunc("month", ZoomInEvent.date) == target_month,
        ZoomInEvent.event_type_id == new_event_type_id,
        ZoomInEvent.thana == new_thana,
        ZoomInEvent.id != event_id,
    )
    count_result = await db.execute(event_count_query)
    created_count = count_result.scalar() or 0

    if created_count >= allocation.count:
        raise HTTPException(
            status_code=400,
            detail=f"Allocation limit reached for {new_thana} thana in {target_month.strftime('%B %Y')}. "
                   f"Allocated: {allocation.count}, Already created: {created_count}, Remaining: 0."
        )

    if data.date is not None:
        event.date = data.date
    if data.event_type_id is not None:
        event.event_type_id = data.event_type_id
    if data.activity_id is not None:
        event.activity_id = data.activity_id
    if data.thana is not None:
        event.thana = data.thana
    event.updated_by = current_user.id

    if data.bts_ids is not None:
        await db.execute(
            sa_delete(ZoomInEventBTS).where(ZoomInEventBTS.zoom_in_event_id == event_id)
        )
        for bts_id in data.bts_ids:
            db.add(ZoomInEventBTS(zoom_in_event_id=event.id, bts_id=bts_id))

    if data.rso_ids is not None:
        await db.execute(
            sa_delete(ZoomInEventRSO).where(ZoomInEventRSO.zoom_in_event_id == event_id)
        )
        for rso_id in data.rso_ids:
            db.add(ZoomInEventRSO(zoom_in_event_id=event.id, employee_id=rso_id))

    if data.bp_ids is not None:
        await db.execute(
            sa_delete(ZoomInEventBP).where(ZoomInEventBP.zoom_in_event_id == event_id)
        )
        for bp_id in data.bp_ids:
            db.add(ZoomInEventBP(zoom_in_event_id=event.id, employee_id=bp_id))

    if data.retailer_codes is not None:
        await db.execute(
            sa_delete(ZoomInEventRetailer).where(ZoomInEventRetailer.zoom_in_event_id == event_id)
        )
        for code in data.retailer_codes:
            db.add(ZoomInEventRetailer(zoom_in_event_id=event.id, retailer_code=code))

    await db.commit()
    return {"success": True, "message": "Event updated"}


@router.delete("/events/{event_id}")
async def delete_event(
    event_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.delete")),
):
    result = await db.execute(
        select(ZoomInEvent).where(ZoomInEvent.id == event_id)
    )
    event = result.unique().scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    await db.delete(event)
    await db.commit()
    return {"success": True, "message": "Event deleted"}


# ─── Eligible BTS Endpoints ─────────────────────────────────────


async def resolve_house_context(
    house_id: Optional[int],
    current_user: User,
    db: AsyncSession,
) -> Optional[int]:
    if house_id:
        return house_id
    user_house_ids = [h.id for h in current_user.houses]
    if user_house_ids:
        return user_house_ids[0]
    if is_admin_user(current_user):
        result = await db.execute(select(House.id).limit(1))
        first_house = result.scalar_one_or_none()
        return first_house
    return None


@router.get("/eligible-bts")
async def get_eligible_bts(
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
    house_id: Optional[int] = Depends(get_house_context),
):
    house_ids: list[int] = []
    if house_id:
        house_ids = [house_id]
    else:
        user_hids = [h.id for h in current_user.houses]
        if user_hids:
            house_ids = user_hids
        elif is_admin_user(current_user):
            result = await db.execute(select(House.id))
            house_ids = [row.id for row in result.all()]

    if not house_ids:
        return []

    query = select(MelaEligibleBTS).options(
        joinedload(MelaEligibleBTS.bts),
        joinedload(MelaEligibleBTS.house),
    ).where(
        MelaEligibleBTS.house_id.in_(house_ids)
    )

    if search:
        search_pattern = f"%{search}%"
        query = query.join(BTS, MelaEligibleBTS.bts_id == BTS.id).where(
            (BTS.bts_code.ilike(search_pattern)) |
            (BTS.site_id.ilike(search_pattern)) |
            (BTS.thana.ilike(search_pattern)) |
            (BTS.short_address.ilike(search_pattern))
        )

    query = query.order_by(MelaEligibleBTS.id.desc())
    result = await db.execute(query)
    entries = result.unique().scalars().all()
    return [
        {
            "id": e.id,
            "house_id": e.house_id,
            "house_code": e.house.code if e.house else None,
            "bts_id": e.bts_id,
            "bts": {
                "id": e.bts.id,
                "site_id": e.bts.site_id,
                "bts_code": e.bts.bts_code,
                "site_type": e.bts.site_type,
                "thana": e.bts.thana,
                "thana_bn": e.bts.thana_bn,
                "district": e.bts.district,
                "district_bn": e.bts.district_bn,
                "division": e.bts.division,
                "division_bn": e.bts.division_bn,
                "cluster": e.bts.cluster,
                "cluster_bn": e.bts.cluster_bn,
                "region": e.bts.region,
                "region_bn": e.bts.region_bn,
                "network_mode": e.bts.network_mode,
                "address": e.bts.address,
                "address_bn": e.bts.address_bn,
                "short_address": e.bts.short_address,
                "short_address_bn": e.bts.short_address_bn,
                "longitude": e.bts.longitude,
                "latitude": e.bts.latitude,
                "archetype": e.bts.archetype,
                "market": e.bts.market,
                "distributor_code": e.bts.distributor_code,
                "onair_date_2g": e.bts.onair_date_2g,
                "onair_date_3g": e.bts.onair_date_3g,
                "onair_date_4g": e.bts.onair_date_4g,
                "urban_rural": e.bts.urban_rural,
                "priority": e.bts.priority,
            },
        }
        for e in entries
    ]


@router.get("/eligible-bts/sample")
async def download_eligible_bts_sample(
    current_user: User = Depends(has_permission("zoom_in.view")),
):
    excel_data = generate_eligible_bts_sample_bytes()
    return Response(
        content=excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=eligible_bts_sample.xlsx"}
    )


@router.post("/eligible-bts/import")
async def import_eligible_bts(
    file: UploadFile = File(...),
    current_user: User = Depends(has_permission("zoom_in.import")),
):
    if not os.path.exists("temp_downloads"):
        os.makedirs("temp_downloads")

    filename = file.filename or "upload.xlsx"
    if not validate_excel(filename):
        raise HTTPException(status_code=400, detail="Invalid file type. Only .xlsx and .xls files are allowed.")

    user_house_ids = [h.id for h in current_user.houses] if current_user.houses else None
    if is_admin_user(current_user):
        user_house_ids = None  # admins can import for any house

    file_path = f"temp_downloads/{safe_filename(filename)}"
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        async def progress(msg: str):
            logger.info(f"Eligible BTS Import: {msg}")

        count, error = await process_eligible_bts_excel(file_path, user_house_ids, progress)
        if error:
            raise HTTPException(status_code=400, detail=error)

        return {"message": f"Successfully imported {count} eligible BTS stations", "count": count}
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)


@router.delete("/eligible-bts")
async def clear_eligible_bts(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.delete")),
    house_id: Optional[int] = Depends(get_house_context),
):
    house_id = await resolve_house_context(house_id, current_user, db)
    if not house_id:
        raise HTTPException(status_code=400, detail="No house selected. Please select a house and try again.")

    await db.execute(
        sa_delete(MelaEligibleBTS).where(MelaEligibleBTS.house_id == house_id)
    )
    await db.commit()
    return {"success": True, "message": "Eligible BTS list cleared"}


@router.post("/eligible-bts/attach")
async def attach_eligible_bts(
    body: EligibleBTSAttach,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.create")),
):
    house_code = body.house_code.strip().upper()
    result = await db.execute(select(House.id).where(House.code == house_code))
    house = result.scalar_one_or_none()
    if not house:
        raise HTTPException(status_code=404, detail=f"House '{house_code}' not found")

    if not is_admin_user(current_user):
        user_hids = [h.id for h in current_user.houses]
        if house not in user_hids:
            raise HTTPException(status_code=403, detail="You do not have access to this house")

    bts_code = body.bts_code.strip().upper()
    result = await db.execute(select(BTS.id).where(BTS.bts_code == bts_code))
    bts = result.scalar_one_or_none()
    if not bts:
        raise HTTPException(status_code=404, detail=f"BTS '{bts_code}' not found")

    result = await db.execute(
        select(MelaEligibleBTS.id).where(
            MelaEligibleBTS.house_id == house,
            MelaEligibleBTS.bts_id == bts,
        )
    )
    if result.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="BTS already in eligible list")

    entry = MelaEligibleBTS(house_id=house, bts_id=bts)
    db.add(entry)
    await db.commit()
    return {"success": True, "message": f"BTS {bts_code} attached successfully"}


@router.delete("/eligible-bts/{entry_id}")
async def detach_eligible_bts(
    entry_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.delete")),
):
    result = await db.execute(
        select(MelaEligibleBTS).where(MelaEligibleBTS.id == entry_id)
    )
    entry = result.scalar_one_or_none()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    if not is_admin_user(current_user):
        user_hids = [h.id for h in current_user.houses]
        if entry.house_id not in user_hids:
            raise HTTPException(status_code=403, detail="Access denied")

    await db.delete(entry)
    await db.commit()
    return {"success": True, "message": "BTS removed from eligible list"}


@router.get("/eligible-bts/thana-options")
async def get_available_bts_thana_options(
    house_code: str = Query(..., description="House code to check against"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
):
    result = await db.execute(select(House.id).where(House.code == house_code.strip().upper()))
    house = result.scalar_one_or_none()
    if not house:
        raise HTTPException(status_code=404, detail=f"House '{house_code}' not found")

    if not is_admin_user(current_user):
        user_hids = [h.id for h in current_user.houses]
        if house not in user_hids:
            raise HTTPException(status_code=403, detail="Access denied")

    subq = select(MelaEligibleBTS.bts_id).where(MelaEligibleBTS.house_id == house).subquery()
    query = select(BTS.thana).where(
        BTS.id.notin_(subq),
        BTS.thana.isnot(None),
        BTS.thana != "",
    ).distinct().order_by(BTS.thana)
    result = await db.execute(query)
    return [row[0] for row in result.all()]


@router.get("/eligible-bts/available")
async def get_available_bts(
    search: Optional[str] = Query(None),
    thana: Optional[str] = Query(None),
    house_code: str = Query(..., description="House code to check against"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("zoom_in.view")),
):
    result = await db.execute(select(House.id).where(House.code == house_code.strip().upper()))
    house = result.scalar_one_or_none()
    if not house:
        raise HTTPException(status_code=404, detail=f"House '{house_code}' not found")

    if not is_admin_user(current_user):
        user_hids = [h.id for h in current_user.houses]
        if house not in user_hids:
            raise HTTPException(status_code=403, detail="Access denied")

    subq = select(MelaEligibleBTS.bts_id).where(MelaEligibleBTS.house_id == house).subquery()
    query = select(BTS).where(BTS.id.notin_(subq))
    if thana:
        query = query.where(BTS.thana == thana)
    if search:
        pattern = f"%{search}%"
        query = query.where(
            (BTS.bts_code.ilike(pattern)) |
            (BTS.site_id.ilike(pattern)) |
            (BTS.thana.ilike(pattern)) |
            (BTS.short_address.ilike(pattern))
        )
    query = query.order_by(BTS.bts_code).limit(50)
    result = await db.execute(query)
    return result.scalars().all()
