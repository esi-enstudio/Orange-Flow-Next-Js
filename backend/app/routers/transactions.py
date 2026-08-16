import logging
from datetime import date
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.routers.deps import get_db, has_permission, get_house_context
from app.models.user import User
from app.models.house import House
from app.services.transaction_report_service import TransactionReportService, parse_date
from app.schemas.pagination import PaginationParams

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api", tags=["transactions"])


async def _resolve_house(
    db: AsyncSession,
    current_user: User,
    q_house_id: Optional[int],
    header_house_id: Optional[int],
) -> Optional[int]:
    if q_house_id and q_house_id != header_house_id:
        from app.utils.access_control import is_admin_user

        if not is_admin_user(current_user):
            user_house_ids = [h.id for h in current_user.houses]
            if q_house_id not in user_house_ids:
                raise HTTPException(status_code=403, detail="You do not have access to this house")
    target_house_id = q_house_id or header_house_id
    if not target_house_id:
        user_house_ids = [h.id for h in current_user.houses]
        if user_house_ids:
            target_house_id = user_house_ids[0]
    if not target_house_id:
        house_res = await db.execute(select(House.id).limit(1))
        target_house_id = house_res.scalar_one_or_none()
    return target_house_id


def _build_service(
    db: AsyncSession,
    target_house_id: Optional[int],
    report_type: Optional[str],
    start_date: Optional[date],
    end_date: Optional[date],
    rso_id: Optional[int],
    retailer_id: Optional[int],
) -> TransactionReportService:
    if start_date and end_date and start_date > end_date:
        raise HTTPException(status_code=400, detail="start_date cannot be after end_date")
    return TransactionReportService(
        db,
        target_house_id,
        report_type=report_type,
        start_date=start_date,
        end_date=end_date,
        rso_id=rso_id,
        retailer_id=retailer_id,
    )


@router.get("/reports/transactions")
async def get_transactions_report(
    pagination: PaginationParams = Depends(),
    report_type: Optional[str] = Query(None, pattern="^(C2C|C2S|Balance)$", description="Report type"),
    start_date: Optional[str] = Query(None, description="YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="YYYY-MM-DD"),
    rso_id: Optional[int] = Query(None, description="RSO employee ID"),
    retailer_id: Optional[int] = Query(None, description="Retailer ID"),
    q_house_id: Optional[int] = Query(None, alias="house_id"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("transactions.view")),
    header_house_id: Optional[int] = Depends(get_house_context),
):
    try:
        sd = parse_date(start_date, "start_date")
        ed = parse_date(end_date, "end_date")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    target_house_id = await _resolve_house(db, current_user, q_house_id, header_house_id)
    service = _build_service(db, target_house_id, report_type, sd, ed, rso_id, retailer_id)

    summary = await service.get_summary()
    trend = await service.get_daily_trend()
    groups, total = await service.get_daily_groups(pagination.page, pagination.per_page)

    total_pages = max(1, (total + pagination.per_page - 1) // pagination.per_page)
    return {
        "success": True,
        "house_id": target_house_id,
        "summary": summary,
        "trend": trend,
        "data": groups,
        "pagination": {
            "page": pagination.page,
            "per_page": pagination.per_page,
            "total": total,
            "total_pages": total_pages,
            "has_next": pagination.page < total_pages,
            "has_prev": pagination.page > 1,
        },
    }


@router.get("/reports/transactions/entities")
async def get_transaction_entities(
    entity_type: str = Query("rso", pattern="^(rso|retailer)$"),
    search: Optional[str] = Query(None),
    q_house_id: Optional[int] = Query(None, alias="house_id"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("transactions.view")),
    header_house_id: Optional[int] = Depends(get_house_context),
):
    target_house_id = await _resolve_house(db, current_user, q_house_id, header_house_id)
    service = TransactionReportService(db, target_house_id)
    return {"success": True, "data": await service.get_entities(entity_type, search)}


@router.get("/reports/transactions/export")
async def export_transactions_report(
    report_type: Optional[str] = Query(None, pattern="^(C2C|C2S|Balance)$"),
    start_date: Optional[str] = Query(None, description="YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="YYYY-MM-DD"),
    rso_id: Optional[int] = Query(None),
    retailer_id: Optional[int] = Query(None),
    q_house_id: Optional[int] = Query(None, alias="house_id"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(has_permission("transactions.export")),
    header_house_id: Optional[int] = Depends(get_house_context),
):
    try:
        sd = parse_date(start_date, "start_date")
        ed = parse_date(end_date, "end_date")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    target_house_id = await _resolve_house(db, current_user, q_house_id, header_house_id)
    service = _build_service(db, target_house_id, report_type, sd, ed, rso_id, retailer_id)

    summary = await service.get_summary()
    rows = await service.get_export_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    label = f"Transaction Report ({report_type or 'All Types'})"
    fmt_d = lambda d: d.strftime("%d %b %Y") if d else "..."
    date_label = f"{fmt_d(sd)} to {fmt_d(ed)}"
    ws.cell(row=1, column=1, value=f"{label} - {date_label}").font = Font(bold=True, size=14)
    ws.merge_cells("A1:J1")

    ws.cell(row=3, column=1, value="Summary").font = Font(bold=True, size=12)
    summary_headers = ["Metric", "Value"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    summary_rows = [
        ("Total Value (BDT)", summary["total_value"]),
        ("Total Transactions", summary["total_records"]),
        ("Active Days", summary["active_days"]),
        ("Active Retailers", summary["active_retailers"]),
        ("Daily Average", summary["daily_average"]),
    ]
    for i, (label_txt, val) in enumerate(summary_rows):
        ws.cell(row=5 + i, column=1, value=label_txt).border = thin_border
        ws.cell(row=5 + i, column=2, value=val).border = thin_border

    ws.cell(row=12, column=1, value="Transactions").font = Font(bold=True, size=12)
    cols = ["House Code", "Date", "RSO Name", "DMS Code", "RSO Itopup Number", "Retailer Code", "Retailer Itopup Number", "Retailer Name", "Amount", "Report Type"]
    for col, h in enumerate(cols, 1):
        cell = ws.cell(row=13, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, r in enumerate(rows):
        row_idx = 14 + i
        ws.cell(row=row_idx, column=1, value=r["house_code"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=r["date"]).border = thin_border
        ws.cell(row=row_idx, column=3, value=r["rso_name"]).border = thin_border
        ws.cell(row=row_idx, column=4, value=r["rso_dms_code"]).border = thin_border
        ws.cell(row=row_idx, column=5, value=r["rso_itop_number"]).border = thin_border
        ws.cell(row=row_idx, column=6, value=r["retailer_code"]).border = thin_border
        ws.cell(row=row_idx, column=7, value=r["retailer_itop_number"]).border = thin_border
        ws.cell(row=row_idx, column=8, value=r["retailer_name"]).border = thin_border
        ws.cell(row=row_idx, column=9, value=r["amount"]).border = thin_border
        ws.cell(row=row_idx, column=10, value=r["report_type"]).border = thin_border

    for col_idx in range(1, ws.max_column + 1):
        col = ws[get_column_letter(col_idx)]
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=transactions_report.xlsx"},
    )