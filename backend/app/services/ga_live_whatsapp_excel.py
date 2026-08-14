"""Server-side GA Live Report Excel builder.

Ports the frontend ExcelJS layout (frontend/src/lib/export-ga-live-report.ts)
to openpyxl so the scheduled WhatsApp delivery can generate the identical file
without a browser.
"""
import io
import math
from datetime import date

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.formatting.rule import IconSetRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.properties import PageSetupProperties

from app.models.house import House
from app.services.ga_live_service import GaLiveQueryBuilder
from app.services.activation_report_service import ActivationReportService
from app.utils.activation_rules import get_excluded_codes
from app.utils.timezone import now_naive

# ── Style constants (mirror of frontend export) ──
SECTION_BG = "FF9FA8DA"
HEADER_BG = "FFC5CAE9"
TEXT_DARK = "FF1E293B"
TEXT_MUTED = "FF64748B"
WHITE = "FFFFFFFF"
MEDIUM_BG = "FF93C5FD"
LIGHT_BG = "FFDBEAFE"
MEDIUM_ORANGE = "FFFCD34D"
LIGHT_ORANGE = "FFFEF3C7"
GREEN = "FF10B981"
AMBER = "FFF59E0B"
RED = "FFEF4444"

THIN_BORDER = Border(
    left=Side(style="thin", color="FF000000"),
    right=Side(style="thin", color="FF000000"),
    top=Side(style="thin", color="FF000000"),
    bottom=Side(style="thin", color="FF000000"),
)

BODY_FONT = Font(name="Calibri", size=11, color=TEXT_DARK)
BOLD_FONT = Font(name="Calibri", bold=True, size=11, color=TEXT_DARK)
TITLE_FONT = Font(name="Calibri", bold=True, size=18, color=TEXT_DARK)
MUTED_FONT = Font(name="Calibri", bold=True, size=11, color=TEXT_MUTED)

SECTION_FILL = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")
HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
ORANGE_FILL = PatternFill(start_color=MEDIUM_ORANGE, end_color=MEDIUM_ORANGE, fill_type="solid")
LIGHT_ORANGE_FILL = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid")


def _fmt_num(n) -> str:
    try:
        return f"{int(n):,}"
    except (TypeError, ValueError):
        return "0"


def _apply_borders(ws, row, col_start, col_end):
    for ci in range(col_start, col_end + 1):
        ws.cell(row=row, column=ci).border = THIN_BORDER


def _section_title(ws, row, label, col_start, col_end):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=label)
    cell.font = BOLD_FONT
    cell.alignment = Alignment(vertical="center", horizontal="left")
    cell.fill = SECTION_FILL
    cell.border = THIN_BORDER
    _apply_borders(ws, row, col_start, col_end)


def _header_row(ws, row, headers, fill=HEADER_FILL, orange_cols=None):
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = BOLD_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.border = THIN_BORDER
        cell.fill = fill
    for ci in (orange_cols or []):
        ws.cell(row=row, column=ci).fill = ORANGE_FILL


def _data_row(ws, row, values):
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=ci, value=val)
        cell.font = BODY_FONT
        cell.alignment = Alignment(
            vertical="center",
            horizontal="left" if ci == 2 else "center",
        )
        cell.border = THIN_BORDER


def _formula_row(ws, row, formula_cols, data_start, data_end, label, total_cols):
    for ci in range(1, total_cols + 1):
        col_letter = chr(64 + ci)
        cell = ws.cell(row=row, column=ci)
        cell.font = BOLD_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.border = THIN_BORDER
        cell.fill = SECTION_FILL
        if ci == 1:
            cell.value = label
        elif col_letter in formula_cols:
            cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"


def _add_5arrows(ws, col_letter, data_start, data_end):
    if data_start > data_end:
        return
    ws.conditional_formatting.add(
        f"{col_letter}{data_start}:{col_letter}{data_end}",
        IconSetRule("5Arrows", "percent", [0, 20, 40, 60, 80], showValue=None),
    )


async def _load_monthly_summary(db: AsyncSession, house_id: int, today: date) -> dict:
    excluded_codes = await get_excluded_codes(db)
    service = ActivationReportService(
        db,
        house_id,
        today.month,
        today.year,
        exclude_tag_names=[],
        exclude_product_codes=excluded_codes,
    )
    return await service.get_summary()


async def build_ga_live_report_excel(
    db: AsyncSession,
    house_id: int,
) -> bytes:
    today = now_naive().date()

    house_res = await db.execute(select(House).where(House.id == house_id))
    house = house_res.scalar_one_or_none()
    if not house:
        raise ValueError(f"House {house_id} not found")
    house_name = house.name or ""
    house_code = house.code or ""

    builder = GaLiveQueryBuilder(db, house_id, today, today)
    data = await builder.build_all()
    summary = await _load_monthly_summary(db, house_id, today)

    summary_obj = {
        "monthly_target": summary.get("monthly_target", 0),
        "achievement": summary.get("achievement", 0),
        "achievement_percentage": summary.get("achievement_percentage", 0),
        "remaining": summary.get("remaining", 0),
        "daily_required": summary.get("daily_required", 0),
        "daily_required_with_friday": summary.get("daily_required_with_friday", 0),
        "days_remaining": summary.get("days_remaining", 0),
    }

    total_activations = data["summary"].get("total_activations", 0)
    rsos = data.get("rsos", [])
    bps = data.get("bps", [])
    ccs = data.get("ccs", [])
    supervisors = data.get("supervisors", [])

    month_year = today.strftime("%B %Y")
    date_str = today.strftime("%d %B %Y")
    time_str = now_naive().strftime("%I:%M %p").lstrip("0").replace(" 0", " ")

    wb = Workbook()
    ws = wb.active
    ws.title = "GA Live Report"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = 9
    ws.page_setup.margins = {
        "top": 0, "bottom": 0, "left": 0, "right": 0, "header": 0, "footer": 0,
    }

    widths = [6, 28, 20, 18] + [12] * 9
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # ══ Rows 1-2: Title + DD Target headers ══
    ws.merge_cells("A1:C2")
    title_cell = ws.cell(row=1, column=1, value=f"GA Live Report ({month_year})")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center", horizontal="left")

    row2 = ws.row_dimensions[2]
    row2.height = 24
    dd_headers = ["Target", "Ach", "%", "Remain", "DRR"]
    for i, h in enumerate(dd_headers):
        cell = ws.cell(row=2, column=5 + i, value=h)
        cell.font = BOLD_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.fill = SECTION_FILL
        cell.border = THIN_BORDER

    # ══ Row 3: DD values + house info ══
    row3 = ws.row_dimensions[3]
    row3.height = 22

    ach_pct = summary_obj["achievement_percentage"]
    pct_color = (
        MEDIUM_BG
        if ach_pct >= 100
        else GREEN if ach_pct >= 70 else AMBER if ach_pct >= 40 else RED
    )
    dd_values = [
        _fmt_num(summary_obj["monthly_target"]),
        _fmt_num(summary_obj["achievement"]),
        "",
        "",
        _fmt_num(summary_obj["daily_required_with_friday"]),
    ]
    for i, val in enumerate(dd_values):
        cell = ws.cell(row=3, column=5 + i, value=val)
        cell.font = Font(
            name="Calibri",
            size=11,
            bold=(i == 2),
            color=pct_color if i == 2 else TEXT_DARK,
        )
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.border = THIN_BORDER

    ws.cell(row=3, column=7).value = "=IF(E3>0, ROUND(F3/E3*100, 1), 0)"
    ws.cell(row=3, column=7).number_format = '0.0"%"'
    ws.cell(row=3, column=7).alignment = Alignment(vertical="center", horizontal="right")
    ws.cell(row=3, column=8).value = "=E3-F3"

    ws.merge_cells("K1:M4")
    total_cell = ws.cell(row=1, column=11, value=total_activations)
    total_cell.font = Font(name="Calibri", bold=True, size=48, color=TEXT_DARK)
    total_cell.alignment = Alignment(vertical="center", horizontal="center")
    total_cell.border = THIN_BORDER

    ws.merge_cells("A3:C3")
    house_cell = ws.cell(row=3, column=1, value=f"House: {house_name} ({house_code})")
    house_cell.font = MUTED_FONT
    house_cell.alignment = Alignment(vertical="center", horizontal="left")

    ws.merge_cells("A4:C4")
    gen_cell = ws.cell(row=4, column=1, value=f"Generated: {date_str}, {time_str}")
    gen_cell.font = MUTED_FONT
    gen_cell.alignment = Alignment(vertical="center", horizontal="left")

    r = 5
    days_remaining = summary_obj["days_remaining"]
    sorted_rsos = sorted(rsos, key=lambda x: str(x.get("itop_number") or ""))
    sorted_bps = sorted(bps, key=lambda x: str(x.get("pool_number") or ""))
    sorted_ccs = sorted(ccs, key=lambda x: str(x.get("name") or ""))

    # ══ RSO PERFORMANCE ══
    if sorted_rsos:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=13)
        rso_title = ws.cell(row=r, column=1, value="RSO PERFORMANCE")
        rso_title.font = BOLD_FONT
        rso_title.alignment = Alignment(vertical="center", horizontal="left")
        rso_title.fill = SECTION_FILL
        rso_title.border = THIN_BORDER
        _apply_borders(ws, r, 1, 10)
        yest_cell = ws.cell(row=r, column=11, value="Yesterday")
        yest_cell.font = BOLD_FONT
        yest_cell.alignment = Alignment(vertical="center", horizontal="center")
        yest_cell.fill = ORANGE_FILL
        yest_cell.border = THIN_BORDER
        _apply_borders(ws, r, 11, 13)
        r += 1

        _header_row(
            ws,
            r,
            [
                "#", "Name", "ITop Number", "Assisted Code", "Today Target",
                "Own Code", "Market", "Total", "%", "Remain",
                "Own Code", "Market", "Total",
            ],
            orange_cols=[11, 12, 13],
        )
        r += 1

        rso_data_start = r
        for i, item in enumerate(sorted_rsos):
            drr = math.ceil(item["remaining"] / max(days_remaining, 1)) if item["remaining"] > 0 else 0
            _data_row(
                ws,
                r,
                [
                    i + 1,
                    item.get("name", ""),
                    item.get("itop_number", "") or "",
                    item.get("assisted_code", "") or "",
                    drr,
                    item.get("own_activation", 0),
                    item.get("market_activation", 0),
                    item.get("total_activation", 0),
                    "",
                    "",
                    item.get("yesterday_own", 0),
                    item.get("yesterday_market", 0),
                    item.get("yesterday_total", 0),
                ],
            )
            ws.cell(row=r, column=9).value = f"=IF(E{r}>0, H{r}/E{r}, 0)"
            ws.cell(row=r, column=9).number_format = "0%"
            ws.cell(row=r, column=9).alignment = Alignment(vertical="center", horizontal="right")
            ws.cell(row=r, column=10).value = f"=MAX(0, E{r}-H{r})"
            for ci in (11, 12, 13):
                ws.cell(row=r, column=ci).fill = LIGHT_ORANGE_FILL
            r += 1
        rso_data_end = r - 1

        _add_5arrows(ws, "H", rso_data_start, rso_data_end)
        _add_5arrows(ws, "I", rso_data_start, rso_data_end)

        if len(sorted_rsos) > 1:
            _formula_row(ws, r, {"E", "F", "G", "H", "J", "K", "L", "M"}, rso_data_start, rso_data_end, "Total", 13)
            ws.cell(row=r, column=9).value = f"=IF(E{r}>0, H{r}/E{r}, 0)"
            ws.cell(row=r, column=9).number_format = "0%"
            ws.cell(row=r, column=9).alignment = Alignment(vertical="center", horizontal="right")
            for ci in (11, 12, 13):
                ws.cell(row=r, column=ci).fill = LIGHT_ORANGE_FILL
            r += 1
        r += 1  # spacer

    # ══ BP PERFORMANCE ══
    if sorted_bps:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        bp_title = ws.cell(row=r, column=1, value="BP PERFORMANCE")
        bp_title.font = BOLD_FONT
        bp_title.alignment = Alignment(vertical="center", horizontal="left")
        bp_title.fill = SECTION_FILL
        bp_title.border = THIN_BORDER
        _apply_borders(ws, r, 1, 9)
        r += 1

        _header_row(
            ws,
            r,
            ["#", "Name", "Pool Number", "Assisted Code", "Today Target", "Ach", "%", "Remain", "Yest GA"],
            orange_cols=[9],
        )
        r += 1

        bp_data_start = r
        for i, item in enumerate(sorted_bps):
            bp_today_target = math.ceil(item["remaining"] / max(days_remaining, 1)) if item["remaining"] > 0 else 0
            _data_row(
                ws,
                r,
                [
                    i + 1,
                    item.get("name", ""),
                    item.get("pool_number", "") or "",
                    item.get("assisted_code", "") or "",
                    bp_today_target,
                    item.get("own_activation", 0),
                    "",
                    "",
                    item.get("yesterday_activation", 0),
                ],
            )
            ws.cell(row=r, column=7).value = f"=IF(E{r}>0, F{r}/E{r}, 0)"
            ws.cell(row=r, column=7).number_format = "0%"
            ws.cell(row=r, column=7).alignment = Alignment(vertical="center", horizontal="right")
            ws.cell(row=r, column=8).value = f"=MAX(0, E{r}-F{r})"
            ws.cell(row=r, column=9).fill = LIGHT_ORANGE_FILL
            r += 1
        bp_data_end = r - 1

        _add_5arrows(ws, "F", bp_data_start, bp_data_end)
        _add_5arrows(ws, "G", bp_data_start, bp_data_end)

        if len(sorted_bps) > 1:
            _formula_row(ws, r, {"E", "F", "H", "I"}, bp_data_start, bp_data_end, "Total", 9)
            ws.cell(row=r, column=7).value = f"=IF(E{r}>0, F{r}/E{r}, 0)"
            ws.cell(row=r, column=7).number_format = "0%"
            ws.cell(row=r, column=7).alignment = Alignment(vertical="center", horizontal="right")
            ws.cell(row=r, column=9).fill = LIGHT_ORANGE_FILL
            r += 1
        r += 1  # spacer

    # ══ CC PERFORMANCE ══
    if sorted_ccs:
        _section_title(ws, r, "CC PERFORMANCE", 1, 8)
        r += 1
        _header_row(
            ws,
            r,
            ["#", "Name", "Assisted Code", "Pool Number", "Today GA", "Total GA", "Yesterday GA", "Day Count"],
        )
        r += 1

        cc_data_start = r
        for i, item in enumerate(sorted_ccs):
            _data_row(
                ws,
                r,
                [
                    i + 1,
                    item.get("name", ""),
                    item.get("assisted_code", "") or "",
                    item.get("pool_number", "") or "",
                    item.get("own_activation", 0),
                    item.get("total_ga", 0) or 0,
                    item.get("yesterday_activation", 0) or 0,
                    item.get("day_count", 0) or 0,
                ],
            )
            r += 1
        cc_data_end = r - 1

        if len(sorted_ccs) > 1:
            _formula_row(ws, r, {"E", "F", "G", "H"}, cc_data_start, cc_data_end, "Total", 8)
            r += 1
        r += 1  # spacer

    # ══ SUPERVISOR PERFORMANCE ══
    if supervisors:
        _section_title(ws, r, "SUPERVISOR PERFORMANCE", 1, 5)
        r += 1
        _header_row(
            ws,
            r,
            ["#", "Name", "Pool Number", "Today GA", "Yest GA"],
        )
        r += 1

        sup_data_start = r
        for i, item in enumerate(supervisors):
            _data_row(
                ws,
                r,
                [
                    i + 1,
                    item.get("name", ""),
                    item.get("pool_number", "") or "",
                    item.get("total_activation", 0),
                    item.get("yesterday_total", 0) or 0,
                ],
            )
            r += 1
        sup_data_end = r - 1

        if len(supervisors) > 1:
            _formula_row(ws, r, {"D", "E"}, sup_data_start, sup_data_end, "Total", 5)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
