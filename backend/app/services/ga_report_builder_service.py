"""GA Report Builder service.

Builds per-retailer activation reports for a configurable date window (event).

Data sourcing rules:
  * Today (when it falls inside the event window) is read from
    `LiveActivation` — but falls back to `Activations` when there is no live
    data for today yet.
  * All other days inside the window are read from `Activations`.

The report is fully configurable: retailers/RSOs, displayed columns (in a
chosen order), sort field/order, and product-code / retailer-tag exclusions.
"""
import io
import math
from datetime import date, timedelta
from typing import Any, Optional

from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.models.activation import Activation
from app.models.live_activation import LiveActivation
from app.models.retailer import Retailer
from app.models.employee import Employee
from app.models.user import User
from app.models.house import House
from app.models.ga_filter import FilterTag, RetailerFilter
from app.models.product_exclusion import ExcludedProductCode
from app.models.ga_report_target import GaReportTarget
from app.utils.activation_rules import exclude_clause
from app.utils.timezone import now_naive

logger = __import__("logging").getLogger("app.services.GaReportBuilder")

ACTIVATION_METRICS = ("activation_count", "today_activation", "yesterday_activation", "live_activation")

SLAB_METRIC_SUFFIXES = ("target", "achievement", "achievement_pct", "remaining")


def slab_column_keys(slabs: int) -> list[str]:
    keys: list[str] = []
    for s in range(1, slabs + 1):
        keys.extend([
            f"slab_{s}_target",
            f"slab_{s}_achievement",
            f"slab_{s}_achievement_pct",
            f"slab_{s}_remaining",
        ])
    return keys


def is_slab_column(key: str) -> bool:
    return key.startswith("slab_") and key.rsplit("_", 1)[-1] in SLAB_METRIC_SUFFIXES


def slab_column_label(key: str) -> str:
    try:
        _, num, metric = key.split("_", 2)
        return f"Slab {num} {metric.replace('_', ' ').title()}"
    except ValueError:
        return key

COLUMN_REGISTRY: list[dict[str, Any]] = [
    # house
    {"key": "house_code", "label": "House Code", "category": "house", "type": "string", "sortable": True},
    {"key": "house_name", "label": "House Name", "category": "house", "type": "string", "sortable": True},
    # rso
    {"key": "rso_name", "label": "RSO Name", "category": "rso", "type": "string", "sortable": True},
    {"key": "rso_itop_number", "label": "RSO iTopUp No", "category": "rso", "type": "string", "sortable": True},
    {"key": "rso_dms_code", "label": "RSO DMS Code", "category": "rso", "type": "string", "sortable": True},
    {"key": "rso_assisted_code", "label": "RSO Assisted Code", "category": "rso", "type": "string", "sortable": True},
    {"key": "rso_pool_number", "label": "RSO Pool No", "category": "rso", "type": "string", "sortable": True},
    # bp
    {"key": "bp_name", "label": "BP Name", "category": "bp", "type": "string", "sortable": True},
    {"key": "bp_pool_number", "label": "BP Pool No", "category": "bp", "type": "string", "sortable": True},
    {"key": "bp_assisted_code", "label": "BP Assisted Code", "category": "bp", "type": "string", "sortable": True},
    # retailer
    {"key": "retailer_code", "label": "Retailer Code", "category": "retailer", "type": "string", "sortable": True},
    {"key": "retailer_name", "label": "Retailer Name", "category": "retailer", "type": "string", "sortable": True},
    {"key": "retailer_itop_number", "label": "Retailer iTopUp No", "category": "retailer", "type": "string", "sortable": True},
    {"key": "retailer_type", "label": "Retailer Type", "category": "retailer", "type": "string", "sortable": True},
    {"key": "retailer_district", "label": "District", "category": "retailer", "type": "string", "sortable": True},
    {"key": "retailer_thana", "label": "Thana", "category": "retailer", "type": "string", "sortable": True},
    {"key": "retailer_address", "label": "Address", "category": "retailer", "type": "string", "sortable": False},
    {"key": "retailer_contact_no", "label": "Contact No", "category": "retailer", "type": "string", "sortable": False},
    # activation metrics
    {"key": "activation_count", "label": "Activations", "category": "activation", "type": "number", "sortable": True},
    {"key": "today_activation", "label": "Today", "category": "activation", "type": "number", "sortable": True},
    {"key": "yesterday_activation", "label": "Yesterday", "category": "activation", "type": "number", "sortable": True},
]

COLUMN_KEYS = {c["key"] for c in COLUMN_REGISTRY}
COLUMN_LABELS = {c["key"]: c["label"] for c in COLUMN_REGISTRY}
COLUMN_LABELS["live_activation"] = "Live Activation"
COLUMN_CATEGORIES = ["house", "rso", "retailer", "activation"]
DEFAULT_COLUMNS = [
    "house_code", "rso_name", "rso_itop_number",
    "retailer_code", "retailer_name", "retailer_itop_number", "activation_count",
]


def _fmt(n) -> str:
    try:
        return f"{int(n):,}"
    except (TypeError, ValueError):
        return "0"


def _is_numeric_value(v) -> bool:
    s = str(v).replace(",", "").replace(".", "").strip()
    return s.isdigit()


class ReportConfig:
    """Plain-object view of the builder config (router validates payload)."""

    def __init__(self, payload: dict):
        self.house_id = payload.get("house_id")
        self.event_id = payload.get("event_id")
        self.event_name = payload.get("event_name")
        self.start_date = self._parse_date(payload.get("start_date"))
        self.end_date = self._parse_date(payload.get("end_date"))
        self.retailer_codes: list[str] = payload.get("retailer_codes") or []
        self.rso_ids: list[int] = payload.get("rso_ids") or []
        self.bp_ids: list[int] = payload.get("bp_ids") or []
        self.target_type: Optional[str] = payload.get("target_type") or None  # rso | bp | retailer
        try:
            self.slabs: int = max(1, int(payload.get("slabs") or 1))
        except (TypeError, ValueError):
            self.slabs = 1
        raw_columns = payload.get("columns") or DEFAULT_COLUMNS
        # Accept either a list of keys or a list of {"key": ...} objects.
        self.columns: list[str] = [
            c["key"] if isinstance(c, dict) else str(c)
            for c in raw_columns
            if (c["key"] if isinstance(c, dict) else str(c)) in COLUMN_KEYS
        ]
        filters = payload.get("filters") or {}
        self.exclude_product_codes: list[str] = filters.get("exclude_product_codes") or []
        self.exclude_retailer_tags: list[str] = filters.get("exclude_retailer_tags") or []
        self.sort_by = payload.get("sort_by") or "activation_count"
        self.sort_order = payload.get("sort_order") or "desc"
        self.targets: list = payload.get("targets") or []

    @staticmethod
    def _parse_date(value):
        if value is None:
            return None
        if isinstance(value, date):
            return value
        try:
            return date.fromisoformat(str(value))
        except ValueError:
            return None


class GaReportBuilderService:

    def __init__(self, db: AsyncSession, config: ReportConfig):
        self.db = db
        self.cfg = config

    # ------------------------------------------------------------------ meta

    @staticmethod
    def column_options() -> list[dict]:
        return [dict(c) for c in COLUMN_REGISTRY]

    async def get_entities(self, entity_type: str, search: Optional[str], limit: int = 50) -> list[dict]:
        p = f"%{search}%" if search else None
        if entity_type == "retailer":
            query = select(Retailer).options(
                joinedload(Retailer.employee).joinedload(Employee.user)
            ).where(Retailer.house_id == self.cfg.house_id)
            if p:
                query = query.where(
                    or_(
                        Retailer.retailer_code.ilike(p),
                        Retailer.name.ilike(p),
                        Retailer.itop_number.ilike(p),
                    )
                )
            res = await self.db.execute(query.order_by(Retailer.name.asc()).limit(limit))
            retailers = res.scalars().all()
            return [
                {
                    "id": r.id,
                    "code": r.retailer_code,
                    "name": r.name,
                    "itop_number": r.itop_number or "",
                    "rso_name": r.employee.user.name if r.employee and r.employee.user else (r.employee.dms_code if r.employee else ""),
                }
                for r in retailers
            ]
        query = select(Employee).options(joinedload(Employee.user))
        if entity_type == "bp":
            query = query.where(Employee.employee_type == "bp", Employee.status == "Active")
        else:
            query = query.where(Employee.employee_type == "rso", Employee.status == "Active")
        if self.cfg.house_id:
            query = query.where(Employee.house_id == self.cfg.house_id)
        if p:
            query = query.where(
                or_(
                    Employee.dms_code.ilike(p),
                    Employee.itop_number.ilike(p),
                    Employee.pool_number.ilike(p),
                    Employee.user.has(User.name.ilike(p)),
                )
            )
        res = await self.db.execute(query.order_by(Employee.dms_code.asc()).limit(limit))
        employees = res.scalars().all()
        return [
            {
                "id": e.id,
                "code": e.dms_code or "",
                "name": e.user.name if e.user else (e.dms_code or f"#{e.id}"),
                "itop_number": e.itop_number or "",
                "pool_number": e.pool_number or "",
                "assisted_code": e.assisted_retailer_code or "",
            }
            for e in employees
        ]

    async def get_exclusion_options(self) -> dict:
        prod_res = await self.db.execute(select(ExcludedProductCode.product_code))
        codes: set[str] = {row[0] for row in prod_res.all()}
        tag_res = await self.db.execute(
            select(FilterTag.id, FilterTag.name)
            .where(FilterTag.house_id == self.cfg.house_id)
            .order_by(FilterTag.name)
        )
        return {
            "product_codes": sorted(codes),
            "retailer_tags": [{"id": t.id, "name": t.name} for t in tag_res.all()],
        }

    # ------------------------------------------------------------ data access

    async def _global_excluded_codes(self) -> set[str]:
        res = await self.db.execute(select(ExcludedProductCode.product_code))
        return {row[0] for row in res.all()}

    async def _excluded_retailer_ids(self) -> set[int]:
        if not self.cfg.exclude_retailer_tags:
            return set()
        res = await self.db.execute(
            select(RetailerFilter.retailer_id)
            .join(FilterTag, RetailerFilter.tag_id == FilterTag.id)
            .where(
                FilterTag.name.in_(self.cfg.exclude_retailer_tags),
                RetailerFilter.house_id == self.cfg.house_id,
            )
        )
        return {row[0] for row in res.all()}

    def _event_window(self) -> tuple[date, date]:
        start = self.cfg.start_date or date(2020, 1, 1)
        end = self.cfg.end_date or now_naive().date()
        return start, end

    def _excluded_codes(self) -> set[str]:
        return self.cfg.exclude_product_codes or set()

    async def _count_activations_by_retailer(
        self, model, start: date, end: date, excluded_retailer_ids: set[int]
    ) -> dict[int, int]:
        query = (
            select(model.retailer_id, func.count(model.id))
            .where(
                model.house_id == self.cfg.house_id,
                model.retailer_id != None,  # noqa: E711
                model.activation_date >= start,
                model.activation_date <= end,
            )
            .group_by(model.retailer_id)
        )
        codes = self._excluded_codes()
        clause = exclude_clause(model, codes)
        if clause is not None:
            query = query.where(clause)
        if excluded_retailer_ids:
            query = query.where(model.retailer_id.notin_(excluded_retailer_ids))
        res = await self.db.execute(query)
        return {row[0]: int(row[1]) for row in res.all()}

    async def _has_live_today(self, today: date) -> bool:
        res = await self.db.execute(
            select(func.count(LiveActivation.id)).where(
                LiveActivation.house_id == self.cfg.house_id,
                LiveActivation.activation_date == today,
            )
        )
        return (res.scalar() or 0) > 0

    async def _load_retailer_rows(self, excluded_retailer_ids: set[int]) -> list[Retailer]:
        query = select(Retailer).options(
            joinedload(Retailer.employee).joinedload(Employee.user)
        ).where(Retailer.house_id == self.cfg.house_id)
        if self.cfg.retailer_codes:
            query = query.where(Retailer.retailer_code.in_(self.cfg.retailer_codes))
        if self.cfg.rso_ids:
            query = query.where(Retailer.employee_id.in_(self.cfg.rso_ids))
        if excluded_retailer_ids:
            query = query.where(Retailer.id.notin_(excluded_retailer_ids))
        res = await self.db.execute(query)
        return res.scalars().all()

    async def _load_targets(self) -> dict:
        targets: dict = {}
        if self.cfg.event_id:
            res = await self.db.execute(
                select(GaReportTarget).where(
                    GaReportTarget.event_id == self.cfg.event_id,
                    GaReportTarget.house_id == self.cfg.house_id,
                    GaReportTarget.target_type == (self.cfg.target_type or ""),
                    GaReportTarget.is_deleted == False,  # noqa: E712
                )
            )
            for t in res.scalars().all():
                key = t.retailer_code if t.target_type == "retailer" else str(t.entity_id)
                targets[(key, t.slab)] = float(t.target_value or 0)

        for entry in self.cfg.targets:
            if not isinstance(entry, dict):
                continue
            if self.cfg.target_type == "retailer":
                key = str(entry.get("retailer_code") or "")
            else:
                entity_id = entry.get("entity_id")
                if entity_id is None:
                    continue
                key = str(entity_id)
            try:
                slab = int(entry.get("slab") or 1)
            except (TypeError, ValueError):
                slab = 1
            try:
                value = float(entry.get("target_value") or 0)
            except (TypeError, ValueError):
                value = 0
            targets[(key, slab)] = value
        return targets

    async def _activation_counts(
        self, today: date, start: date, end: date, excluded_retailer_ids: set[int]
    ) -> tuple[dict[int, int], dict[int, int], dict[int, int]]:
        history_end = min(today - timedelta(days=1), end)
        history_counts: dict[int, int] = {}
        yesterday_counts: dict[int, int] = {}
        if start <= history_end:
            history_counts = await self._count_activations_by_retailer(
                Activation, start, history_end, excluded_retailer_ids
            )
            yesterday = today - timedelta(days=1)
            if start <= yesterday <= history_end:
                yesterday_counts = await self._count_activations_by_retailer(
                    Activation, yesterday, yesterday, excluded_retailer_ids
                )
        today_counts: dict[int, int] = {}
        if start <= today <= end:
            use_live = await self._has_live_today(today)
            model = LiveActivation if use_live else Activation
            today_counts = await self._count_activations_by_retailer(
                model, today, today, excluded_retailer_ids
            )
        return history_counts, today_counts, yesterday_counts

    async def _count_activations_by_retailer_code(
        self, start: date, end: date, excluded_retailer_ids: set[int]
    ) -> dict[str, int]:
        today = now_naive().date()
        counts: dict[str, int] = {}

        async def _count_model(model, s, e):
            query = (
                select(model.retailer_code, func.count(model.id))
                .where(
                    model.house_id == self.cfg.house_id,
                    model.retailer_code != None,  # noqa: E711
                    model.activation_date >= s,
                    model.activation_date <= e,
                )
                .group_by(model.retailer_code)
            )
            clause = exclude_clause(model, self._excluded_codes())
            if clause is not None:
                query = query.where(clause)
            if excluded_retailer_ids:
                query = query.where(model.retailer_id.notin_(excluded_retailer_ids))
            res = await self.db.execute(query)
            for row in res.all():
                counts[row[0]] = counts.get(row[0], 0) + int(row[1])

        if start > end:
            return counts
        use_live = (start <= today <= end) and await self._has_live_today(today)
        if use_live:
            if start <= min(today - timedelta(days=1), end):
                await _count_model(Activation, start, min(today - timedelta(days=1), end))
            await _count_model(LiveActivation, today, end)
        else:
            await _count_model(Activation, start, end)
        return counts

    async def _count_live_by_retailer(self, excluded_retailer_ids: set[int]) -> dict[int, int]:
        return await self._count_activations_by_retailer(
            LiveActivation, now_naive().date(), now_naive().date(), excluded_retailer_ids
        )

    async def _count_live_by_retailer_code(self, excluded_retailer_ids: set[int]) -> dict[str, int]:
        query = (
            select(LiveActivation.retailer_code, func.count(LiveActivation.id))
            .where(
                LiveActivation.house_id == self.cfg.house_id,
                LiveActivation.retailer_code != None,  # noqa: E711
                LiveActivation.activation_date == now_naive().date(),
            )
            .group_by(LiveActivation.retailer_code)
        )
        clause = exclude_clause(LiveActivation, self._excluded_codes())
        if clause is not None:
            query = query.where(clause)
        if excluded_retailer_ids:
            query = query.where(LiveActivation.retailer_id.notin_(excluded_retailer_ids))
        res = await self.db.execute(query)
        return {row[0]: int(row[1]) for row in res.all()}

    @staticmethod
    def _apply_slab_columns(row: dict, achievement: float, targets: dict, entity_key: str, slabs: int) -> dict:
        for s in range(1, slabs + 1):
            target = targets.get((entity_key, s), 0)
            row[f"slab_{s}_target"] = target
            row[f"slab_{s}_achievement"] = achievement
            row[f"slab_{s}_achievement_pct"] = round((achievement / target * 100), 1) if target else 0
            row[f"slab_{s}_remaining"] = max(0, target - achievement)
        return row

    @staticmethod
    def _compute_totals(rows: list[dict], columns: list[str], slabs: int) -> dict:
        totals: dict = {}
        for key in columns:
            if key in ACTIVATION_METRICS:
                totals[key] = sum(float(r.get(key, 0) or 0) for r in rows)
            elif is_slab_column(key) and not key.endswith("_achievement_pct"):
                totals[key] = sum(float(r.get(key, 0) or 0) for r in rows)
        for s in range(1, slabs + 1):
            tgt = totals.get(f"slab_{s}_target", 0)
            ach = totals.get(f"slab_{s}_achievement", 0)
            totals[f"slab_{s}_achievement_pct"] = round((ach / tgt * 100), 1) if tgt else 0
        return totals

    # -------------------------------------------------------------- reporting

    async def build_report(self) -> dict:
        today = now_naive().date()
        start, end = self._event_window()
        window = self._window_info(today, start, end)

        if self.cfg.target_type == "rso":
            return await self._build_employee_report("rso", today, start, end, window)
        if self.cfg.target_type == "bp":
            return await self._build_employee_report("bp", today, start, end, window)
        return await self._build_retailer_report(today, start, end, window)

    async def _build_retailer_report(self, today: date, start: date, end: date, window: dict) -> dict:
        excluded_retailer_ids = await self._excluded_retailer_ids()
        retailers = await self._load_retailer_rows(excluded_retailer_ids)
        targets = await self._load_targets()
        columns = [c for c in self.cfg.columns] + slab_column_keys(self.cfg.slabs) + ["live_activation"]
        if not retailers:
            return {"columns": columns, "rows": [], "totals": {}, "window": window}

        history_counts, today_counts, yesterday_counts = await self._activation_counts(
            today, start, end, excluded_retailer_ids
        )
        live_counts = await self._count_live_by_retailer(excluded_retailer_ids)

        house_cache: dict[int, House] = {}
        if self.cfg.house_id:
            house_res = await self.db.execute(select(House).where(House.id == self.cfg.house_id))
            house = house_res.scalar_one_or_none()
            if house:
                house_cache[house.id] = house

        rows: list[dict] = []
        for r in retailers:
            count = history_counts.get(r.id, 0) + today_counts.get(r.id, 0)
            row: dict = {
                "retailer_code": r.retailer_code or "",
                "retailer_name": r.name or "",
                "retailer_itop_number": r.itop_number or "",
                "retailer_type": r.type or "",
                "retailer_district": r.district or "",
                "retailer_thana": r.thana or "",
                "retailer_address": r.address or "",
                "retailer_contact_no": r.contact_no or "",
                "rso_name": r.employee.user.name if r.employee and r.employee.user else (r.employee.dms_code if r.employee else ""),
                "rso_itop_number": r.employee.itop_number or "" if r.employee else "",
                "rso_dms_code": r.employee.dms_code or "" if r.employee else "",
                "rso_assisted_code": r.employee.assisted_retailer_code or "" if r.employee else "",
                "rso_pool_number": r.employee.pool_number or "" if r.employee else "",
                "activation_count": count,
                "today_activation": today_counts.get(r.id, 0),
                "yesterday_activation": yesterday_counts.get(r.id, 0),
                "live_activation": live_counts.get(r.id, 0),
            }
            self._apply_slab_columns(row, count, targets, r.retailer_code or "", self.cfg.slabs)
            if house_cache:
                h = house_cache.get(r.house_id)
                if h:
                    row["house_code"] = h.code or ""
                    row["house_name"] = h.name or ""
                else:
                    row["house_code"] = ""
                    row["house_name"] = ""
            rows.append(row)

        if self.cfg.sort_by in columns:
            reverse = self.cfg.sort_order != "asc"
            col_type = next((c["type"] for c in COLUMN_REGISTRY if c["key"] == self.cfg.sort_by), "string")
            if col_type == "number":
                rows.sort(key=lambda x: x.get(self.cfg.sort_by, 0) or 0, reverse=reverse)
            else:
                rows.sort(key=lambda x: str(x.get(self.cfg.sort_by, "") or ""), reverse=reverse)

        totals = self._compute_totals(rows, columns, self.cfg.slabs)
        return {"columns": columns, "rows": rows, "totals": totals, "window": window}

    async def _build_employee_report(self, entity_type: str, today: date, start: date, end: date, window: dict) -> dict:
        id_filter = self.cfg.rso_ids if entity_type == "rso" else self.cfg.bp_ids
        query = select(Employee).options(joinedload(Employee.user)).where(
            Employee.house_id == self.cfg.house_id,
            Employee.employee_type == entity_type,
            Employee.status == "Active",
        )
        if id_filter:
            query = query.where(Employee.id.in_(id_filter))
        employees = (await self.db.execute(query)).scalars().all()

        identity = (
            ["rso_name", "rso_dms_code", "rso_itop_number"]
            if entity_type == "rso"
            else ["bp_name", "bp_pool_number", "bp_assisted_code"]
        )
        columns = identity + slab_column_keys(self.cfg.slabs) + ["live_activation"]
        if not employees:
            return {"columns": columns, "rows": [], "totals": {}, "window": window}

        excluded_retailer_ids = await self._excluded_retailer_ids()
        targets = await self._load_targets()
        history_counts, today_counts, yesterday_counts = await self._activation_counts(
            today, start, end, excluded_retailer_ids
        )
        live_counts = await self._count_live_by_retailer(excluded_retailer_ids)
        live_code_counts = await self._count_live_by_retailer_code(excluded_retailer_ids)

        achievement_map: dict[int, float] = {}
        live_map: dict[int, int] = {}
        if entity_type == "rso":
            emp_ids = [e.id for e in employees]
            retailer_rows = (
                await self.db.execute(
                    select(Retailer.id, Retailer.employee_id).where(
                        Retailer.house_id == self.cfg.house_id,
                        Retailer.employee_id.in_(emp_ids),
                        Retailer.employee_id.isnot(None),
                    )
                )
            ).all()
            emp_retailer_ids: dict[int, set[int]] = {}
            for rid, eid in retailer_rows:
                emp_retailer_ids.setdefault(eid, set()).add(rid)
            for emp in employees:
                rid_set = emp_retailer_ids.get(emp.id, set())
                achievement_map[emp.id] = sum(
                    history_counts.get(rid, 0) + today_counts.get(rid, 0) for rid in rid_set
                )
                live_map[emp.id] = sum(live_counts.get(rid, 0) for rid in rid_set)
        else:
            code_counts = await self._count_activations_by_retailer_code(start, end, excluded_retailer_ids)
            for emp in employees:
                achievement_map[emp.id] = code_counts.get(emp.assisted_retailer_code or "", 0)
                live_map[emp.id] = live_code_counts.get(emp.assisted_retailer_code or "", 0)

        rows: list[dict] = []
        for emp in employees:
            if entity_type == "rso":
                row = {
                    "rso_name": emp.user.name if emp.user else (emp.dms_code or f"#{emp.id}"),
                    "rso_dms_code": emp.dms_code or "",
                    "rso_itop_number": emp.itop_number or "",
                }
            else:
                row = {
                    "bp_name": emp.user.name if emp.user else (emp.dms_code or f"#{emp.id}"),
                    "bp_pool_number": emp.pool_number or "",
                    "bp_assisted_code": emp.assisted_retailer_code or "",
                }
            self._apply_slab_columns(row, achievement_map.get(emp.id, 0), targets, str(emp.id), self.cfg.slabs)
            row["live_activation"] = live_map.get(emp.id, 0)
            rows.append(row)

        sort_key = self.cfg.sort_by if self.cfg.sort_by in columns else "slab_1_achievement"
        rows.sort(key=lambda x: x.get(sort_key, 0) or 0, reverse=self.cfg.sort_order != "asc")

        totals = self._compute_totals(rows, columns, self.cfg.slabs)
        return {"columns": columns, "rows": rows, "totals": totals, "window": window}

    def _window_info(self, today: date, start: date, end: date) -> dict:
        return {
            "start": start.isoformat(),
            "end": end.isoformat(),
            "today": today.isoformat(),
            "today_source": "live" if start <= today <= end else None,
        }

    # -------------------------------------------------------------- renderers

    def _select_columns(self, columns: list[str], rows: list[dict], totals: dict) -> tuple[list[str], list[list[Any]], list[Any]]:
        header = ["#"] + [self._column_label(key) for key in columns]
        body = [
            [i + 1] + [r.get(key, "") if key not in ACTIVATION_METRICS else (r.get(key, 0) or 0) for key in columns]
            for i, r in enumerate(rows)
        ]
        total_row = ["Total"] + [totals.get(key, "") for key in columns]
        return header, body, total_row

    @staticmethod
    def _column_label(key: str) -> str:
        if is_slab_column(key):
            return slab_column_label(key)
        return COLUMN_LABELS.get(key, key)

    async def build_report_excel(self) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        report = await self.build_report()
        header, body, total_row = self._select_columns(report["columns"], report["rows"], report["totals"])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GA Report"

        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="C2410C", end_color="C2410C", fill_type="solid")
        thin = Border(*[Side(style="thin")] * 4)

        ws.cell(row=1, column=1, value=self.cfg.event_name or "GA Report").font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(header), 1))
        for ci, h in enumerate(header, 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin
        for ri, row in enumerate(body, 4):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = thin
        if total_row and any(str(v) for v in total_row):
            tr = 4 + len(body)
            for ci, val in enumerate(total_row, 1):
                cell = ws.cell(row=tr, column=ci, value=val)
                cell.font = Font(bold=True)
                cell.border = thin

        for ci in range(1, len(header) + 1):
            col = ws[get_column_letter(ci)]
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 45)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    async def build_report_text(self) -> list[str]:
        report = await self.build_report()
        header, body, total_row = self._select_columns(report["columns"], report["rows"], report["totals"])
        lines: list[str] = []
        title = self.cfg.event_name or "GA Report"
        lines.append(f"*{title} ({report['window']['start']} to {report['window']['end']})*")
        lines.append(f"Retailers: {len(body)}")
        lines.append("")
        lines.append("```")
        widths = []
        for ci, h in enumerate(header):
            w = max(len(str(h)), *[len(str(r[ci])) for r in body] or [1])
            widths.append(min(w + 2, 30))
        lines.extend(_table_lines(header, body, widths, total_row))
        lines.append("```")
        return _chunk_lines(lines)

    async def build_report_image(self) -> bytes:
        from app.services.ga_report_image import build_report_image

        report = await self.build_report()
        header, body, total_row = self._select_columns(report["columns"], report["rows"], report["totals"])
        title = self.cfg.event_name or "GA Report"
        return build_report_image(
            title=title,
            subtitle=f"{report['window']['start']} to {report['window']['end']} · Retailers: {len(body)}",
            header=header,
            rows=body,
            total_row=total_row,
        )


def _table_lines(header: list[str], rows: list[list], widths: list[int], total_row: list | None = None) -> list[str]:
    fmt_header = "  ".join(h.ljust(w)[:w] for h, w in zip(header, widths))
    sep = "  ".join("-" * w for w in widths)
    lines = [fmt_header, sep]
    for row in rows:
        cells = []
        for val, w in zip(row, widths):
            s = str(val)
            cells.append(s.ljust(w)[:w])
        lines.append("  ".join(cells))
    if total_row and any(str(v) for v in total_row):
        lines.append(sep)
        cells = []
        for val, w in zip(total_row, widths):
            s = str(val)
            cells.append(s.ljust(w)[:w])
        lines.append("  ".join(cells))
    return lines


def _chunk_lines(lines: list[str], limit: int = 3800) -> list[str]:
    chunks: list[str] = []
    current: list[str] = []
    current_len = 0
    for line in lines:
        if current and current_len + len(line) + 1 > limit:
            chunks.append("\n".join(current))
            current = []
            current_len = 0
        current.append(line)
        current_len += len(line) + 1
    if current:
        chunks.append("\n".join(current))
    return chunks