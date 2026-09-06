import io
from datetime import date, timedelta
from sqlalchemy import select, func, and_, false
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

from app.models.live_activation import LiveActivation
from app.models.activation import Activation
from app.models.retailer import Retailer
from app.models.employee import Employee
from app.models.user import User
from app.models.bp_retailer_code import BpRetailerCode
from app.models.bp_target import BpTarget
from app.models.ga_section_config import GaSectionConfig
from app.models.role import Role
from app.services.retailer_marking_service import get_active_retailer_ids_for_marking
from app.utils.activation_rules import exclude_clause

# ── Style constants (matching frontend activations export) ──
HEADER_BG = "1E293B"
SUBHEADER_BG = "F1F5F9"
ROW_ALT = "F8FAFC"
BORDER_COLOR = "E2E8F0"
TEXT_DARK = "1E293B"
TEXT_MUTED = "64748B"

THIN_BORDER = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR),
)
HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color=SUBHEADER_BG, end_color=SUBHEADER_BG, fill_type="solid")
ROW_ALT_FILL = PatternFill(start_color=ROW_ALT, end_color=ROW_ALT, fill_type="solid")

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
SUBHEADER_FONT = Font(name="Calibri", bold=True, color=TEXT_DARK, size=10)
TITLE_FONT = Font(name="Calibri", bold=True, color=TEXT_DARK, size=14)
SUBTITLE_FONT = Font(name="Calibri", color=TEXT_MUTED, size=10)
BODY_FONT = Font(name="Calibri", color=TEXT_DARK, size=10)
BOLD_FONT = Font(name="Calibri", bold=True, color=TEXT_DARK, size=10)


async def _load_export_section_configs(db: AsyncSession, house_id: int) -> dict[str, dict]:
    result = await db.execute(
        select(GaSectionConfig).where(GaSectionConfig.house_id == house_id)
    )
    configs: dict[str, dict] = {}
    for cfg in result.scalars().all():
        configs[cfg.section_key] = {
            "exclude_product_codes": cfg.exclude_product_codes or [],
            "exclude_retailer_tags": cfg.exclude_retailer_tags or [],
            "selected_employee_ids": cfg.selected_employee_ids or [],
        }
    return configs


async def _get_excluded_retailer_ids_by_tags(
    db: AsyncSession, house_id: int, tag_names: list[str]
) -> set[int]:
    excluded: set[int] = set()
    for tag_name in tag_names:
        excluded |= await get_active_retailer_ids_for_marking(db, house_id, tag_name)
    return excluded


def _apply_exclusions_to_query(query, model, exclude_product_codes: list[str], excluded_retailer_ids: set[int]):
    if exclude_product_codes:
        clause = exclude_clause(model, set(exclude_product_codes))
        if clause is not None:
            query = query.where(clause)
    if excluded_retailer_ids:
        query = query.where(
            and_(
                model.retailer_id != None,
                model.retailer_id.notin_(excluded_retailer_ids),
            )
        )
    return query


def _build_sheet(ws, title_text, subtitle_text, headers, rows, numeric_cols):
    nc = len(headers)

    # Row 1: Title (merged across all columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
    c = ws.cell(row=1, column=1, value=title_text)
    c.font = TITLE_FONT
    c.alignment = Alignment(vertical="center", horizontal="left")
    for ci in range(1, nc + 1):
        ws.cell(row=1, column=ci).border = THIN_BORDER

    # Row 2: Subtitle (merged across all columns)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nc)
    c = ws.cell(row=2, column=1, value=subtitle_text)
    c.font = SUBTITLE_FONT
    c.alignment = Alignment(vertical="center", horizontal="left")
    for ci in range(1, nc + 1):
        ws.cell(row=2, column=ci).border = THIN_BORDER

    # Row 4: Column headers
    hr = 4
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=ci, value=h)
        c.font = SUBHEADER_FONT
        c.fill = SUBHEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

    # Data rows (starting at row 5)
    dr = hr + 1
    totals = [0] * nc
    for ri, row_data in enumerate(rows):
        alt = ri % 2 == 1
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=dr, column=ci, value=val)
            c.font = BODY_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(
                horizontal="left" if ci == 1 else "center",
                vertical="center"
            )
            if alt:
                c.fill = ROW_ALT_FILL
            if isinstance(val, (int, float)):
                totals[ci - 1] += val
        dr += 1

    # Total row
    for ci in range(1, nc + 1):
        c = ws.cell(row=dr, column=ci)
        c.font = BOLD_FONT
        c.border = THIN_BORDER
        c.alignment = Alignment(
            horizontal="left" if ci == 1 else "center",
            vertical="center"
        )
        if ci == 1:
            c.value = "Total"
        elif ci in numeric_cols:
            c.value = totals[ci - 1]

    # Column widths
    for ci in range(1, nc + 1):
        col_letter = chr(64 + ci)
        ws.column_dimensions[col_letter].width = 20


async def export_ga_live_performance_excel(
    db: AsyncSession,
    house_id: int,
    today: date,
) -> bytes:
    yesterday = today - timedelta(days=1)

    section_configs = await _load_export_section_configs(db, house_id)

    exclude_products_total: list[str] = []
    exclude_tags_total: list[str] = []

    for section_key in ["total_activation", "distribution", "supervisors", "rsos", "bps"]:
        cfg = section_configs.get(section_key)
        if cfg:
            exclude_products_total.extend(cfg["exclude_product_codes"])
            exclude_tags_total.extend(cfg["exclude_retailer_tags"])

    excluded_retailer_ids = await _get_excluded_retailer_ids_by_tags(db, house_id, exclude_tags_total)

    emp_rows = await db.execute(
        select(Employee.id, Employee.user_id, Employee.dms_code, Employee.itop_number,
               Employee.personal_number, Employee.assisted_retailer_code, Employee.pool_number,
               Employee.employee_type)
        .where(Employee.house_id == house_id, Employee.status == "Active")
    )
    all_employees = emp_rows.all()

    user_ids = [e.user_id for e in all_employees if e.user_id]
    user_name_map: dict[int, str] = {}
    if user_ids:
        users_res = await db.execute(select(User.id, User.name).where(User.id.in_(user_ids)))
        for uid, uname in users_res.all():
            user_name_map[uid] = uname

    ret_rows = await db.execute(
        select(Retailer.id, Retailer.retailer_code, Retailer.name, Retailer.employee_id)
        .where(Retailer.house_id == house_id)
    )
    retailers = ret_rows.all()
    emp_retailer_map: dict[int, set[int]] = {}
    for r in retailers:
        if r.employee_id:
            emp_retailer_map.setdefault(r.employee_id, set()).add(r.id)

    bp_code_rows = await db.execute(
        select(BpRetailerCode.bp_employee_id, BpRetailerCode.retailer_code)
        .where(BpRetailerCode.house_id == house_id)
    )
    bp_code_map: dict[int, list[str]] = {}
    all_bp_codes: set[str] = set()
    for bp_emp_id, r_code in bp_code_rows.all():
        bp_code_map.setdefault(bp_emp_id, []).append(r_code)
        all_bp_codes.add(r_code)

    rso_list = [e for e in all_employees if e.employee_type == "rso"]
    bp_list = [e for e in all_employees if e.employee_type == "bp"]
    sup_list = [e for e in all_employees if e.employee_type == "supervisor"]

    # ── Supervisor → RSO user mapping (filter by RSO role) ──
    sup_user_ids = [e.user_id for e in sup_list if e.user_id]
    sup_user_to_rso_user_ids: dict[int, list[int]] = {}
    if sup_user_ids:
        user_rows = await db.execute(
            select(User).options(selectinload(User.roles)).where(User.parent_id.in_(sup_user_ids))
        )
        for u in user_rows.unique().scalars().all():
            if "rso" in [r.name.lower() for r in u.roles]:
                sup_user_to_rso_user_ids.setdefault(u.parent_id, []).append(u.id)
    rso_user_id_to_emp_id = {e.user_id: e.id for e in rso_list if e.user_id}

    async def _today_count(retailer_ids: set[int], bp_filter: bool = True):
        if not retailer_ids:
            return 0
        q = select(func.count()).where(
            LiveActivation.retailer_id.in_(retailer_ids),
            LiveActivation.house_id == house_id,
        )
        q = _apply_exclusions_to_query(q, LiveActivation, exclude_products_total, excluded_retailer_ids)
        if bp_filter and all_bp_codes:
            q = q.where(LiveActivation.retailer_code.notin_(all_bp_codes))
        res = await db.execute(q)
        return res.scalar() or 0

    async def _today_own_count(retailer_ids: set[int], code: str | None, bp_filter: bool = True):
        if not retailer_ids or not code:
            return 0
        q = select(func.count()).where(
            and_(
                LiveActivation.retailer_code == code,
                LiveActivation.retailer_id.in_(retailer_ids),
            ),
            LiveActivation.house_id == house_id,
        )
        q = _apply_exclusions_to_query(q, LiveActivation, exclude_products_total, excluded_retailer_ids)
        if bp_filter and all_bp_codes:
            q = q.where(LiveActivation.retailer_code.notin_(all_bp_codes))
        res = await db.execute(q)
        return res.scalar() or 0

    async def _yesterday_count(retailer_ids: set[int], bp_filter: bool = True):
        if not retailer_ids:
            return 0
        q = select(func.count()).where(
            Activation.retailer_id.in_(retailer_ids),
            Activation.house_id == house_id,
            Activation.activation_date == yesterday,
        )
        q = _apply_exclusions_to_query(q, Activation, exclude_products_total, excluded_retailer_ids)
        if bp_filter and all_bp_codes:
            q = q.where(Activation.retailer_code.notin_(all_bp_codes))
        res = await db.execute(q)
        return res.scalar() or 0

    async def _yesterday_own_count(retailer_ids: set[int], code: str | None, bp_filter: bool = True):
        if not retailer_ids or not code:
            return 0
        q = select(func.count()).where(
            and_(
                Activation.retailer_code == code,
                Activation.retailer_id.in_(retailer_ids),
            ),
            Activation.house_id == house_id,
            Activation.activation_date == yesterday,
        )
        q = _apply_exclusions_to_query(q, Activation, exclude_products_total, excluded_retailer_ids)
        if bp_filter and all_bp_codes:
            q = q.where(Activation.retailer_code.notin_(all_bp_codes))
        res = await db.execute(q)
        return res.scalar() or 0

    async def _bp_today(codes: list[str]):
        if not codes:
            return 0
        q = select(func.count()).where(
            LiveActivation.retailer_code.in_(codes),
            LiveActivation.house_id == house_id,
        )
        q = _apply_exclusions_to_query(q, LiveActivation, exclude_products_total, excluded_retailer_ids)
        res = await db.execute(q)
        return res.scalar() or 0

    async def _bp_yesterday(codes: list[str]):
        if not codes:
            return 0
        q = select(func.count()).where(
            Activation.retailer_code.in_(codes),
            Activation.house_id == house_id,
            Activation.activation_date == yesterday,
        )
        q = _apply_exclusions_to_query(q, Activation, exclude_products_total, excluded_retailer_ids)
        res = await db.execute(q)
        return res.scalar() or 0

    # ── Load BP targets for current month ──
    month_start = date(today.year, today.month, 1)
    _, last_day = __import__("calendar").monthrange(today.year, today.month)
    month_end = date(today.year, today.month, last_day)
    bp_emp_ids = [e.id for e in bp_list]
    bp_target_map: dict[int, int] = {}
    if bp_emp_ids:
        bp_target_rows = await db.execute(
            select(BpTarget).where(
                BpTarget.employee_id.in_(bp_emp_ids),
                BpTarget.target_date >= month_start,
                BpTarget.target_date <= month_end,
            )
        )
        for t in bp_target_rows.scalars().all():
            bp_target_map[t.employee_id] = t.ga_target or 0

    # ── BP MTD activation counts for remaining ──
    yesterday_for_mtd = today - timedelta(days=1)
    bp_mtd_code_counts: dict[str, int] = {}
    all_bp_codes_mtd = list(all_bp_codes)
    if all_bp_codes_mtd and yesterday_for_mtd >= month_start:
        mtd_bp_q = select(Activation.retailer_code, func.count()).where(
            Activation.house_id == house_id,
            Activation.activation_date >= month_start,
            Activation.activation_date <= yesterday_for_mtd,
            Activation.retailer_code.in_(all_bp_codes_mtd),
        )
        mtd_bp_q = _apply_exclusions_to_query(mtd_bp_q, Activation, exclude_products_total, excluded_retailer_ids)
        mtd_bp_q = mtd_bp_q.group_by(Activation.retailer_code)
        mtd_bp_rows = await db.execute(mtd_bp_q)
        for code, cnt in mtd_bp_rows.all():
            bp_mtd_code_counts[code] = cnt

    # ── Build Excel ──
    wb = Workbook()
    date_str = str(today)

    # ── Sheet 1: RSO Report ──
    ws_rso = wb.active
    ws_rso.title = "RSO Report"
    rso_headers = [
        "Name", "ITop Number", "Assisted Code",
        "Today Own", "Today Market", "Today Total",
        "Yesterday Own", "Yesterday Market", "Yesterday Total",
        "%",
    ]
    rso_numeric_cols = {4, 5, 6, 7, 8, 9, 10}
    rso_rows = []
    rso_grand_total = 0
    rso_total_map: dict[int, dict[str, int]] = {}
    for e in rso_list:
        emp_id = e.id
        ret_ids = emp_retailer_map.get(emp_id, set())
        code = e.assisted_retailer_code
        t_own = await _today_own_count(ret_ids, code)
        t_total = await _today_count(ret_ids)
        y_own = await _yesterday_own_count(ret_ids, code)
        y_total = await _yesterday_count(ret_ids)
        user_name = user_name_map.get(e.user_id) if e.user_id else ""
        rso_total_map[emp_id] = {"today_total": t_total, "yesterday_total": y_total}
        rso_rows.append([
            user_name or e.dms_code or f"#{emp_id}",
            e.itop_number or "",
            code or "",
            t_own,
            t_total - t_own,
            t_total,
            y_own,
            y_total - y_own,
            y_total,
            0.0,
        ])
        rso_grand_total += t_total
    for row in rso_rows:
        row[9] = round((row[5] / rso_grand_total * 100), 1) if rso_grand_total else 0
    _build_sheet(ws_rso, "RSO Performance Report", f"Date: {date_str}", rso_headers, rso_rows, rso_numeric_cols)
    # Conditional formatting color scale on % column (J5:J...)
    if rso_rows:
        last = 4 + len(rso_rows)
        ws_rso.conditional_formatting.add(f"J5:J{last}", ColorScaleRule(
            start_type="min", start_color="F44336",
            mid_type="percent", mid_value=50, mid_color="FFC107",
            end_type="max", end_color="4CAF50",
        ))

    # ── Sheet 2: BP Report ──
    ws_bp = wb.create_sheet("BP Report")
    days_in_month = last_day
    days_elapsed = today.day - 1
    days_remaining = days_in_month - days_elapsed
    bp_headers = ["Name", "Pool Number", "Assisted Code", "Today Target", "Ach", "%", "Remain", "Yest GA"]
    bp_numeric_cols = {4, 5, 6, 7, 8}
    bp_rows = []
    for e in bp_list:
        codes = bp_code_map.get(e.id, [])
        if e.assisted_retailer_code and e.assisted_retailer_code not in codes:
            codes.append(e.assisted_retailer_code)
        today_count = await _bp_today(codes)
        yest_count = await _bp_yesterday(codes)
        user_name = user_name_map.get(e.user_id) if e.user_id else ""
        bp_target_val = bp_target_map.get(e.id, 0)
        bp_mtd = sum(bp_mtd_code_counts.get(code, 0) for code in codes)
        bp_remaining = max(0, bp_target_val - bp_mtd)
        bp_today_target = int(bp_remaining / max(days_remaining, 1)) if bp_remaining > 0 else 0
        bp_rows.append([
            user_name or e.dms_code or f"#{e.id}",
            e.pool_number or "",
            e.assisted_retailer_code or "",
            bp_today_target,
            today_count,
            0.0,  # % formula
            0,    # remain formula
            yest_count,
        ])
        # Column indices: 4=Today Target, 5=Ach, 6=%, 7=Remain
    for row in bp_rows:
        row[5] = round((row[4] / row[3] * 100), 1) if row[3] > 0 else 0  # % = Ach / Today Target
        row[6] = max(0, row[3] - row[4])  # Remain = Today Target - Ach
    _build_sheet(ws_bp, "BP Performance Report", f"Date: {date_str}", bp_headers, bp_rows, bp_numeric_cols)
    # Conditional formatting color scale on % column (G5:G...)
    if bp_rows:
        last = 4 + len(bp_rows)
        ws_bp.conditional_formatting.add(f"G5:G{last}", ColorScaleRule(
            start_type="min", start_color="F44336",
            mid_type="percent", mid_value=50, mid_color="FFC107",
            end_type="max", end_color="4CAF50",
        ))

    # ── Sheet 3: Supervisor Report (aggregated from RSO data) ──
    ws_sup = wb.create_sheet("Supervisor Report")
    sup_headers = ["Name", "Pool Number", "Today Activation", "Yesterday Activation"]
    sup_numeric_cols = {3, 4}
    sup_rows = []
    for sup_emp in sup_list:
        user_name = user_name_map.get(sup_emp.user_id) if sup_emp.user_id else ""
        today_total = 0
        yesterday_total = 0
        sup_uid = sup_emp.user_id
        if sup_uid:
            for rso_uid in sup_user_to_rso_user_ids.get(sup_uid, []):
                rso_emp_id = rso_user_id_to_emp_id.get(rso_uid)
                if rso_emp_id and rso_emp_id in rso_total_map:
                    today_total += rso_total_map[rso_emp_id]["today_total"]
                    yesterday_total += rso_total_map[rso_emp_id]["yesterday_total"]
        sup_rows.append([
            user_name or sup_emp.dms_code or f"#{sup_emp.id}",
            sup_emp.pool_number or "",
            today_total,
            yesterday_total,
        ])
    _build_sheet(ws_sup, "Supervisor Performance Report", f"Date: {date_str}", sup_headers, sup_rows, sup_numeric_cols)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
