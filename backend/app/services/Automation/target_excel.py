import pandas as pd
import logging
from datetime import datetime
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy import func, select
from app.models.house_target import HouseTarget
from app.models.supervisor_target import SupervisorTarget
from app.models.rso_target import RSOTarget
from app.services.db_service import async_session
from app.utils.helpers import bn_num

logger = logging.getLogger(__name__)

def normalize_code(code):
    if not code: return None
    # Remove leading zeros and spaces for flexible matching
    c = str(code).strip().upper()
    while c.startswith('0') and len(c) > 1:
        c = c[1:]
    return c

def normalize_msisdn(msisdn):
    if not msisdn: return None
    s = str(msisdn).strip()
    if s.endswith('.0'): s = s[:-2]
    if s.startswith('880'): s = s[3:]
    if s.startswith('0'): s = s[1:]
    return s

def clean_val(val):
    if pd.isna(val):
        return None
    v = str(val).strip()
    if v.startswith("'"): v = v[1:]
    if v.endswith("'"): v = v[:-1]
    if v == "" or v.lower() in ["nan", "none", "null"]:
        return None
    return v

def clean_float(val):
    if pd.isna(val):
        return 0.0
    try:
        v = str(val).replace(",", "").strip()
        if v.startswith("'"): v = v[1:]
        return float(v)
    except:
        return 0.0

def clean_int(val):
    if pd.isna(val):
        return 0
    try:
        v = str(val).replace(",", "").strip()
        if v.startswith("'"): v = v[1:]
        return int(float(v))
    except:
        return 0

def clean_date(val):
    if pd.isna(val):
        return None
    try:
        if isinstance(val, datetime):
            return val
        if hasattr(val, 'to_pydatetime'):
            return val.to_pydatetime()
        v = str(val).strip()
        for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%b-%Y", "%Y%m%d",
                     "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f"]:
            try:
                return datetime.strptime(v, fmt)
            except:
                pass
        try:
            from datetime import timedelta
            return datetime(1899, 12, 30) + timedelta(days=float(v))
        except:
            pass
        return None
    except:
        return None

HOUSE_COLUMN_MAP = {
    'DD_CODE': 'house_code',
    'D_CODE': 'house_code',
    'HOUSE_CODE': 'house_code',
    'EV_C2C_TARGET': ('ev_c2c_target', clean_float),
    'SC_PRIMARY_TARGET': ('sc_primary_target', clean_float),
    'TOTAL_RECHARGE_TARGET': ('total_recharge_target', clean_float),
    'TOTAL_RECHAGE_TARGET_(EV_C2C+_SC_PRIMARY)': ('total_recharge_target', clean_float),
    'TOTAL_RECHARGE_TARGET_(EV_C2C+_SC_PRIMARY)': ('total_recharge_target', clean_float),
    'TOTAL_GA_TARGET': ('total_ga_target', clean_int),
    'BP_GA': ('bp_ga', clean_int),
    'RSO_GA': ('rso_ga', clean_int),
    'EV_SCR': ('ev_scr', clean_float),
    'SSO': ('sso', clean_int),
    'LSO': ('lso', clean_int),
    'ALSO': ('lso', clean_int),
    'BSO': ('bso', clean_int),
    'DDSO': ('ddso', clean_int)
}

# Metadata columns to exclude from extra_targets
HOUSE_META_COLUMNS = [
    'CLUSTER', 'REGION', 'D_CODE', 'D_NAME', 'DD_CODE', 'DD_NAME', 
    'MONTH', 'YEAR', 'HOUSE_CODE', 'HOUSE_NAME', 'TARGET_DATE'
]

SUPERVISOR_COLUMN_MAP = {
    'DD_CODE': 'house_code',
    'HOUSE_CODE': 'house_code',
    'SUPERVISOR_MSISDN': 'supervisor_msisdn',
    'RS0_SUPERVISOR_MSISDN': 'supervisor_msisdn',
    'RSO_SUPERVISOR_MSISDN': 'supervisor_msisdn',
    'EV_SECONDARY': ('ev_secondary', clean_float),
    'SC_SECONDARY': ('sc_secondary', clean_float),
    'TOTAL_RECHARGE': ('total_recharge', clean_float),
    'TOTAL_RECHAGE': ('total_recharge', clean_float),
    'TOTAL_RECHAGE_(EV_SECONDARY+SC_SECONDARY)': ('total_recharge', clean_float),
    'TOTAL_GA': ('total_ga', clean_int),
    'BP_GA': ('bp_ga', clean_int),
    'RSO_GA': ('rso_ga', clean_int),
    'GA_(RSO)': ('rso_ga', clean_int),
    'SSO': ('sso', clean_int),
    'ASSO': ('sso', clean_int),
    'LSO': ('lso', clean_int),
    'ALSO': ('lso', clean_int),
    'BSO': ('bso', clean_int),
    'DDSO': ('ddso', clean_int)
}

SUPERVISOR_META_COLUMNS = [
    'CLUSTER', 'REGION', 'DD_CODE', 'DD_NAME', 'RS0_SUPERVISOR_NAME', 
    'RS0_SUPERVISOR_MSISDN', 'RSO_SUPERVISOR_NAME', 'RSO_SUPERVISOR_MSISDN',
    'SUPERVISOR_MSISDN', 'HOUSE_CODE', 'HOUSE_NAME', 'SUPERVISOR_NAME', 'TARGET_DATE'
]

RSO_COLUMN_MAP = {
    'DD_CODE': 'house_code',
    'HOUSE_CODE': 'house_code',
    'RS0_CODE': 'rso_code',
    'RSO_CODE': 'rso_code',
    'RSO_ITOPUP_NUMBER': 'rso_itop',
    'SUPERVISOR_MSISDN': 'supervisor_msisdn',
    'RS0_SUPERVISOR_MSISDN': 'supervisor_msisdn',
    'RSO_SUPERVISOR_MSISDN': 'supervisor_msisdn',
    'EV_SECONDARY': ('ev_secondary', clean_float),
    'SC_SECONDARY': ('sc_secondary', clean_float),
    'TOTAL_RECHARGE': ('total_recharge', clean_float),
    'TOTAL_RECHAGE': ('total_recharge', clean_float),
    'TOTAL_RECHAGE_(EV_SECONDARY+SC_SECONDARY)': ('total_recharge', clean_float),
    'GA': ('ga', clean_int),
    'GA_(RSO)': ('ga', clean_int),
    'SSO': ('sso', clean_int),
    'ASSO': ('sso', clean_int),
    'LSO': ('lso', clean_int),
    'ALSO': ('lso', clean_int),
    'BSO': ('bso', clean_int),
    'DDSO': ('ddso', clean_int),
    'SERVICE_ROUTE': 'service_route',
    'MAIN_HOUSE/OSDO/RESIDENTIAL_RSO': 'market_type',
    'MARKET_TYPE': 'market_type',
    'THANA_NAME_(ONLY_FOR_OSDO)': 'thana_name',
    'THANA_NAME': 'thana_name',
    'GA_TARGET_(MODIFIED)': ('ga_target_modified', clean_int),
    'EV_SECONDARY_(MODIFIED)': ('ev_secondary_modified', clean_float),
    'SC_SECONDARY_(MODIFIED)': ('sc_secondary_modified', clean_float),
    'RECHARGE_TARGET_(MODIFIED)': ('recharge_target_modified', clean_float),
    'LSO_TARGET_(MODIFIED)': ('lso_target_modified', clean_int),
    'SSO_TARGET_(MODIFIED)': ('sso_target_modified', clean_int),
    'BSO_TARGET_(MODIFIED)': ('bso_target_modified', clean_int),
    'DAILY_DSO_TARGET_(MODIFIED)': ('daily_dso_target_modified', clean_int)
}

RSO_META_COLUMNS = [
    'CLUSTER', 'REGION', 'DD_CODE', 'DD_NAME', 'RS0_CODE', 'RSO_CODE',
    "RS0_MSISDN_[I'TOP-UP_NUMBER]", 'RS0_MSISDN', 'RS0_NAME', 'RSO_NAME',
    'RS0_SUPERVISOR_NAME', 'RS0_SUPERVISOR_MSISDN', 'RSO_SUPERVISOR_NAME', 'RSO_SUPERVISOR_MSISDN',
    'SUPERVISOR_MSISDN', 'HOUSE_CODE', 'HOUSE_NAME', 'DD_MANAGER_NAME', 'DD_MANAGER_CONTACT_NUMBER',
    'NEW_MARKET_TYPE', 'ARCHETYPE', 'TYPE_OF_THANA', 'TARGET_DATE'
]

from app.models.house import House
from app.models.employee import Employee

async def process_target_excel_unified(file_path, target_date, progress_callback=None):
    """Auto-process each sheet in Excel file ✅"""
    try:
        xl = pd.ExcelFile(file_path)
        sheet_names = xl.sheet_names
        
        results = []
        total_processed = 0

        async with async_session() as session:
            # Pre-fetch ONLY ACTIVE houses
            houses_res = await session.execute(select(House).where(House.is_active == True))
            all_houses = houses_res.scalars().all()
            house_map = {normalize_code(h.code): h.id for h in all_houses}
            
            # Fetch ALL employees for mapping
            emp_res = await session.execute(select(Employee))
            all_emp = emp_res.scalars().all()
            
            # Active Employees for TARGET UPLOAD
            active_emp = [f for f in all_emp if f.status == 'Active']
            
            # RSO map for TARGET UPLOAD (Strictly Type='RSO')
            rso_map_active = {}
            rso_map_by_itop = {}
            for f in active_emp:
                f_type_upper = (f.employee_type or "").upper()
                if f.dms_code and f_type_upper == 'RSO':
                    rso_map_active[normalize_code(f.dms_code)] = f.id
                if f.itop_number and f_type_upper == 'RSO':
                    rso_map_by_itop[normalize_msisdn(f.itop_number)] = f.id
            
            # Supervisor map for RSO linkage (Only Active, Only pool_number)
            sup_map_all = {}
            for f in active_emp:
                if f.pool_number:
                    sup_map_all[normalize_msisdn(f.pool_number)] = f.id

            # Supervisor map for TARGET UPLOAD (Strictly ACTIVE, POOL_NUMBER, and Type='Supervisor')
            sup_map_active_pool = {}
            for f in active_emp:
                f_type_upper = (f.employee_type or "").upper()
                if f.pool_number and f_type_upper == 'SUPERVISOR':
                    sup_map_active_pool[normalize_msisdn(f.pool_number)] = f.id
            
            debug_info = (
                f"🔍 DB Status (Active Only):\n"
                f"🏠 Houses: {len(house_map)}\n"
                f"👤 RSOs: {len(rso_map_active)}\n"
                f"👨‍💼 Supervisors: {len(sup_map_active_pool)}\n\n"
            )

            for sheet_name in sheet_names:
                name_upper = sheet_name.upper()
                df = pd.read_excel(xl, sheet_name=sheet_name)
                if df.empty: continue

                # Clean columns for detection
                orig_cols = df.columns.tolist()
                clean_cols = [str(c).strip().upper().replace(" ", "_").replace("\n", "_") for c in orig_cols]
                df.columns = clean_cols
                col_name_map = dict(zip(clean_cols, orig_cols))

                # Target type detection (Refined)
                target_type = None
                if sheet_name == "Supervisor Target":
                    target_type = 'supervisor'
                    model = SupervisorTarget
                    col_map = SUPERVISOR_COLUMN_MAP
                    meta_cols = SUPERVISOR_META_COLUMNS
                    conflict_elements = ['employee_id', 'target_date']
                elif 'HOUSE' in name_upper or 'DD' in name_upper or 'TOTAL_GA_TARGET' in clean_cols:
                    target_type = 'house'
                    model = HouseTarget
                    col_map = HOUSE_COLUMN_MAP
                    meta_cols = HOUSE_META_COLUMNS
                    conflict_elements = ['house_id', 'target_date']
                elif 'RSO' in name_upper or 'RSO_CODE' in clean_cols or 'RS0_CODE' in clean_cols:
                    target_type = 'rso'
                    model = RSOTarget
                    col_map = RSO_COLUMN_MAP
                    meta_cols = RSO_META_COLUMNS
                    conflict_elements = ['employee_id', 'target_date']
                
                if not target_type:
                    logger.warning(f"Skipping unknown sheet: {sheet_name}")
                    continue

                if progress_callback:
                    await progress_callback(f"⏳ Processing: <b>{sheet_name}</b> ({target_type})...")

                count = 0
                batch_data = []

                for index, row in df.iterrows():
                    row_target_date = target_date
                    if 'TARGET_DATE' in clean_cols:
                        parsed_date = clean_date(row.get('TARGET_DATE'))
                        if parsed_date:
                            row_target_date = parsed_date
                    values = {"target_date": row_target_date}
                    excel_values = {}
                    extra_targets = {}
                    
                    # 1. Collect values per mapping
                    mapped_excel_cols = set()
                    for excel_header, db_field in col_map.items():
                        if excel_header in clean_cols:
                            mapped_excel_cols.add(excel_header)
                            if isinstance(db_field, tuple):
                                field_name, cleaner = db_field
                                excel_values[field_name] = cleaner(row.get(excel_header))
                            else:
                                excel_values[db_field] = clean_val(row.get(excel_header))

                    # 2. Extra column processing
                    for col in clean_cols:
                        if col not in mapped_excel_cols and col not in meta_cols:
                            val = row.get(col)
                            if pd.notna(val):
                                extra_targets[col_name_map[col]] = val
                    values['extra_targets'] = extra_targets

                    # 3. ID lookup and validation
                    h_code_raw = excel_values.get('house_code')
                    h_id = house_map.get(normalize_code(h_code_raw))
                    if not h_id: continue # Skip if House not found
                    values['house_id'] = h_id

                    if target_type == 'house':
                        for k, v in excel_values.items():
                            if k != 'house_code': values[k] = v
                    
                    elif target_type == 'supervisor':
                        s_msisdn = normalize_msisdn(excel_values.get('supervisor_msisdn'))
                        emp_id = sup_map_active_pool.get(s_msisdn)
                        if not emp_id: continue # Skip if Active and Pool Number don't match
                        values['employee_id'] = emp_id
                        for k, v in excel_values.items():
                            if k not in ['house_code', 'supervisor_msisdn']:
                                values[k] = v

                    elif target_type == 'rso':
                        r_code = normalize_code(excel_values.get('rso_code'))
                        emp_id = rso_map_active.get(r_code)
                        if not emp_id:
                            r_itop = normalize_msisdn(excel_values.get('rso_itop'))
                            emp_id = rso_map_by_itop.get(r_itop)
                        if not emp_id: continue 
                        values['employee_id'] = emp_id
                        
                        # Ensure supervisor_id exists in dict even if None
                        values['supervisor_id'] = None
                        s_msisdn = normalize_msisdn(excel_values.get('supervisor_msisdn'))
                        if s_msisdn:
                            sid = sup_map_all.get(s_msisdn)
                            if sid: values['supervisor_id'] = sid

                        for k, v in excel_values.items():
                            if k not in ['house_code', 'rso_code', 'rso_itop', 'supervisor_msisdn']:
                                values[k] = v

                    batch_data.append(values)

                    if len(batch_data) >= 100:
                        await do_bulk_upsert_target(session, model, batch_data, conflict_elements)
                        count += len(batch_data)
                        batch_data = []

                if batch_data:
                    await do_bulk_upsert_target(session, model, batch_data, conflict_elements)
                    count += len(batch_data)
                
                results.append(f"✅ {sheet_name}: {bn_num(count)} records")
                total_processed += count

            await session.commit()
            summary = debug_info + "\n".join(results)
            if progress_callback:
                await progress_callback(summary)
            return total_processed, None

    except Exception as e:
        logger.error(f"Error in unified target excel: {str(e)}", exc_info=True)
        return 0, f"Error: {str(e)}"

async def do_bulk_upsert_target(session, model, batch_data, conflict_elements):
    """Bulk upsert logic for targets"""
    stmt = insert(model).values(batch_data)
    excluded = stmt.excluded
    update_dict = {
        k: getattr(excluded, k) 
        for k in batch_data[0].keys() 
        if k not in conflict_elements
    }
    update_dict['updated_at'] = func.now()
    
    stmt = stmt.on_conflict_do_update(
        index_elements=conflict_elements,
        set_=update_dict
    )
    await session.execute(stmt)

async def update_progress_target(count, total_rows, target_type, progress_callback):
    """Target progress update helper"""
    percent = round((count / total_rows) * 100)
    await progress_callback(
        f"📊 <b>Target Upload ({target_type}):</b> {bn_num(percent)}%\n"
        f"📈 Processed: <code>{bn_num(count)}</code> / <code>{bn_num(total_rows)}</code>"
    )


import io
from openpyxl import Workbook

def generate_house_target_sample_bytes():
    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "House Targets"
    headers = ["TARGET_DATE", "HOUSE_CODE", "EV_C2C_TARGET", "SC_PRIMARY_TARGET", "TOTAL_RECHARGE_TARGET",
               "TOTAL_GA_TARGET", "BP_GA", "RSO_GA", "EV_SCR", "SSO", "LSO", "BSO", "DDSO"]
    ws.append(headers)
    ws.append(["2026-07-01", "DD001", "500", "300", "800", "50", "20", "10", "100", "5", "3", "2", "1"])
    ws.append(["2026-07-01", "DD002", "600", "350", "950", "60", "25", "12", "120", "6", "4", "3", "2"])
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def generate_supervisor_target_sample_bytes():
    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Supervisor Target"
    headers = ["TARGET_DATE", "HOUSE_CODE", "SUPERVISOR_MSISDN", "EV_SECONDARY", "SC_SECONDARY", "TOTAL_RECHARGE",
               "TOTAL_GA", "BP_GA", "RSO_GA", "SSO", "LSO", "BSO", "DDSO"]
    ws.append(headers)
    ws.append(["2026-07-01", "DD001", "01712345678", "200", "150", "350", "30", "10", "5", "3", "2", "1", "1"])
    ws.append(["2026-07-01", "DD002", "01787654321", "250", "180", "430", "35", "12", "6", "4", "3", "2", "1"])
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def generate_rso_target_sample_bytes():
    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "RSO Target"
    headers = ["TARGET_DATE", "HOUSE_CODE", "RSO_CODE", "SUPERVISOR_MSISDN", "EV_SECONDARY", "SC_SECONDARY",
               "TOTAL_RECHARGE", "GA", "SSO", "LSO", "BSO", "DDSO",
               "SERVICE_ROUTE", "MARKET_TYPE", "THANA_NAME"]
    ws.append(headers)
    ws.append(["2026-07-01", "DD001", "RSO001", "01712345678", "100", "80", "180", "15", "2", "1", "1", "1", "Route A", "OSDO", "Thana A"])
    ws.append(["2026-07-01", "DD002", "RSO002", "01787654321", "120", "90", "210", "18", "3", "2", "1", "1", "Route B", "Residential", "Thana B"])
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

async def export_house_targets_excel(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "House Targets"
    headers = ["House Code", "EV C2C Target", "SC Primary Target", "Total Recharge Target",
               "Total GA Target", "BP GA", "RSO GA", "EV SCR", "SSO", "LSO", "BSO", "DDSO",
               "Target Date"]
    ws.append(headers)
    for r in records:
        ws.append([
            r.house.code if r.house else "", r.ev_c2c_target, r.sc_primary_target,
            r.total_recharge_target, r.total_ga_target, r.bp_ga, r.rso_ga, r.ev_scr,
            r.sso, r.lso, r.bso, r.ddso, str(r.target_date) if r.target_date else ""
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

async def export_supervisor_targets_excel(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Supervisor Targets"
    headers = ["Employee ID", "House Code", "EV Secondary", "SC Secondary", "Total Recharge",
               "Total GA", "BP GA", "RSO GA", "SSO", "LSO", "BSO", "DDSO", "Target Date"]
    ws.append(headers)
    for r in records:
        ws.append([
            r.employee_id, r.house.code if r.house else "", r.ev_secondary, r.sc_secondary,
            r.total_recharge, r.total_ga, r.bp_ga, r.rso_ga, r.sso, r.lso, r.bso,
            r.ddso, str(r.target_date) if r.target_date else ""
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

async def export_rso_targets_excel(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "RSO Targets"
    headers = ["Employee ID", "Supervisor ID", "House Code", "EV Secondary", "SC Secondary",
               "Total Recharge", "GA", "SSO", "LSO", "BSO", "DDSO", "Service Route",
               "Market Type", "Thana", "Target Date"]
    ws.append(headers)
    for r in records:
        ws.append([
            r.employee_id, r.supervisor_id, r.house.code if r.house else "",
            r.ev_secondary, r.sc_secondary, r.total_recharge, r.ga, r.sso, r.lso,
            r.bso, r.ddso, r.service_route, r.market_type, r.thana_name,
            str(r.target_date) if r.target_date else ""
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
